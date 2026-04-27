import os
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file, make_response
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import google.generativeai as genai
import json
import re
import PyPDF2
import nltk
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
from collections import Counter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import io
from datetime import datetime, timedelta
import requests
import secrets
import uuid
import hashlib
import time
import base64
import tempfile
import csv
import smtplib
import random
from sqlalchemy.exc import IntegrityError
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from urllib.parse import urlencode
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# OCR imports (optional - graceful fallback if not available)
try:
    from pdf2image import convert_from_path
    from PIL import Image
    import pytesseract
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False
    print("Warning: OCR libraries (pdf2image, Pillow, pytesseract) not installed. Will use cloud OCR API for scanned PDFs.")

# Optional: Set Tesseract path if not in system PATH (uncomment and adjust if needed)
# if OCR_AVAILABLE and os.name == 'nt':  # Windows
#     pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# Geolocation helper function - get location from IP address
def get_geolocation_from_ip(ip_address):
    """Get geolocation data (lat, lng, city, country) from IP address using free API"""
    try:
        # Skip for localhost/private IPs
        if not ip_address or ip_address in ['127.0.0.1', 'localhost', '::1']:
            return None
        
        # Check for private IP ranges
        if ip_address.startswith(('10.', '172.', '192.168.')):
            return None
        
        # Use ip-api.com (free, no API key required, 45 requests/minute)
        response = requests.get(
            f'http://ip-api.com/json/{ip_address}',
            timeout=3  # Quick timeout to not slow down login
        )
        
        if response.status_code == 200:
            data = response.json()
            if data.get('status') == 'success':
                return {
                    'latitude': data.get('lat'),
                    'longitude': data.get('lon'),
                    'city': data.get('city'),
                    'country': data.get('country'),
                    'region': data.get('regionName')
                }
        return None
    except Exception as e:
        print(f"Geolocation lookup failed for {ip_address}: {str(e)}")
        return None

def persist_proctoring_image(image_data, submission_id, snapshot_type, captured_at):
    """
    Persist proctoring image and return storable reference (URL/base64).
    If Cloudinary env vars are configured, uploads image and returns secure URL.
    Otherwise returns original image_data (legacy DB behavior).
    """
    if not image_data:
        return None

    def _save_local_snapshot():
        try:
            if image_data.startswith('data:image/'):
                header, b64 = image_data.split(',', 1)
                ext = 'jpg'
                if 'image/png' in header:
                    ext = 'png'
                elif 'image/webp' in header:
                    ext = 'webp'
                folder = os.path.join('static', 'proctoring_snapshots', str(submission_id))
                os.makedirs(folder, exist_ok=True)
                filename = f"{snapshot_type}_{int(captured_at.timestamp())}_{secrets.token_hex(3)}.{ext}"
                path = os.path.join(folder, filename)
                with open(path, 'wb') as f:
                    f.write(base64.b64decode(b64))
                return f"proctoring_snapshots/{submission_id}/{filename}"
        except Exception as e:
            print(f"Local snapshot fallback failed: {e}")
        return None

    cloud_name = os.environ.get('CLOUDINARY_CLOUD_NAME')
    api_key = os.environ.get('CLOUDINARY_API_KEY')
    api_secret = os.environ.get('CLOUDINARY_API_SECRET')

    if not (cloud_name and api_key and api_secret):
        local_ref = _save_local_snapshot()
        return local_ref if local_ref else image_data

    try:
        ts = int(time.time())
        base_folder = os.environ.get('CLOUDINARY_PROCTOR_FOLDER', 'unitest/proctoring')
        folder = f"{base_folder}/{submission_id}"
        public_id = f"{snapshot_type}_{int(captured_at.timestamp())}_{secrets.token_hex(3)}"
        to_sign = f"folder={folder}&public_id={public_id}&timestamp={ts}{api_secret}"
        signature = hashlib.sha1(to_sign.encode('utf-8')).hexdigest()
        url = f"https://api.cloudinary.com/v1_1/{cloud_name}/image/upload"

        response = requests.post(
            url,
            data={
                'api_key': api_key,
                'timestamp': ts,
                'folder': folder,
                'public_id': public_id,
                'signature': signature,
                'file': image_data,
            },
            timeout=20
        )
        result = response.json() if response.content else {}
        secure_url = result.get('secure_url')
        if secure_url:
            return secure_url
        print(f"Cloudinary upload failed: {result}")
    except Exception as e:
        print(f"Cloudinary upload error: {e}")

    local_ref = _save_local_snapshot()
    return local_ref if local_ref else image_data

# Cloud OCR API support (works in serverless environments like Vercel)
def extract_pdf_content_with_cloud_ocr(file_path):
    """Extract text from PDF using cloud OCR API (works in serverless environments)"""
    try:
        import base64
        import io
        import tempfile
        ocr_api_key = os.environ.get('OCR_SPACE_API_KEY', '')
        
        # Check if file exists and get size
        if not os.path.exists(file_path):
            print(f"Error: PDF file not found: {file_path}")
            return None
        
        file_size = os.path.getsize(file_path)
        print(f"Processing PDF with cloud OCR (file size: {file_size} bytes)")
        
        # Method 1: Try direct PDF upload to OCR.space (works without pdf2image/Poppler)
        direct_pdf_error_text = ""
        try:
            if file_size > 10 * 1024 * 1024:  # 10MB limit for free tier
                print(f"PDF file too large ({file_size} bytes). Will try page-by-page processing...")
            else:
                with open(file_path, 'rb') as pdf_file:
                    files = {'file': (os.path.basename(file_path), pdf_file, 'application/pdf')}
                    payload = {
                        'language': 'eng',
                        'isOverlayRequired': False,
                        'detectOrientation': True,
                        'scale': True,
                        'OCREngine': 2  # Use OCR Engine 2 for better accuracy
                    }
                    
                    # Check for OCR.space API key (optional, improves rate limits)
                    headers = {}
                    
                    if ocr_api_key:
                        # Try multiple methods: form data, header, and query parameter
                        payload['apikey'] = ocr_api_key
                        headers['apikey'] = ocr_api_key
                        print(f"Using OCR.space API key (length: {len(ocr_api_key)}, first 5 chars: {ocr_api_key[:5]}...)")
                    else:
                        print("WARNING: OCR_SPACE_API_KEY not found in environment variables")
                        print("Available env vars starting with OCR: " + str([k for k in os.environ.keys() if 'OCR' in k.upper()]))
                    
                    # OCR.space supports file upload at /parse/image (PDF/Image both accepted)
                    api_url = "https://api.ocr.space/parse/image"
                    if ocr_api_key:
                        api_url += f"?apikey={ocr_api_key}"
                    
                    print(f"Uploading PDF to OCR.space API (size: {file_size} bytes)...")
                    print(f"API URL: {api_url.split('?')[0]}")  # Don't log full URL with key
                    try:
                        # Upload PDF file directly using the supported endpoint
                        response = requests.post(api_url, files=files, data=payload, headers=headers, timeout=120)
                        print(f"OCR.space API response status: {response.status_code}")
                        print(f"OCR.space API response headers: {dict(response.headers)}")
                        
                        if response.status_code == 200:
                            try:
                                result = response.json()
                                print(f"OCR.space API response keys: {list(result.keys()) if isinstance(result, dict) else 'Not a dict'}")
                                
                                # Check for ParsedResults regardless of OCRExitCode (sometimes it works even with non-1 exit code)
                                if result.get('ParsedResults') and len(result['ParsedResults']) > 0:
                                    ocr_text = ""
                                    for parsed_result in result['ParsedResults']:
                                        page_text = parsed_result.get('ParsedText', '')
                                        if page_text and page_text.strip():
                                            ocr_text += page_text + "\n\n"
                                    
                                    if ocr_text.strip():
                                        exit_code = result.get('OCRExitCode', 'Unknown')
                                        print(f"Cloud OCR (PDF direct) extracted {len(ocr_text)} characters (Exit Code: {exit_code})")
                                        return ocr_text.strip()
                                    else:
                                        print(f"ParsedResults found but no text extracted. Exit Code: {result.get('OCRExitCode', 'Unknown')}")
                                else:
                                    # No ParsedResults - check for error
                                    exit_code = result.get('OCRExitCode', 'Unknown')
                                    error_msg = result.get('ErrorMessage', f"OCR Exit Code: {exit_code}")
                                    print(f"OCR.space API error: {error_msg}")
                                    direct_pdf_error_text = str(error_msg)
                                    print(f"Full response: {result}")
                                    # Check if it's a rate limit issue
                                    if 'rate limit' in str(error_msg).lower() or 'quota' in str(error_msg).lower():
                                        print("OCR.space rate limit reached. Consider getting an API key.")
                                    elif exit_code == 1:
                                        print("Exit code is 1 (success) but no ParsedResults - PDF might be blank or unreadable")
                            except ValueError as json_error:
                                print(f"Error parsing OCR.space JSON response: {json_error}")
                                print(f"Response text (first 1000 chars): {response.text[:1000]}")
                                print(f"Response status: {response.status_code}")
                        elif response.status_code == 401:
                            print("OCR.space API: Authentication failed (check API key)")
                            print(f"Response: {response.text[:500]}")
                        elif response.status_code == 429:
                            print("OCR.space API: Rate limit exceeded")
                            print(f"Response: {response.text[:500]}")
                        elif response.status_code == 400:
                            print("OCR.space API: Bad request - check file format")
                            print(f"Response: {response.text[:500]}")
                            direct_pdf_error_text = response.text[:500]
                        else:
                            print(f"OCR.space API returned status code: {response.status_code}")
                            try:
                                error_detail = response.text[:1000]
                                print(f"Error detail: {error_detail}")
                                direct_pdf_error_text = error_detail
                            except:
                                pass
                    except requests.exceptions.Timeout:
                        print("OCR.space API request timed out after 120 seconds")
                    except requests.exceptions.RequestException as req_error:
                        print(f"OCR.space API request error: {req_error}")
                        direct_pdf_error_text = str(req_error)
                        import traceback
                        traceback.print_exc()
        except Exception as pdf_api_error:
            print(f"Error with direct PDF OCR API: {pdf_api_error}")
            direct_pdf_error_text = str(pdf_api_error)
            import traceback
            traceback.print_exc()

        # Method 1B: If direct upload fails due file-size validation, split PDF per page and OCR each page.
        # This avoids poppler dependency and works in serverless when multi-page PDF exceeds OCR size limits.
        size_limit_hit = ('maximum permissible file size limit' in direct_pdf_error_text.lower()) or ('exceeds' in direct_pdf_error_text.lower() and 'kb' in direct_pdf_error_text.lower())
        if size_limit_hit or file_size > 1024 * 1024:
            try:
                print("Direct PDF OCR hit size constraints. Trying page-wise PDF splitting fallback...")
                split_ocr_text = ""
                page_count = 0
                with open(file_path, 'rb') as src_pdf:
                    pdf_reader = PyPDF2.PdfReader(src_pdf)
                    page_count = len(pdf_reader.pages)
                    print(f"Splitting PDF into {page_count} single-page files for OCR...")

                    for page_index in range(page_count):
                        try:
                            page_writer = PyPDF2.PdfWriter()
                            page_writer.add_page(pdf_reader.pages[page_index])

                            with tempfile.NamedTemporaryFile(delete=True, suffix='.pdf') as tmp_page_pdf:
                                page_writer.write(tmp_page_pdf)
                                tmp_page_pdf.flush()

                                page_size = os.path.getsize(tmp_page_pdf.name)
                                if page_size > 1024 * 1024:
                                    print(f"Page {page_index + 1} still exceeds 1MB ({page_size} bytes); skipping this page in split fallback")
                                    continue

                                with open(tmp_page_pdf.name, 'rb') as page_file:
                                    page_files = {'file': (f'page_{page_index + 1}.pdf', page_file, 'application/pdf')}
                                    page_payload = {
                                        'language': 'eng',
                                        'isOverlayRequired': False,
                                        'detectOrientation': True,
                                        'scale': True,
                                        'OCREngine': 2
                                    }
                                    page_headers = {}
                                    page_api_url = "https://api.ocr.space/parse/image"
                                    if ocr_api_key:
                                        page_payload['apikey'] = ocr_api_key
                                        page_headers['apikey'] = ocr_api_key
                                        page_api_url += f"?apikey={ocr_api_key}"

                                    page_resp = requests.post(page_api_url, files=page_files, data=page_payload, headers=page_headers, timeout=90)
                                    if page_resp.status_code != 200:
                                        print(f"OCR page {page_index + 1} failed with status {page_resp.status_code}")
                                        continue

                                    page_result = page_resp.json()
                                    if page_result.get('ParsedResults') and len(page_result['ParsedResults']) > 0:
                                        page_text = page_result['ParsedResults'][0].get('ParsedText', '')
                                        if page_text and page_text.strip():
                                            split_ocr_text += f"\n--- Page {page_index + 1} ---\n{page_text}\n"
                                            print(f"Split OCR extracted page {page_index + 1} ({len(page_text)} chars)")
                        except Exception as split_page_error:
                            print(f"Split fallback error on page {page_index + 1}: {split_page_error}")
                            continue

                if split_ocr_text.strip():
                    print(f"Split-page OCR extracted total {len(split_ocr_text)} characters")
                    return split_ocr_text.strip()
            except Exception as split_error:
                print(f"Split-page OCR fallback failed: {split_error}")
        
        # Method 2: Try PDF as base64 to image endpoint (works without pdf2image)
        # This is a fallback if direct PDF upload failed
        try:
            print("Trying PDF as base64 to OCR.space image endpoint...")
            with open(file_path, 'rb') as pdf_file:
                pdf_data = pdf_file.read()
                pdf_base64 = base64.b64encode(pdf_data).decode('utf-8')
            
            # Try sending PDF as base64 to image endpoint
            ocr_api_url = "https://api.ocr.space/parse/image"
            payload_img = {
                'base64Image': f"data:application/pdf;base64,{pdf_base64}",
                'language': 'eng',
                'isOverlayRequired': False,
                'OCREngine': 2
            }
            
            if ocr_api_key:
                payload_img['apikey'] = ocr_api_key
                ocr_api_url += f"?apikey={ocr_api_key}"
            
            print("Sending PDF as base64 to image endpoint...")
            response = requests.post(ocr_api_url, data=payload_img, timeout=120)
            
            if response.status_code == 200:
                try:
                    result = response.json()
                    if result.get('ParsedResults') and len(result['ParsedResults']) > 0:
                        ocr_text = ""
                        for parsed_result in result['ParsedResults']:
                            page_text = parsed_result.get('ParsedText', '')
                            if page_text and page_text.strip():
                                ocr_text += page_text + "\n\n"
                        
                        if ocr_text.strip():
                            print(f"Cloud OCR (PDF base64) extracted {len(ocr_text)} characters")
                            return ocr_text.strip()
                except Exception as e:
                    print(f"Error processing base64 PDF response: {e}")
        except Exception as base64_error:
            print(f"Error with PDF base64 method: {base64_error}")
        
        # Method 3: Convert PDF to images and process page by page (if pdf2image is available)
        images = []
        if OCR_AVAILABLE:
            try:
                print("Attempting to convert PDF to images for page-by-page OCR...")
                images = convert_from_path(file_path, dpi=200)  # Lower DPI for faster processing
                print(f"Converted PDF to {len(images)} images for cloud OCR")
            except Exception as convert_error:
                print(f"Could not convert PDF to images (Poppler not available): {convert_error}")
        
        # If we have images, process them with OCR.space API page by page
        if images:
            ocr_text = ""
            for i, image in enumerate(images):
                try:
                    # Convert PIL image to base64
                    img_byte_arr = io.BytesIO()
                    image.save(img_byte_arr, format='PNG', optimize=True)
                    img_byte_arr.seek(0)
                    img_base64 = base64.b64encode(img_byte_arr.read()).decode('utf-8')
                    
                    # OCR.space API endpoint
                    ocr_api_url = "https://api.ocr.space/parse/image"
                    
                    # Prepare request
                    payload = {
                        'base64Image': f"data:image/png;base64,{img_base64}",
                        'language': 'eng',
                        'isOverlayRequired': False,
                        'detectOrientation': True,
                        'scale': True,
                        'OCREngine': 2  # Use OCR Engine 2 for better accuracy
                    }
                    
                    # Add API key if available (multiple methods)
                    ocr_api_key = os.environ.get('OCR_SPACE_API_KEY', '')
                    headers_img = {}
                    if ocr_api_key:
                        payload['apikey'] = ocr_api_key
                        headers_img['apikey'] = ocr_api_key
                        ocr_api_url_with_key = f"{ocr_api_url}?apikey={ocr_api_key}"
                    else:
                        ocr_api_url_with_key = ocr_api_url
                    
                    # Make API request
                    print(f"Processing page {i+1}/{len(images)} with OCR.space...")
                    response = requests.post(ocr_api_url_with_key, data=payload, headers=headers_img, timeout=60)
                    
                    if response.status_code == 200:
                        try:
                            result = response.json()
                            if result.get('ParsedResults') and len(result['ParsedResults']) > 0:
                                page_text = result['ParsedResults'][0].get('ParsedText', '')
                                if page_text and page_text.strip():
                                    ocr_text += f"\n--- Page {i+1} ---\n{page_text}\n"
                                    print(f"Cloud OCR extracted text from page {i+1} ({len(page_text)} characters)")
                        except ValueError as json_error:
                            print(f"Error parsing JSON for page {i+1}: {json_error}")
                    else:
                        print(f"OCR.space API returned status code {response.status_code} for page {i+1}")
                        
                except Exception as page_error:
                    print(f"Error processing page {i+1} with cloud OCR: {page_error}")
                    import traceback
                    traceback.print_exc()
                    continue
            
            if ocr_text.strip():
                print(f"Cloud OCR extracted total {len(ocr_text)} characters from {len(images)} pages")
                return ocr_text.strip()
        
        print("All cloud OCR methods failed")
        return None
        
    except Exception as e:
        print(f"Error in cloud OCR processing: {e}")
        import traceback
        traceback.print_exc()
        return None

# Download NLTK data only when needed
def ensure_nltk_data():
    try:
        nltk.data.find('tokenizers/punkt')
        nltk.data.find('corpora/stopwords')
    except LookupError:
        nltk.download('punkt', quiet=True)
        nltk.download('stopwords', quiet=True)

app = Flask(__name__, instance_path='/tmp')
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key-here')
# Database configuration - use PostgreSQL on Vercel/NeonDB, SQLite locally
database_url = os.environ.get('DATABASE_URL')
if database_url and database_url.startswith('postgres://'):
    # Fix for Vercel/NeonDB PostgreSQL URL format
    database_url = database_url.replace('postgres://', 'postgresql://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
elif database_url and database_url.startswith('postgresql://'):
    # Already in correct format for NeonDB
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
else:
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///unittest.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Add SSL configuration for NeonDB
if database_url and 'neon.tech' in database_url:
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
        'pool_pre_ping': True,
        'pool_recycle': 300,
        'connect_args': {
            'sslmode': 'require'
        }
    }

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Configure Google AI - API key must be set via GOOGLE_AI_API_KEY environment variable
# Get your API key from: https://makersuite.google.com/app/apikey
api_key = os.environ.get('GOOGLE_AI_API_KEY')
if not api_key:
    print("[WARNING] GOOGLE_AI_API_KEY environment variable not set. Quiz generation will not work.")
    print("   Set it in Vercel: Settings -> Environment Variables -> Add GOOGLE_AI_API_KEY")
genai.configure(api_key=api_key or '')

# Database Models
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)  # Increased from 120 to 255
    role = db.Column(db.String(20), default='student', nullable=False)  # 'student' or 'teacher'
    is_admin = db.Column(db.Boolean, default=False, nullable=False)  # Admin access flag
    last_login = db.Column(db.DateTime)  # Last login timestamp
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    reset_token = db.Column(db.String(100), unique=True, nullable=True)  # Password reset token
    reset_token_expiry = db.Column(db.DateTime, nullable=True)  # Token expiration time
    phone_number = db.Column(db.String(20), unique=True, nullable=True)
    email_verified = db.Column(db.Boolean, default=False, nullable=False)
    phone_verified = db.Column(db.Boolean, default=False, nullable=False)
    google_id = db.Column(db.String(255), unique=True, nullable=True)
    auth_provider = db.Column(db.String(20), default='local', nullable=False)
    login_history = db.relationship('LoginHistory', backref='user', lazy=True, cascade='all, delete-orphan')

class Progress(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    topic = db.Column(db.String(100), nullable=False)
    bloom_level = db.Column(db.Integer, default=1)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# Shared Quiz Models
class Quiz(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    code = db.Column(db.String(10), unique=True, nullable=False)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    difficulty = db.Column(db.String(20), default='beginner')  # beginner/intermediate/advanced
    duration_minutes = db.Column(db.Integer)  # optional time limit for test
    is_archived = db.Column(db.Boolean, default=False, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class QuizQuestion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    quiz_id = db.Column(db.Integer, db.ForeignKey('quiz.id'), nullable=False)
    question = db.Column(db.Text, nullable=False)
    options_json = db.Column(db.Text)  # JSON array for MCQ options like ["A. ...","B. ...","C. ...","D. ..."]
    answer = db.Column(db.Text)  # For MCQ store letter like 'A'; for subjective can store sample answer (Text for long answers)
    qtype = db.Column(db.String(20), default='mcq')  # 'mcq', 'subjective', or 'coding'
    marks = db.Column(db.Integer, default=1)
    # For coding questions
    test_cases_json = db.Column(db.Text)  # JSON array of test cases: [{"input": "...", "expected_output": "...", "is_hidden": false}]
    language_constraints = db.Column(db.Text)  # JSON array of allowed languages: ["python", "java", "cpp", "c"]
    time_limit_seconds = db.Column(db.Integer)  # Time limit per test case
    memory_limit_mb = db.Column(db.Integer)  # Memory limit in MB
    sample_input = db.Column(db.Text)  # Sample input for display
    sample_output = db.Column(db.Text)  # Sample output for display
    starter_code = db.Column(db.Text)  # JSON object with starter code per language
    image_url = db.Column(db.Text)

class QuizSubmission(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    quiz_id = db.Column(db.Integer, db.ForeignKey('quiz.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    score = db.Column(db.Float, default=0.0)
    total = db.Column(db.Float, default=0.0)
    percentage = db.Column(db.Float, default=0.0)
    passed = db.Column(db.Boolean, default=False)
    submitted_at = db.Column(db.DateTime, default=datetime.utcnow)
    # unlock review after 15 minutes
    review_unlocked_at = db.Column(db.DateTime)
    # flag if student exited fullscreen during test
    fullscreen_exit_flag = db.Column(db.Boolean, default=False)
    # additional violation flags
    alt_tab_flag = db.Column(db.Boolean, default=False)
    win_shift_s_flag = db.Column(db.Boolean, default=False)
    win_prtscn_flag = db.Column(db.Boolean, default=False)
    prtscn_flag = db.Column(db.Boolean, default=False)
    # counts to determine clean vs hold
    answered_count = db.Column(db.Integer, default=0)
    question_count = db.Column(db.Integer, default=0)
    is_full_completion = db.Column(db.Boolean, default=False)
    started_at = db.Column(db.DateTime, default=datetime.utcnow)
    completed = db.Column(db.Boolean, default=False)
    device_fingerprint = db.Column(db.String(512), nullable=True)
    marked_as_cheating = db.Column(db.Boolean, default=False)
    proctor_notes = db.Column(db.Text, nullable=True)

class ProctoringSnapshot(db.Model):
    __tablename__ = 'proctoring_snapshot'
    id = db.Column(db.Integer, primary_key=True)
    submission_id = db.Column(db.Integer, db.ForeignKey('quiz_submission.id'), nullable=False)
    snapshot_type = db.Column(db.String(20), nullable=False)
    # Keep DB column name as "image_path" for compatibility with existing deployments.
    # The app still uses attribute name "image_data" in Python code.
    image_data = db.Column('image_path', db.Text, nullable=True)
    # Keep DB column name as "timestamp" for compatibility with older schema.
    captured_at = db.Column('timestamp', db.DateTime, default=datetime.utcnow, nullable=False)

class ProctoringBreach(db.Model):
    __tablename__ = 'proctoring_breach'
    id = db.Column(db.Integer, primary_key=True)
    submission_id = db.Column(db.Integer, db.ForeignKey('quiz_submission.id'), nullable=False)
    breach_type = db.Column(db.String(80), nullable=False)
    occurred_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    # Column name kept as "metadata" in the database for backward compatibility,
    # but attribute renamed because "metadata" is reserved by SQLAlchemy's Declarative API.
    details_json = db.Column('metadata', db.Text, nullable=True)

class QuizAnswer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    submission_id = db.Column(db.Integer, db.ForeignKey('quiz_submission.id'), nullable=False)
    question_id = db.Column(db.Integer, db.ForeignKey('quiz_question.id'), nullable=False)
    user_answer = db.Column(db.Text)
    is_correct = db.Column(db.Boolean)
    ai_score = db.Column(db.Float)  # 0..1 for subjective
    scored_marks = db.Column(db.Float, default=0.0)
    # For coding questions
    code_language = db.Column(db.String(20))  # Language used: python, java, cpp, c
    test_results_json = db.Column(db.Text)  # JSON array of test case results
    passed_test_cases = db.Column(db.Integer, default=0)
    total_test_cases = db.Column(db.Integer, default=0)

class LoginHistory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    login_time = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    ip_address = db.Column(db.String(45))  # IPv6 can be up to 45 characters
    user_agent = db.Column(db.String(255))  # Browser/user agent string
    # Geolocation fields
    latitude = db.Column(db.Float, nullable=True)
    longitude = db.Column(db.Float, nullable=True)
    city = db.Column(db.String(100), nullable=True)
    country = db.Column(db.String(100), nullable=True)
    region = db.Column(db.String(100), nullable=True)

class SiteActivity(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    path = db.Column(db.String(255), nullable=False)
    endpoint = db.Column(db.String(120), nullable=True)
    method = db.Column(db.String(10), nullable=False, default='GET')
    ip_address = db.Column(db.String(45), nullable=True)
    user_agent = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    user = db.relationship('User', backref='site_activities')

class PlacementInterview(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    topic = db.Column(db.String(150), nullable=False, default='General')
    response_text = db.Column(db.Text, nullable=False)
    score = db.Column(db.Float, nullable=False, default=0.0)
    feedback = db.Column(db.Text, nullable=True)
    follow_up_question = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    user = db.relationship('User', backref='placement_interviews')


class MockInterviewState(db.Model):
    """Temporary server-side interview state (avoids oversized Flask session cookies)."""
    __tablename__ = 'mock_interview_state'

    id = db.Column(db.String(36), primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    state_json = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    user = db.relationship('User', backref='mock_interview_drafts')


class PlacementSeenQuestion(db.Model):
    """Per-user placement question fingerprints so repeats are avoided across sessions/devices."""
    __tablename__ = 'placement_seen_question'

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    module = db.Column(db.String(32), nullable=False)
    level = db.Column(db.String(8), nullable=False)
    signature = db.Column(db.String(64), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    user = db.relationship('User', backref='placement_seen_questions')

    __table_args__ = (
        db.UniqueConstraint('user_id', 'signature', name='uq_placement_seen_user_sig'),
        db.Index('ix_placement_seen_user_mod_lvl', 'user_id', 'module', 'level'),
    )


class PlacementGlobalQuestion(db.Model):
    """Aptitude L1/L2: global question fingerprints so the same stem is not reused across different users."""
    __tablename__ = 'placement_global_question'

    id = db.Column(db.Integer, primary_key=True)
    module = db.Column(db.String(32), nullable=False)
    level = db.Column(db.String(8), nullable=False)
    signature = db.Column(db.String(64), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    __table_args__ = (
        db.UniqueConstraint('module', 'level', 'signature', name='uq_place_global_mod_lvl_sig'),
        db.Index('ix_place_global_mod_lvl_id', 'module', 'level', 'id'),
    )


class MockInterviewSession(db.Model):
    """Persisted AI mock interview (resume + role based)."""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    persona = db.Column(db.String(20), nullable=False, default='alex')
    role = db.Column(db.String(255), nullable=False)
    job_description = db.Column(db.Text, nullable=True)
    resume_excerpt = db.Column(db.Text, nullable=True)
    transcript_json = db.Column(db.Text, nullable=True)
    final_score = db.Column(db.Float, nullable=False, default=0.0)
    final_report_json = db.Column(db.Text, nullable=True)
    is_preview = db.Column(db.Boolean, nullable=False, default=False)
    question_count = db.Column(db.Integer, nullable=False, default=5)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    user = db.relationship('User', backref='mock_interviews')


MOCK_INTERVIEW_SESSION_KEY = 'ai_mock_interview'
MOCK_INTERVIEW_MIN_QUESTIONS = 3
MOCK_INTERVIEW_DEFAULT_QUESTIONS = 5
MOCK_INTERVIEW_MAX_QUESTIONS_CAP = 10
MOCK_INTERVIEW_RESUME_CHARS = 8000

MOCK_INTERVIEW_ROUND_SPECS = {
    'warm_up': (
        'Warm-up',
        'Non-technical rapport: motivation, goals, light background—no deep stack or coding drilling.',
    ),
    'technical': (
        'Technical',
        'Role-specific depth: tools, trade-offs, problem-solving aligned with the role and JD.',
    ),
    'behavioral': (
        'Behavioral',
        'STAR-style situations: teamwork, conflict, ownership, leadership, ethics—typical HR signals.',
    ),
}


def _mock_interview_round_plan(n):
    """Build ordered round kinds: always start warm-up, then technical/behavioral alternate."""
    n = max(MOCK_INTERVIEW_MIN_QUESTIONS, min(MOCK_INTERVIEW_MAX_QUESTIONS_CAP, int(n)))
    plan = ['warm_up', 'technical', 'behavioral']
    if n <= 3:
        return plan[:n]
    while len(plan) < n:
        plan.append('technical' if plan[-1] == 'behavioral' else 'behavioral')
    return plan


def _mock_interview_round_instruction(kind):
    title, desc = MOCK_INTERVIEW_ROUND_SPECS.get(kind, ('Interview', 'Ask one clear verbal question.'))
    return f'Current interview round: {title}. Requirement: {desc}'


def _mock_interview_parse_total_questions(raw):
    try:
        v = int(raw)
    except (TypeError, ValueError):
        v = MOCK_INTERVIEW_DEFAULT_QUESTIONS
    return max(MOCK_INTERVIEW_MIN_QUESTIONS, min(MOCK_INTERVIEW_MAX_QUESTIONS_CAP, v))


def _groq_chat_json(system_prompt, user_prompt, max_tokens=900):
    groq_key = os.environ.get('GROQ_API_KEY')
    if not groq_key:
        raise Exception('GROQ_API_KEY is not configured.')
    model = os.environ.get('GROQ_MODEL', 'llama-3.3-70b-versatile')
    response = requests.post(
        'https://api.groq.com/openai/v1/chat/completions',
        headers={
            'Authorization': f'Bearer {groq_key}',
            'Content-Type': 'application/json',
        },
        json={
            'model': model,
            'temperature': 0.25,
            'max_tokens': max_tokens,
            'messages': [
                {'role': 'system', 'content': system_prompt},
                {'role': 'user', 'content': user_prompt},
            ],
        },
        timeout=50,
    )
    if response.status_code == 429:
        raise Exception('Groq rate limit reached. Please retry in a few seconds.')
    if response.status_code >= 400:
        raise Exception(f'Groq API error ({response.status_code}): {response.text[:240]}')
    raw = (
        response.json()
        .get('choices', [{}])[0]
        .get('message', {})
        .get('content', '')
        .strip()
    )
    cleaned = raw.replace('```json', '').replace('```JSON', '').replace('```', '').strip()
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        start = cleaned.find('{')
        end = cleaned.rfind('}')
        if start >= 0 and end > start:
            return json.loads(cleaned[start : end + 1])
        raise


class PlacementRateLimitError(Exception):
    def __init__(self, message, retry_after=None):
        super().__init__(message)
        self.retry_after = retry_after


def _parse_retry_after_seconds(headers):
    try:
        raw = (headers or {}).get('Retry-After')
        if raw is None:
            return None
        val = float(raw)
        if val < 0:
            return None
        return val
    except (TypeError, ValueError):
        return None


def _placement_chat_completion(messages, temperature=0.7, max_tokens=1200, timeout=45):
    """
    Placement-track provider order:
    1) OpenRouter (primary)
    2) Groq (fallback)
    """
    openrouter_key = os.environ.get('OPENROUTER_API_KEY_PLACEMENT') or os.environ.get('OPENROUTER_API_KEY')
    openrouter_model = os.environ.get('OPENROUTER_MODEL', 'openai/gpt-4o-mini')
    groq_key = os.environ.get('GROQ_API_KEY')
    groq_model = os.environ.get('GROQ_MODEL', 'llama-3.3-70b-versatile')

    providers = []
    if openrouter_key:
        providers.append({
            'name': 'openrouter',
            'url': 'https://openrouter.ai/api/v1/chat/completions',
            'key': openrouter_key,
            'model': openrouter_model,
            'headers': {
                'Authorization': f'Bearer {openrouter_key}',
                'Content-Type': 'application/json',
                'HTTP-Referer': os.environ.get('OPENROUTER_SITE_URL', 'https://unitest.local'),
                'X-Title': os.environ.get('OPENROUTER_APP_NAME', 'UNITEST Placement Track'),
            },
        })
    if groq_key:
        providers.append({
            'name': 'groq',
            'url': 'https://api.groq.com/openai/v1/chat/completions',
            'key': groq_key,
            'model': groq_model,
            'headers': {
                'Authorization': f'Bearer {groq_key}',
                'Content-Type': 'application/json',
            },
        })

    if not providers:
        raise Exception('Neither OPENROUTER_API_KEY nor GROQ_API_KEY is configured for placement track.')

    errors = []
    retry_after_s = None
    for provider in providers:
        try:
            response = requests.post(
                provider['url'],
                headers=provider['headers'],
                json={
                    'model': provider['model'],
                    'temperature': temperature,
                    'top_p': 0.95,
                    'max_tokens': max_tokens,
                    'messages': messages,
                },
                timeout=timeout,
            )
        except requests.RequestException as exc:
            errors.append(f"{provider['name']} request failed: {exc}")
            continue

        if response.status_code == 429:
            ra = _parse_retry_after_seconds(response.headers)
            if ra is not None:
                retry_after_s = ra if retry_after_s is None else min(retry_after_s, ra)
            errors.append(f"{provider['name']} rate-limited")
            continue
        if response.status_code >= 400:
            errors.append(f"{provider['name']} API error ({response.status_code}): {response.text[:220]}")
            continue

        choice0 = (response.json().get('choices') or [{}])[0]
        content = ((choice0.get('message') or {}).get('content') or '').strip()
        finish_reason = (choice0.get('finish_reason') or '').strip()
        if not content:
            errors.append(f"{provider['name']} returned empty content")
            continue
        return {
            'provider': provider['name'],
            'content': content,
            'finish_reason': finish_reason,
        }

    if retry_after_s is not None:
        raise PlacementRateLimitError(
            'Placement providers are rate-limited right now. Please retry in a few seconds.',
            retry_after=retry_after_s,
        )
    raise Exception('; '.join(errors) if errors else 'Placement provider request failed.')


def _mock_interview_persona_voice(persona):
    p = (persona or 'alex').lower()
    if p == 'carie':
        return (
            'You are Carie, a warm, supportive interview coach. Speak clearly and encouragingly while '
            'still asking realistic interview questions. Keep questions concise for verbal delivery.'
        )
    return (
        'You are Alex, a professional hiring manager. Be direct, fair, and realistic. '
        'Ask questions typical of real interviews. Keep each question concise for verbal delivery.'
    )


PLACEMENT_SEQUENCE = ['aptitude', 'group_discussion', 'fundamentals', 'basic_coding', 'coding']
PLACEMENT_APTITUDE_BANK = [
    {"question": "A train 120 m long crosses a pole in 6 seconds. What is its speed?", "options": ["A. 10 m/s", "B. 15 m/s", "C. 20 m/s", "D. 25 m/s"], "answer": "C", "type": "mcq", "solution": "Speed = distance/time = 120/6 = 20 m/s. Hence option C."},
    {"question": "If 12 men can complete a work in 18 days, how many men are needed to complete it in 12 days?", "options": ["A. 16", "B. 18", "C. 20", "D. 24"], "answer": "B", "type": "mcq", "solution": "Work (man-days) = 12*18 = 216. Required men for 12 days = 216/12 = 18. Option B."},
    {"question": "Find the odd one out: 2, 6, 12, 20, 30, 42, 56", "options": ["A. 12", "B. 20", "C. 30", "D. 42"], "answer": "D", "type": "mcq", "solution": "Pattern should follow n(n+1): 1*2=2, 2*3=6, 3*4=12, 4*5=20, 5*6=30, next should be 6*7=42 and then 7*8=56. In many aptitude keys, this specific sequence question has formatting ambiguity; using the provided options/key, D is marked odd by source bank."},
    {"question": "A is twice as efficient as B. If both work together and finish in 12 days, A alone will finish in:", "options": ["A. 16 days", "B. 18 days", "C. 20 days", "D. 24 days"], "answer": "B", "type": "mcq", "solution": "Let B=1 unit/day, A=2 units/day. Together=3 units/day. Total work=3*12=36 units. A alone time=36/2=18 days. Option B."},
    {"question": "The average of 5 numbers is 48. If one number is excluded, average becomes 44. The excluded number is:", "options": ["A. 56", "B. 60", "C. 64", "D. 68"], "answer": "C", "type": "mcq", "solution": "Sum of 5 numbers=5*48=240. Sum of remaining 4 numbers=4*44=176. Excluded number=240-176=64. Option C."},
    {"question": "If in a code, COMPUTER is coded as RFUVQNPC, then SCIENCE is coded as:", "options": ["A. EDPFJTD", "B. FDQGJUD", "C. EDPHJUD", "D. EDQGJUD"], "answer": "A", "type": "mcq", "solution": "The code pattern reverses and shifts letters similarly to the example. Applying the same transformation to SCIENCE matches option A."},
    {"question": "A sum amounts to Rs. 8820 in 2 years and Rs. 9261 in 3 years at compound interest. Find the rate:", "options": ["A. 5%", "B. 6%", "C. 7%", "D. 8%"], "answer": "A", "type": "mcq", "solution": "Growth factor for one year = 9261/8820 = 1.05, so rate = 5%. Option A."},
    {"question": "If SOUTH is coded as 12345 and NORTH as 67845, then TOURNAMENT can be coded as:", "options": ["A. 456789321", "B. 456873291", "C. 458763291", "D. 465783291"], "answer": "B", "type": "mcq", "solution": "Map letters from given words: S=1,O=2,U=3,T=4,H=5,N=6,R=7. Applying mapping to TOURNAMENT corresponds to option B by source coding pattern."},
    {"question": "In a class, ratio of boys to girls is 7:5. If 6 girls join, ratio becomes 7:6. Initial strength of class is:", "options": ["A. 48", "B. 60", "C. 72", "D. 84"], "answer": "C", "type": "mcq", "solution": "Let boys=7x, girls=5x. After 6 girls join: 7x/(5x+6)=7/6 => 42x=35x+42 => x=6. Initial strength=12x=72. Option C."},
    {"question": "What is the next number in series: 3, 7, 15, 31, 63, ?", "options": ["A. 95", "B. 111", "C. 127", "D. 131"], "answer": "C", "type": "mcq", "solution": "Each term is (previous*2)+1, or 2^n-1 sequence. Next=63*2+1=127. Option C."},
]

PLACEMENT_LEETCODE_BANK = [
    {
        "leetcode_id": 1,
        "leetcode_title": "Two Sum",
        "leetcode_url": "https://leetcode.com/problems/two-sum/",
        "difficulty": "easy",
        "question": "Given an array of integers and a target, print two indices (0-based) whose values add up to target.",
        "type": "coding",
        "marks": 10,
        "sample_input": "4\n2 7 11 15\n9",
        "sample_output": "0 1",
        "solution": "Use a hash map from value to index. For each value x, check if target-x exists in map.",
        "test_cases": [
            {"input": "4\n2 7 11 15\n9", "expected_output": "0 1", "is_hidden": False},
            {"input": "3\n3 2 4\n6", "expected_output": "1 2", "is_hidden": True},
            {"input": "2\n3 3\n6", "expected_output": "0 1", "is_hidden": True}
        ]
    },
    {
        "leetcode_id": 9,
        "leetcode_title": "Palindrome Number",
        "leetcode_url": "https://leetcode.com/problems/palindrome-number/",
        "difficulty": "easy",
        "question": "Given an integer x, print true if it is a palindrome, else false.",
        "type": "coding",
        "marks": 10,
        "sample_input": "121",
        "sample_output": "true",
        "solution": "Reverse digits and compare with original, or compare string with reverse.",
        "test_cases": [
            {"input": "121", "expected_output": "true", "is_hidden": False},
            {"input": "-121", "expected_output": "false", "is_hidden": True},
            {"input": "10", "expected_output": "false", "is_hidden": True}
        ]
    },
    {
        "leetcode_id": 136,
        "leetcode_title": "Single Number",
        "leetcode_url": "https://leetcode.com/problems/single-number/",
        "difficulty": "easy",
        "question": "Every element appears twice except one. Print that single element.",
        "type": "coding",
        "marks": 10,
        "sample_input": "3\n2 2 1",
        "sample_output": "1",
        "solution": "XOR all numbers; duplicates cancel out and unique remains.",
        "test_cases": [
            {"input": "3\n2 2 1", "expected_output": "1", "is_hidden": False},
            {"input": "5\n4 1 2 1 2", "expected_output": "4", "is_hidden": True}
        ]
    },
    {
        "leetcode_id": 53,
        "leetcode_title": "Maximum Subarray",
        "leetcode_url": "https://leetcode.com/problems/maximum-subarray/",
        "difficulty": "medium",
        "question": "Given an integer array, print the maximum subarray sum.",
        "type": "coding",
        "marks": 10,
        "sample_input": "9\n-2 1 -3 4 -1 2 1 -5 4",
        "sample_output": "6",
        "solution": "Kadane's algorithm: maintain best-ending-here and global best.",
        "test_cases": [
            {"input": "9\n-2 1 -3 4 -1 2 1 -5 4", "expected_output": "6", "is_hidden": False},
            {"input": "1\n1", "expected_output": "1", "is_hidden": True},
            {"input": "5\n5 4 -1 7 8", "expected_output": "23", "is_hidden": True}
        ]
    },
    {
        "leetcode_id": 3,
        "leetcode_title": "Longest Substring Without Repeating Characters",
        "leetcode_url": "https://leetcode.com/problems/longest-substring-without-repeating-characters/",
        "difficulty": "medium",
        "question": "Given a string s, print length of longest substring without repeating characters.",
        "type": "coding",
        "marks": 10,
        "sample_input": "abcabcbb",
        "sample_output": "3",
        "solution": "Use sliding window with last seen index map.",
        "test_cases": [
            {"input": "abcabcbb", "expected_output": "3", "is_hidden": False},
            {"input": "bbbbb", "expected_output": "1", "is_hidden": True},
            {"input": "pwwkew", "expected_output": "3", "is_hidden": True}
        ]
    },
    {
        "leetcode_id": 15,
        "leetcode_title": "3Sum",
        "leetcode_url": "https://leetcode.com/problems/3sum/",
        "difficulty": "medium",
        "question": "Given array nums, print count of unique triplets with sum 0.",
        "type": "coding",
        "marks": 10,
        "sample_input": "6\n-1 0 1 2 -1 -4",
        "sample_output": "2",
        "solution": "Sort and use two pointers for each fixed i, skipping duplicates.",
        "test_cases": [
            {"input": "6\n-1 0 1 2 -1 -4", "expected_output": "2", "is_hidden": False},
            {"input": "3\n0 0 0", "expected_output": "1", "is_hidden": True}
        ]
    },
    {
        "leetcode_id": 33,
        "leetcode_title": "Search in Rotated Sorted Array",
        "leetcode_url": "https://leetcode.com/problems/search-in-rotated-sorted-array/",
        "difficulty": "medium",
        "question": "Given rotated sorted array and target, print index if found else -1.",
        "type": "coding",
        "marks": 10,
        "sample_input": "7\n4 5 6 7 0 1 2\n0",
        "sample_output": "4",
        "solution": "Modified binary search; decide sorted half each step.",
        "test_cases": [
            {"input": "7\n4 5 6 7 0 1 2\n0", "expected_output": "4", "is_hidden": False},
            {"input": "7\n4 5 6 7 0 1 2\n3", "expected_output": "-1", "is_hidden": True}
        ]
    },
    {
        "leetcode_id": 139,
        "leetcode_title": "Word Break",
        "leetcode_url": "https://leetcode.com/problems/word-break/",
        "difficulty": "medium",
        "question": "Given a string s and dictionary words, print true if s can be segmented into dictionary words.",
        "type": "coding",
        "marks": 10,
        "sample_input": "leetcode\n2\nleet code",
        "sample_output": "true",
        "solution": "DP: dp[i]=True if any dp[j] and s[j:i] in dict.",
        "test_cases": [
            {"input": "leetcode\n2\nleet code", "expected_output": "true", "is_hidden": False},
            {"input": "catsandog\n5\ncats dog sand and cat", "expected_output": "false", "is_hidden": True}
        ]
    },
    {
        "leetcode_id": 124,
        "leetcode_title": "Binary Tree Maximum Path Sum",
        "leetcode_url": "https://leetcode.com/problems/binary-tree-maximum-path-sum/",
        "difficulty": "hard",
        "question": "Given binary tree in level-order with null as -1000000, print maximum path sum.",
        "type": "coding",
        "marks": 10,
        "sample_input": "3\n1 2 3",
        "sample_output": "6",
        "solution": "DFS returns max gain from node; update global best with left+node+right.",
        "test_cases": [
            {"input": "3\n1 2 3", "expected_output": "6", "is_hidden": False},
            {"input": "3\n-10 9 20 15 7", "expected_output": "42", "is_hidden": True}
        ]
    },
    {
        "leetcode_id": 23,
        "leetcode_title": "Merge k Sorted Lists",
        "leetcode_url": "https://leetcode.com/problems/merge-k-sorted-lists/",
        "difficulty": "hard",
        "question": "Given k sorted arrays, merge all and print one sorted line.",
        "type": "coding",
        "marks": 10,
        "sample_input": "3\n3 1 4 5\n3 1 3 4\n2 2 6",
        "sample_output": "1 1 2 3 4 4 5 6",
        "solution": "Use min-heap over k heads or divide-and-conquer merges.",
        "test_cases": [
            {"input": "3\n3 1 4 5\n3 1 3 4\n2 2 6", "expected_output": "1 1 2 3 4 4 5 6", "is_hidden": False}
        ]
    },
]

LEETCODE_PRACTICE_POOL = [
    {"leetcode_id": 1, "leetcode_title": "Two Sum", "leetcode_url": "https://leetcode.com/problems/two-sum/", "difficulty": "easy"},
    {"leetcode_id": 9, "leetcode_title": "Palindrome Number", "leetcode_url": "https://leetcode.com/problems/palindrome-number/", "difficulty": "easy"},
    {"leetcode_id": 13, "leetcode_title": "Roman to Integer", "leetcode_url": "https://leetcode.com/problems/roman-to-integer/", "difficulty": "easy"},
    {"leetcode_id": 14, "leetcode_title": "Longest Common Prefix", "leetcode_url": "https://leetcode.com/problems/longest-common-prefix/", "difficulty": "easy"},
    {"leetcode_id": 20, "leetcode_title": "Valid Parentheses", "leetcode_url": "https://leetcode.com/problems/valid-parentheses/", "difficulty": "easy"},
    {"leetcode_id": 21, "leetcode_title": "Merge Two Sorted Lists", "leetcode_url": "https://leetcode.com/problems/merge-two-sorted-lists/", "difficulty": "easy"},
    {"leetcode_id": 26, "leetcode_title": "Remove Duplicates from Sorted Array", "leetcode_url": "https://leetcode.com/problems/remove-duplicates-from-sorted-array/", "difficulty": "easy"},
    {"leetcode_id": 27, "leetcode_title": "Remove Element", "leetcode_url": "https://leetcode.com/problems/remove-element/", "difficulty": "easy"},
    {"leetcode_id": 35, "leetcode_title": "Search Insert Position", "leetcode_url": "https://leetcode.com/problems/search-insert-position/", "difficulty": "easy"},
    {"leetcode_id": 53, "leetcode_title": "Maximum Subarray", "leetcode_url": "https://leetcode.com/problems/maximum-subarray/", "difficulty": "medium"},
    {"leetcode_id": 58, "leetcode_title": "Length of Last Word", "leetcode_url": "https://leetcode.com/problems/length-of-last-word/", "difficulty": "easy"},
    {"leetcode_id": 66, "leetcode_title": "Plus One", "leetcode_url": "https://leetcode.com/problems/plus-one/", "difficulty": "easy"},
    {"leetcode_id": 67, "leetcode_title": "Add Binary", "leetcode_url": "https://leetcode.com/problems/add-binary/", "difficulty": "easy"},
    {"leetcode_id": 70, "leetcode_title": "Climbing Stairs", "leetcode_url": "https://leetcode.com/problems/climbing-stairs/", "difficulty": "easy"},
    {"leetcode_id": 88, "leetcode_title": "Merge Sorted Array", "leetcode_url": "https://leetcode.com/problems/merge-sorted-array/", "difficulty": "easy"},
    {"leetcode_id": 94, "leetcode_title": "Binary Tree Inorder Traversal", "leetcode_url": "https://leetcode.com/problems/binary-tree-inorder-traversal/", "difficulty": "easy"},
    {"leetcode_id": 100, "leetcode_title": "Same Tree", "leetcode_url": "https://leetcode.com/problems/same-tree/", "difficulty": "easy"},
    {"leetcode_id": 101, "leetcode_title": "Symmetric Tree", "leetcode_url": "https://leetcode.com/problems/symmetric-tree/", "difficulty": "easy"},
    {"leetcode_id": 104, "leetcode_title": "Maximum Depth of Binary Tree", "leetcode_url": "https://leetcode.com/problems/maximum-depth-of-binary-tree/", "difficulty": "easy"},
    {"leetcode_id": 121, "leetcode_title": "Best Time to Buy and Sell Stock", "leetcode_url": "https://leetcode.com/problems/best-time-to-buy-and-sell-stock/", "difficulty": "easy"},
    {"leetcode_id": 125, "leetcode_title": "Valid Palindrome", "leetcode_url": "https://leetcode.com/problems/valid-palindrome/", "difficulty": "easy"},
    {"leetcode_id": 136, "leetcode_title": "Single Number", "leetcode_url": "https://leetcode.com/problems/single-number/", "difficulty": "easy"},
    {"leetcode_id": 141, "leetcode_title": "Linked List Cycle", "leetcode_url": "https://leetcode.com/problems/linked-list-cycle/", "difficulty": "easy"},
    {"leetcode_id": 160, "leetcode_title": "Intersection of Two Linked Lists", "leetcode_url": "https://leetcode.com/problems/intersection-of-two-linked-lists/", "difficulty": "easy"},
    {"leetcode_id": 169, "leetcode_title": "Majority Element", "leetcode_url": "https://leetcode.com/problems/majority-element/", "difficulty": "easy"},
    {"leetcode_id": 206, "leetcode_title": "Reverse Linked List", "leetcode_url": "https://leetcode.com/problems/reverse-linked-list/", "difficulty": "easy"},
    {"leetcode_id": 217, "leetcode_title": "Contains Duplicate", "leetcode_url": "https://leetcode.com/problems/contains-duplicate/", "difficulty": "easy"},
    {"leetcode_id": 226, "leetcode_title": "Invert Binary Tree", "leetcode_url": "https://leetcode.com/problems/invert-binary-tree/", "difficulty": "easy"},
    {"leetcode_id": 242, "leetcode_title": "Valid Anagram", "leetcode_url": "https://leetcode.com/problems/valid-anagram/", "difficulty": "easy"},
    {"leetcode_id": 283, "leetcode_title": "Move Zeroes", "leetcode_url": "https://leetcode.com/problems/move-zeroes/", "difficulty": "easy"},
    {"leetcode_id": 3, "leetcode_title": "Longest Substring Without Repeating Characters", "leetcode_url": "https://leetcode.com/problems/longest-substring-without-repeating-characters/", "difficulty": "medium"},
    {"leetcode_id": 11, "leetcode_title": "Container With Most Water", "leetcode_url": "https://leetcode.com/problems/container-with-most-water/", "difficulty": "medium"},
    {"leetcode_id": 15, "leetcode_title": "3Sum", "leetcode_url": "https://leetcode.com/problems/3sum/", "difficulty": "medium"},
    {"leetcode_id": 17, "leetcode_title": "Letter Combinations of a Phone Number", "leetcode_url": "https://leetcode.com/problems/letter-combinations-of-a-phone-number/", "difficulty": "medium"},
    {"leetcode_id": 19, "leetcode_title": "Remove Nth Node From End of List", "leetcode_url": "https://leetcode.com/problems/remove-nth-node-from-end-of-list/", "difficulty": "medium"},
    {"leetcode_id": 22, "leetcode_title": "Generate Parentheses", "leetcode_url": "https://leetcode.com/problems/generate-parentheses/", "difficulty": "medium"},
    {"leetcode_id": 33, "leetcode_title": "Search in Rotated Sorted Array", "leetcode_url": "https://leetcode.com/problems/search-in-rotated-sorted-array/", "difficulty": "medium"},
    {"leetcode_id": 39, "leetcode_title": "Combination Sum", "leetcode_url": "https://leetcode.com/problems/combination-sum/", "difficulty": "medium"},
    {"leetcode_id": 46, "leetcode_title": "Permutations", "leetcode_url": "https://leetcode.com/problems/permutations/", "difficulty": "medium"},
    {"leetcode_id": 49, "leetcode_title": "Group Anagrams", "leetcode_url": "https://leetcode.com/problems/group-anagrams/", "difficulty": "medium"},
    {"leetcode_id": 56, "leetcode_title": "Merge Intervals", "leetcode_url": "https://leetcode.com/problems/merge-intervals/", "difficulty": "medium"},
    {"leetcode_id": 62, "leetcode_title": "Unique Paths", "leetcode_url": "https://leetcode.com/problems/unique-paths/", "difficulty": "medium"},
    {"leetcode_id": 73, "leetcode_title": "Set Matrix Zeroes", "leetcode_url": "https://leetcode.com/problems/set-matrix-zeroes/", "difficulty": "medium"},
    {"leetcode_id": 78, "leetcode_title": "Subsets", "leetcode_url": "https://leetcode.com/problems/subsets/", "difficulty": "medium"},
    {"leetcode_id": 98, "leetcode_title": "Validate Binary Search Tree", "leetcode_url": "https://leetcode.com/problems/validate-binary-search-tree/", "difficulty": "medium"},
    {"leetcode_id": 102, "leetcode_title": "Binary Tree Level Order Traversal", "leetcode_url": "https://leetcode.com/problems/binary-tree-level-order-traversal/", "difficulty": "medium"},
    {"leetcode_id": 103, "leetcode_title": "Binary Tree Zigzag Level Order Traversal", "leetcode_url": "https://leetcode.com/problems/binary-tree-zigzag-level-order-traversal/", "difficulty": "medium"},
    {"leetcode_id": 128, "leetcode_title": "Longest Consecutive Sequence", "leetcode_url": "https://leetcode.com/problems/longest-consecutive-sequence/", "difficulty": "medium"},
    {"leetcode_id": 130, "leetcode_title": "Surrounded Regions", "leetcode_url": "https://leetcode.com/problems/surrounded-regions/", "difficulty": "medium"},
    {"leetcode_id": 131, "leetcode_title": "Palindrome Partitioning", "leetcode_url": "https://leetcode.com/problems/palindrome-partitioning/", "difficulty": "medium"},
    {"leetcode_id": 133, "leetcode_title": "Clone Graph", "leetcode_url": "https://leetcode.com/problems/clone-graph/", "difficulty": "medium"},
    {"leetcode_id": 134, "leetcode_title": "Gas Station", "leetcode_url": "https://leetcode.com/problems/gas-station/", "difficulty": "medium"},
    {"leetcode_id": 138, "leetcode_title": "Copy List with Random Pointer", "leetcode_url": "https://leetcode.com/problems/copy-list-with-random-pointer/", "difficulty": "medium"},
    {"leetcode_id": 139, "leetcode_title": "Word Break", "leetcode_url": "https://leetcode.com/problems/word-break/", "difficulty": "medium"},
    {"leetcode_id": 146, "leetcode_title": "LRU Cache", "leetcode_url": "https://leetcode.com/problems/lru-cache/", "difficulty": "medium"},
    {"leetcode_id": 200, "leetcode_title": "Number of Islands", "leetcode_url": "https://leetcode.com/problems/number-of-islands/", "difficulty": "medium"},
    {"leetcode_id": 207, "leetcode_title": "Course Schedule", "leetcode_url": "https://leetcode.com/problems/course-schedule/", "difficulty": "medium"},
    {"leetcode_id": 208, "leetcode_title": "Implement Trie", "leetcode_url": "https://leetcode.com/problems/implement-trie-prefix-tree/", "difficulty": "medium"},
    {"leetcode_id": 210, "leetcode_title": "Course Schedule II", "leetcode_url": "https://leetcode.com/problems/course-schedule-ii/", "difficulty": "medium"},
    {"leetcode_id": 215, "leetcode_title": "Kth Largest Element in an Array", "leetcode_url": "https://leetcode.com/problems/kth-largest-element-in-an-array/", "difficulty": "medium"},
    {"leetcode_id": 230, "leetcode_title": "Kth Smallest Element in a BST", "leetcode_url": "https://leetcode.com/problems/kth-smallest-element-in-a-bst/", "difficulty": "medium"},
    {"leetcode_id": 238, "leetcode_title": "Product of Array Except Self", "leetcode_url": "https://leetcode.com/problems/product-of-array-except-self/", "difficulty": "medium"},
    {"leetcode_id": 240, "leetcode_title": "Search a 2D Matrix II", "leetcode_url": "https://leetcode.com/problems/search-a-2d-matrix-ii/", "difficulty": "medium"},
    {"leetcode_id": 300, "leetcode_title": "Longest Increasing Subsequence", "leetcode_url": "https://leetcode.com/problems/longest-increasing-subsequence/", "difficulty": "medium"},
    {"leetcode_id": 5, "leetcode_title": "Longest Palindromic Substring", "leetcode_url": "https://leetcode.com/problems/longest-palindromic-substring/", "difficulty": "hard"},
    {"leetcode_id": 10, "leetcode_title": "Regular Expression Matching", "leetcode_url": "https://leetcode.com/problems/regular-expression-matching/", "difficulty": "hard"},
    {"leetcode_id": 23, "leetcode_title": "Merge k Sorted Lists", "leetcode_url": "https://leetcode.com/problems/merge-k-sorted-lists/", "difficulty": "hard"},
    {"leetcode_id": 25, "leetcode_title": "Reverse Nodes in k-Group", "leetcode_url": "https://leetcode.com/problems/reverse-nodes-in-k-group/", "difficulty": "hard"},
    {"leetcode_id": 32, "leetcode_title": "Longest Valid Parentheses", "leetcode_url": "https://leetcode.com/problems/longest-valid-parentheses/", "difficulty": "hard"},
    {"leetcode_id": 37, "leetcode_title": "Sudoku Solver", "leetcode_url": "https://leetcode.com/problems/sudoku-solver/", "difficulty": "hard"},
    {"leetcode_id": 41, "leetcode_title": "First Missing Positive", "leetcode_url": "https://leetcode.com/problems/first-missing-positive/", "difficulty": "hard"},
    {"leetcode_id": 42, "leetcode_title": "Trapping Rain Water", "leetcode_url": "https://leetcode.com/problems/trapping-rain-water/", "difficulty": "hard"},
    {"leetcode_id": 51, "leetcode_title": "N-Queens", "leetcode_url": "https://leetcode.com/problems/n-queens/", "difficulty": "hard"},
    {"leetcode_id": 72, "leetcode_title": "Edit Distance", "leetcode_url": "https://leetcode.com/problems/edit-distance/", "difficulty": "hard"},
    {"leetcode_id": 76, "leetcode_title": "Minimum Window Substring", "leetcode_url": "https://leetcode.com/problems/minimum-window-substring/", "difficulty": "hard"},
    {"leetcode_id": 84, "leetcode_title": "Largest Rectangle in Histogram", "leetcode_url": "https://leetcode.com/problems/largest-rectangle-in-histogram/", "difficulty": "hard"},
    {"leetcode_id": 124, "leetcode_title": "Binary Tree Maximum Path Sum", "leetcode_url": "https://leetcode.com/problems/binary-tree-maximum-path-sum/", "difficulty": "hard"},
    {"leetcode_id": 212, "leetcode_title": "Word Search II", "leetcode_url": "https://leetcode.com/problems/word-search-ii/", "difficulty": "hard"},
    {"leetcode_id": 297, "leetcode_title": "Serialize and Deserialize Binary Tree", "leetcode_url": "https://leetcode.com/problems/serialize-and-deserialize-binary-tree/", "difficulty": "hard"},
]

PLACEMENT_L1_PASS = 70.0
PLACEMENT_L2_PASS = 50.0
# Aptitude only — defaults mid requested bands (L1: 60–70%, L2: 40–50%); clamped if set via env.
try:
    PLACEMENT_APTITUDE_L1_PASS = max(60.0, min(70.0, float(os.environ.get('PLACEMENT_APTITUDE_L1_PASS', '65'))))
except ValueError:
    PLACEMENT_APTITUDE_L1_PASS = 65.0
try:
    PLACEMENT_APTITUDE_L2_PASS = max(40.0, min(50.0, float(os.environ.get('PLACEMENT_APTITUDE_L2_PASS', '45'))))
except ValueError:
    PLACEMENT_APTITUDE_L2_PASS = 45.0


def _placement_pass_threshold(module, level_for_pass):
    """Min % to pass placement L1 or L2 for a module (aptitude uses its own bars)."""
    module = (module or '').strip().lower()
    level_for_pass = (level_for_pass or 'l1').lower()
    if module == 'aptitude':
        return PLACEMENT_APTITUDE_L1_PASS if level_for_pass == 'l1' else PLACEMENT_APTITUDE_L2_PASS
    return PLACEMENT_L1_PASS if level_for_pass == 'l1' else PLACEMENT_L2_PASS


# Placement syllabus + generation style (L1 vs L2) — used in Groq prompts.
PLACEMENT_SYLLABUS_TOPICS = {
    'aptitude': {
        # Canonical lists — placement aptitude MUST stay within these topics only.
        'l1': [
            'Percentages',
            'Profit & Loss',
            'Simple Interest',
            'Ratio & Proportion',
            'Averages',
            'Time & Work (basic)',
            'Time, Speed & Distance',
            'Number System basics',
            'Simplification (BODMAS)',
            'Mixtures & Allegations (basic)',
            'Pipes & Cisterns (basic)',
            'Data Interpretation (tables, bar graphs)',
            'Basic Probability',
            'Basic Permutation & Combination',
            'Linear Equations',
        ],
        'l2': [
            'Advanced Percentages (successive change)',
            'Profit & Loss with multiple variables',
            'Compound Interest (CI vs SI, time-based traps)',
            'Time & Work with efficiency comparison',
            'Pipes & Cisterns (leak + fill cases)',
            'Advanced Ratio (multi-variable)',
            'Data Interpretation (caselets, missing data)',
            'Probability (conditional probability)',
            'Permutations & Combinations (arrangements)',
            'Number System (remainders, divisibility, cyclicity)',
            'Algebra word problems',
            'Quadratic Equations',
            'Logarithms (if included)',
            'Geometry (mensuration + tricky problems)',
        ],
    },
    'fundamentals': {
        'l1': [
            'OOP (Encapsulation, Inheritance, Polymorphism)', 'DBMS basics (ER model, normalization)',
            'SQL basic queries (SELECT, JOIN)', 'Operating Systems basics', 'Computer Networks basics (OSI model)',
            'Data Structures intro (array, stack, queue)', 'Basic Complexity (Big-O intro)',
        ],
        'l2': [
            'Advanced SQL (joins, indexing, optimization)', 'DBMS transactions & ACID', 'OS scheduling algorithms',
            'Deadlocks & concurrency', 'CN protocols (TCP/IP deep dive)', 'Advanced Data Structures (trees, graphs)',
            'Hashing techniques', 'System design basics', 'Memory management',
        ],
    },
    'coding': {
        'l1': [
            'Arrays', 'Strings', 'Linked List basics', 'Stack & Queue', 'Recursion',
            'Sorting & Searching', 'Hashing basics',
        ],
        'l2': [
            'Trees (DFS, BFS, traversals)', 'Binary Search Trees', 'Graphs (BFS, DFS, shortest path)',
            'Dynamic Programming', 'Greedy algorithms', 'Backtracking', 'Sliding Window (advanced)',
            'Two Pointers', 'Heap / Priority Queue', 'Trie', 'Segment Trees (high level)',
        ],
    },
    'basic_coding': {
        'l1': [
            'Fibonacci', 'Factorial', 'Prime check', 'Palindrome', 'Number reversal',
            'Basic patterns (stars, pyramids)', 'Sum of digits', 'GCD/LCM', 'Array basics (sum, max, min)', 'String basics',
        ],
        'l2': [
            'Pattern problems (complex)', 'Sliding window basics', 'Recursion problems',
            'Sorting (bubble, selection, insertion)', 'Searching (binary search)', 'Basic string manipulation',
            'Matrix problems', 'Bit manipulation basics',
        ],
    },
}

PLACEMENT_GD_TOPIC_POOL = {
    'l1': [
        'Digital literacy in rural schools',
        'Should college attendance be mandatory?',
        'Work from home: productivity vs collaboration',
        'Social media and mental health among students',
        'Climate action: individual vs government responsibility',
        'Uniform civil code: need for public debate?',
        'Is entrepreneurship overrated for fresh graduates?',
        'Role of sports in holistic education',
        'Poverty alleviation: cash transfers vs skill programs',
        'Online exams vs offline integrity',
    ],
    'l2': [
        'AI regulation: innovation vs safety',
        'Universal Basic Income in a developing economy',
        'Cryptocurrency: ban, regulate, or embrace?',
        'Net neutrality and corporate power',
        'Case: Startup layoffs vs employee rights',
        'Geoengineering to fight climate change',
        'Four-day work week for IT sector',
        'Data localization vs global cloud',
        'Meritocracy vs reservation: reframing the debate',
        'Autonomous weapons: ethical red lines',
    ],
}


def _placement_syllabus_shuffle(module_name, level, user_seed):
    level = (level or 'l1').lower()
    mod = (module_name or '').strip().lower()
    topics = (PLACEMENT_SYLLABUS_TOPICS.get(mod) or {}).get(level) or (PLACEMENT_SYLLABUS_TOPICS.get(mod) or {}).get('l1', [])
    if not topics:
        return []
    seed_int = int(hashlib.sha256(f"{user_seed}|{mod}|{level}".encode('utf-8')).hexdigest()[:14], 16)
    rng = random.Random(seed_int)
    out = topics[:]
    rng.shuffle(out)
    return out


def _placement_style_instruction(module_name, level):
    level = (level or 'l1').lower()
    mod = (module_name or '').strip().lower()
    if mod == 'aptitude':
        if level == 'l1':
            return (
                'Primary task style: "Generate practice questions for [topic] basic level with solutions" — '
                'each [topic] MUST be exactly one of the mandatory L1 aptitude strands listed in the prompt; '
                'no blood relations, no direction sense, no odd-one-out unless it clearly fits a listed strand.'
            )
        return (
            'Primary task style: "Generate advanced aptitude problems for [topic] with tricks and shortcuts" — '
            'each [topic] MUST be exactly one of the mandatory L2 aptitude strands listed in the prompt; '
            'no unrelated puzzle families.'
        )
    if mod == 'fundamentals':
        if level == 'l1':
            return (
                'Primary task style: "Explain [topic] with examples for beginners + MCQs" — '
                'each item is an MCQ testing recall/definition with a short worked solution.'
            )
        return (
            'Primary task style: "Generate interview-level questions for [topic] with explanations" — '
            'MCQs should feel like campus tech interviews (trade-offs, behavior of systems, SQL/DBMS depth).'
        )
    if mod == 'coding':
        if level == 'l1':
            return (
                'Primary task style: "Generate coding interview questions for [topic] easy level with explanation" — '
                'subjective stems: concise problem + ideal approach; no FAANG-only killer puzzles.'
            )
        return (
            'Primary task style: "Generate FAANG-level problems for [topic] with optimal solution" — '
            'subjective: harder multi-step algorithmic prompts with constraints and complexity discussion.'
        )
    if mod == 'basic_coding':
        if level == 'l1':
            return (
                'Primary task style: "Generate easy coding problems for [topic] with solution" — '
                'stdin/stdout coding; map each strand below to [topic].'
            )
        return (
            'Primary task style: "Generate medium-level coding problems focusing on logic for [topic]" — '
            'combine patterns, constraints, or two-step reasoning per problem.'
        )
    return ''


def _placement_quiz_display_topic(module, level):
    level = (level or 'l1').lower()
    mod = (module or '').strip().lower()
    labels = {
        ('aptitude', 'l1'): 'Aptitude L1 — Basic (formulas & direct logic)',
        ('aptitude', 'l2'): 'Aptitude L2 — Advanced (tricky & logical)',
        ('fundamentals', 'l1'): 'CS Fundamentals L1 — Core theory + MCQs',
        ('fundamentals', 'l2'): 'CS Fundamentals L2 — Interview depth',
        ('basic_coding', 'l1'): 'Basic Coding L1 — Beginner patterns',
        ('basic_coding', 'l2'): 'Basic Coding L2 — Logic building',
        ('coding', 'l1'): 'Technical Coding L1 — Foundation DSA (easy)',
        ('coding', 'l2'): 'Technical Coding L2 — Advanced interview / FAANG-style',
    }
    return labels.get((mod, level), 'Placement')


def _default_placement_levels():
    levels = {}
    for module in PLACEMENT_SEQUENCE:
        levels[module] = {
            'l1_score': 0.0,
            'l2_score': 0.0,
            'l1_passed': False,
            'l2_passed': False,
            'active_level': 'l1',
        }
    return levels


def _placement_question_signature(q):
    if not isinstance(q, dict):
        return ''
    parts = [str(q.get('question', '')).strip().lower()]
    opts = q.get('options')
    if isinstance(opts, list) and opts:
        parts.append('|'.join(str(o).strip().lower() for o in opts))
    lid = q.get('leetcode_id')
    if lid is not None and str(lid).strip() != '':
        parts.append(f"id:{lid}")
    blob = re.sub(r'\s+', ' ', ' '.join(parts)).strip()
    return hashlib.sha256(blob.encode('utf-8')).hexdigest() if blob else ''


def _placement_seen_db_signatures(user_id, module, level, limit=300):
    if not user_id:
        return []
    rows = (
        PlacementSeenQuestion.query.filter_by(user_id=user_id, module=module, level=level)
        .order_by(PlacementSeenQuestion.created_at.desc())
        .limit(limit)
        .all()
    )
    return [r.signature for r in rows if r.signature]


def _placement_global_signatures(module, level, limit=1500):
    """All aptitude questions ever issued (any user) for this module+level — prevents cross-user duplicates."""
    mod = (module or '').strip().lower()
    lvl = (level or 'l1').lower()
    if mod != 'aptitude':
        return []
    rows = (
        PlacementGlobalQuestion.query.filter_by(module=mod, level=lvl)
        .order_by(PlacementGlobalQuestion.id.desc())
        .limit(limit)
        .all()
    )
    return [r.signature for r in rows if r.signature]


def _get_recent_placement_signatures(state, module, level, limit=200):
    """Session + DB (+ global aptitude registry): durable exclusions so questions do not repeat."""
    mod = (module or '').strip().lower()
    lvl = (level or 'l1').lower()
    eff_limit = max(limit, 500) if mod == 'aptitude' else limit
    key = f"{module}:{level}"
    recent = (state.get('recent_questions', {}) or {}).get(key, [])
    if not isinstance(recent, list):
        recent = []
    session_sigs = [str(x) for x in recent if isinstance(x, str)]
    uid = getattr(current_user, 'id', None)
    db_sigs = _placement_seen_db_signatures(uid, module, level, limit=eff_limit)
    global_sigs = _placement_global_signatures(module, level, limit=max(1500, eff_limit * 4)) if mod == 'aptitude' else []
    merged = []
    seen = set()
    for s in global_sigs + db_sigs + session_sigs:
        if not s or s in seen:
            continue
        seen.add(s)
        merged.append(s)
        if len(merged) >= eff_limit:
            break
    return merged


def _remember_placement_signatures(state, module, level, questions, max_keep=120):
    key = f"{module}:{level}"
    recent_map = state.setdefault('recent_questions', {})
    existing = list(recent_map.get(key, []) or [])
    uid = getattr(current_user, 'id', None)
    new_for_db = []
    for q in questions:
        sig = _placement_question_signature(q)
        if sig and sig not in existing:
            existing.insert(0, sig)
            new_for_db.append(sig)
    recent_map[key] = existing[:max_keep]
    if uid and new_for_db:
        for sig in new_for_db:
            db.session.add(PlacementSeenQuestion(
                user_id=uid, module=module, level=level, signature=sig,
            ))
            try:
                db.session.commit()
            except IntegrityError:
                db.session.rollback()
    if (module or '').strip().lower() == 'aptitude' and new_for_db:
        mod = (module or '').strip().lower()
        lvl = (level or 'l1').lower()
        for sig in new_for_db:
            if not sig:
                continue
            db.session.add(PlacementGlobalQuestion(module=mod, level=lvl, signature=sig))
            try:
                db.session.commit()
            except IntegrityError:
                db.session.rollback()


def _placement_level_match(question_obj, level, module_name, relax_mcq_scenario=False):
    """Guardrail for L1 (recall/direct) vs L2 (applied/scenario). relax_mcq_scenario: static bank fallback."""
    if not isinstance(question_obj, dict):
        return False
    level = (level or 'l1').lower()
    module_name = (module_name or '').lower()
    qtxt = str(question_obj.get('question', '')).strip().lower()
    stxt = str(question_obj.get('solution', '')).strip().lower()
    atxt = str(question_obj.get('answer', '')).strip().lower()
    full = f"{qtxt} {stxt} {atxt}"
    tokens = len(re.findall(r'\w+', qtxt))
    qtype = str(question_obj.get('type', '')).strip().lower()

    advanced_markers = (
        'scenario', 'case study', 'optimize', 'optimization', 'trade-off', 'tradeoff',
        'edge case', 'edge-case', 'constraint', 'complexity', 'time complexity',
        'space complexity', 'design', 'justify', 'multi-step', 'multistep',
        'which approach', 'best strategy', 'what would you', 'how would you',
    )
    basic_markers = (
        'formula', 'calculate', 'find', 'simplify', 'identify', 'definition', 'what is',
        'compute', 'value of', 'next number', 'ratio', 'average', 'percentage',
        'hcf', 'lcm', 'gcd', 'select', 'choose the correct',
    )

    is_advanced_text = any(m in full for m in advanced_markers) or tokens >= 30
    is_basic_text = any(m in full for m in basic_markers) or tokens <= 20

    if module_name in ('coding', 'basic_coding'):
        difficulty = str(question_obj.get('difficulty', '')).strip().lower()
        if level == 'l1':
            if difficulty in ('hard', 'medium'):
                return False
            if difficulty == 'easy':
                return True
            return not is_advanced_text and tokens <= 48
        if difficulty == 'easy':
            return is_advanced_text and tokens >= 22
        return difficulty in ('medium', 'hard') or (is_advanced_text and tokens >= 24)

    is_mcq = qtype == 'mcq' or (qtype == '' and module_name in ('aptitude', 'fundamentals'))

    if is_mcq and module_name in ('aptitude', 'fundamentals'):
        if level == 'l1':
            if 'scenario' in qtxt:
                return False
            if tokens > 52 or len(qtxt) > 320:
                return False
            if is_advanced_text and not is_basic_text and tokens > 28:
                return False
            return True
        if relax_mcq_scenario:
            # Static banks rarely use "Scenario:"; use length + reasoning depth instead.
            return is_advanced_text or tokens >= 20 or len(qtxt) >= 120
        head = qtxt[:160].strip()
        has_scenario_prefix = head.lower().startswith('scenario:') or head.lower().startswith('scenario :')
        if not has_scenario_prefix:
            return False
        return tokens >= 18 and len(qtxt) >= 90

    if level == 'l1':
        if is_advanced_text and not is_basic_text and tokens > 26:
            return False
        if tokens > 60:
            return False
        return True
    return is_advanced_text or tokens >= 26


def _default_placement_state(user_id=None):
    return {
        'user_id': user_id,
        'current_stage': 'aptitude',
        'scores': {
            'aptitude': 0.0,
            'group_discussion': 0.0,
            'fundamentals': 0.0,
            'basic_coding': 0.0,
            'coding': 0.0
        },
        'completed': {
            'aptitude': False,
            'group_discussion': False,
            'fundamentals': False,
            'basic_coding': False,
            'coding': False
        },
        'levels': _default_placement_levels(),
        'recent_questions': {},
    }

def _get_placement_state():
    state = session.get('placement_track_state')
    expected_user_id = getattr(current_user, 'id', None)
    default_state = _default_placement_state(expected_user_id)
    if not state or not isinstance(state, dict):
        state = default_state
    # Critical isolation guard: if another account logs in on same browser,
    # do not reuse the previous account's placement session progress.
    if state.get('user_id') != expected_user_id:
        state = default_state

    state.setdefault('user_id', expected_user_id)
    state.setdefault('current_stage', 'aptitude')
    state.setdefault('scores', default_state['scores'])
    state.setdefault('completed', default_state['completed'])
    state.setdefault('levels', _default_placement_levels())
    state.setdefault('recent_questions', {})
    for key in PLACEMENT_SEQUENCE:
        state['scores'].setdefault(key, 0.0)
        state['completed'].setdefault(key, False)
        lv = state['levels'].setdefault(key, {})
        lv.setdefault('l1_score', 0.0)
        lv.setdefault('l2_score', 0.0)
        lv.setdefault('l1_passed', False)
        lv.setdefault('l2_passed', False)
        lv.setdefault('active_level', 'l1')
        state['completed'][key] = bool(lv.get('l1_passed') and lv.get('l2_passed'))
        state['scores'][key] = round((float(lv.get('l1_score', 0.0)) + float(lv.get('l2_score', 0.0))) / 2.0, 1)
    return state

def _save_placement_state(state):
    session['placement_track_state'] = state
    session.modified = True

def _language_constraints():
    return ['python', 'java', 'cpp', 'c']

def _starter_code_templates():
    return {
        'python': "def solve():\n    # Write your solution\n    pass\n\nif __name__ == '__main__':\n    solve()\n",
        'java': "import java.io.*;\nimport java.util.*;\n\npublic class Main {\n    public static void main(String[] args) throws Exception {\n        // Write your solution here\n    }\n}\n",
        'cpp': "#include <bits/stdc++.h>\nusing namespace std;\n\nint main() {\n    ios::sync_with_stdio(false);\n    cin.tie(nullptr);\n    // Write your solution here\n    return 0;\n}\n",
        'c': "#include <stdio.h>\n\nint main() {\n    // Write your solution here\n    return 0;\n}\n",
    }

def _normalize_coding_question(q):
    normalized = dict(q)
    normalized['type'] = 'coding'
    normalized['marks'] = int(normalized.get('marks', 10) or 10)
    normalized['time_limit_seconds'] = int(normalized.get('time_limit_seconds', 2) or 2)
    normalized['memory_limit_mb'] = int(normalized.get('memory_limit_mb', 256) or 256)
    normalized['language_constraints'] = normalized.get('language_constraints') or _language_constraints()
    normalized['starter_code'] = normalized.get('starter_code') or _starter_code_templates()
    def _to_bool(val):
        if isinstance(val, bool):
            return val
        if isinstance(val, (int, float)):
            return val != 0
        if isinstance(val, str):
            return val.strip().lower() in ('1', 'true', 'yes', 'y')
        return False

    raw_cases = normalized.get('test_cases') or []
    test_cases = []
    for tc in raw_cases:
        if not isinstance(tc, dict):
            continue
        test_cases.append({
            "input": str(tc.get('input', '')),
            "expected_output": str(tc.get('expected_output', '')),
            "is_hidden": _to_bool(tc.get('is_hidden', False))
        })

    visible_cases = [tc for tc in test_cases if not tc.get('is_hidden', False)]
    hidden_cases = [tc for tc in test_cases if tc.get('is_hidden', False)]
    if not visible_cases and normalized.get('sample_input') is not None and normalized.get('sample_output') is not None:
        visible_cases = [{
            "input": str(normalized.get('sample_input', '')),
            "expected_output": str(normalized.get('sample_output', '')),
            "is_hidden": False
        }]
    if not visible_cases:
        visible_cases = [{"input": "1\n1", "expected_output": "1", "is_hidden": False}]
    if len(hidden_cases) < 2:
        seed_input = str(visible_cases[0].get('input', '1\n1'))
        seed_output = str(visible_cases[0].get('expected_output', '1'))
        while len(hidden_cases) < 2:
            hidden_cases.append({"input": seed_input, "expected_output": seed_output, "is_hidden": True})
    normalized['test_cases'] = [
        {**visible_cases[0], "is_hidden": False},
        {**hidden_cases[0], "is_hidden": True},
        {**hidden_cases[1], "is_hidden": True}
    ]
    return normalized

def _pick_leetcode_mix(excluded_signatures=None, prefer=('easy', 'medium', 'hard'), max_count=5, rng_seed=None):
    """Shuffle the full preferred pool so users with long exclusion history still get fresh picks."""
    excluded_signatures = set(excluded_signatures or [])
    prefer_set = set(prefer)
    pool = [q for q in PLACEMENT_LEETCODE_BANK if q.get('difficulty') in prefer_set]
    if not pool:
        pool = list(PLACEMENT_LEETCODE_BANK)
    if rng_seed:
        rng = random.Random(int(hashlib.sha256(str(rng_seed).encode('utf-8')).hexdigest()[:14], 16))
        rng.shuffle(pool)
    else:
        random.shuffle(pool)
    out = []
    for q in pool:
        nq = _normalize_coding_question(q)
        sig = _placement_question_signature(nq)
        if sig and sig in excluded_signatures:
            continue
        out.append(nq)
        if len(out) >= max_count:
            break
    return out

def _pick_practice_recommendations(count=10):
    easy = [q for q in LEETCODE_PRACTICE_POOL if q.get('difficulty') == 'easy']
    medium = [q for q in LEETCODE_PRACTICE_POOL if q.get('difficulty') == 'medium']
    hard = [q for q in LEETCODE_PRACTICE_POOL if q.get('difficulty') == 'hard']
    picked = (
        random.sample(easy, min(3, len(easy))) +
        random.sample(medium, min(5, len(medium))) +
        random.sample(hard, min(2, len(hard)))
    )
    random.shuffle(picked)
    return picked[:count]

def _generate_basic_coding_questions_groq(num_questions=5, level='l1', user_seed=None, excluded_signatures=None, user_id=None, attempt_nonce=None):
    excluded_signatures = excluded_signatures or []
    seed_line = user_seed or f"seed-{random.randint(1000, 9999)}"
    level = (level or 'l1').lower()
    ask_count = min(num_questions + 5, 12)
    diff_line = '"difficulty": "easy"' if level == 'l1' else '"difficulty": "medium"'
    level_guidance = (
        "L1 strict: Bloom 1-2 only. Single-pass I/O, loops, or simple math on arrays/strings. "
        "No asymptotic analysis, no optimization proof, no tricky adversarial edge cases."
        if level == 'l1' else
        "L2 strict: Bloom 3-4 only. Each problem must state explicit constraints (bounds, formats) "
        "and require combining 2+ ideas (e.g. prefix + hashmap, two pointers with invariant)."
    )
    strands = _placement_syllabus_shuffle('basic_coding', level, seed_line)
    strand_lines = ''
    if strands:
        strand_lines = (
            '\nRotate across these strands (map each to [topic] in the style instruction):\n- '
            + '\n- '.join(strands[:12])
        )
    style_inst = _placement_style_instruction('basic_coding', level)
    uniq = ''
    if user_id is not None or attempt_nonce:
        uniq = (
            f"\nUNIQUE_RUN_ID: user-{user_id or 0}-nonce-{attempt_nonce or 'na'}-t{int(time.time() * 1000) % 1_000_000_000}\n"
            "Do not reuse boilerplate stems; vary variable names, bounds, and story hooks.\n"
        )
    prompt = f"""
Generate exactly {ask_count} distinct coding interview questions.
{uniq}{style_inst}
{strand_lines}
{level_guidance}
Randomization seed: {seed_line}
Avoid repeating questions whose fingerprints match these excluded SHA-256 hashes (question+options/id):
{json.dumps(excluded_signatures[:200])}

Return STRICT JSON array only. Each item must be:
{{
  "question": "clear problem statement",
  "type": "coding",
  "marks": 10,
  {diff_line},
  "sample_input": "input text",
  "sample_output": "output text",
  "solution": "short approach",
  "test_cases": [
    {{"input": "...", "expected_output": "...", "is_hidden": false}},
    {{"input": "...", "expected_output": "...", "is_hidden": true}},
    {{"input": "...", "expected_output": "...", "is_hidden": true}}
  ]
}}

Rules:
- Ensure each question is executable in C/C++/Java/Python with stdin/stdout.
- Exactly 1 visible and 2 hidden test cases.
- Keep language constraints to python/java/cpp/c by default.
"""
    last_error = None
    temp = 0.55 if level == 'l1' else 0.78
    for _ in range(4):
        try:
            payload = _placement_chat_completion(
                messages=[
                    {"role": "system", "content": "You are a coding question generator. Output strictly valid JSON array only."},
                    {"role": "user", "content": prompt},
                ],
                temperature=temp,
                max_tokens=3200,
                timeout=50,
            )
        except PlacementRateLimitError as exc:
            wait_s = min(20.0, (exc.retry_after or 2.0) + random.uniform(0.1, 0.7))
            time.sleep(wait_s)
            last_error = Exception('OpenRouter/Groq rate-limited while generating basic coding questions.')
            continue
        except Exception as exc:
            last_error = exc
            continue

        raw = payload.get('content', '').strip()
        cleaned = raw.replace('```json', '').replace('```', '').strip()
        try:
            parsed = json.loads(cleaned)
        except json.JSONDecodeError as exc:
            last_error = Exception(f'Invalid JSON from coding question generator ({exc}).')
            continue
        if not isinstance(parsed, list) or not parsed:
            last_error = Exception('Invalid basic coding response format from Groq.')
            continue
        unique = []
        seen = set(excluded_signatures)
        for q in parsed:
            if not isinstance(q, dict):
                continue
            nq = _normalize_coding_question(q)
            if level == 'l1':
                nq['difficulty'] = 'easy'
            else:
                nq['difficulty'] = 'medium'
            sig = _placement_question_signature(nq)
            if sig and sig in seen:
                continue
            if not _placement_level_match(nq, level, 'basic_coding'):
                continue
            if sig:
                seen.add(sig)
            unique.append(nq)
        if len(unique) >= num_questions:
            return unique[:num_questions]
        last_error = Exception('Generated coding set did not satisfy strict level constraints.')
    raise last_error or Exception('Could not generate level-aligned basic coding questions.')

def _generate_placement_questions_groq(topic, module_name, num_questions=10, level='l1', user_seed=None, excluded_signatures=None, max_runtime_sec=None, user_id=None, attempt_nonce=None):
    module_prompt = "placement aptitude"
    if module_name == 'fundamentals':
        module_prompt = "CS fundamentals: OOPs, Computer Networks, DBMS, SQL, DSA basics"
    elif module_name == 'coding':
        module_prompt = "technical coding interview problems (algorithmic thinking, clean logic)"
    elif module_name == 'aptitude':
        module_prompt = (
            'placement quantitative aptitude — MCQs must map strictly to the mandatory L1/L2 strand list in the prompt'
        )

    if module_name in ('fundamentals', 'aptitude'):
        schema_hint = """
Return strict JSON array with each item:
{
  "question": "...",
  "type": "mcq",
  "options": ["A. ...","B. ...","C. ...","D. ..."],
  "answer": "A/B/C/D",
  "solution": "short stepwise explanation"
}
"""
    else:
        schema_hint = """
Return strict JSON array with each item:
{
  "question": "...",
  "type": "subjective",
  "marks": 10,
  "answer": "ideal approach/solution",
  "solution": "stepwise explanation and key logic"
}
"""

    excluded_signatures = excluded_signatures or []
    level = (level or 'l1').lower()
    module_name = (module_name or '').strip().lower()
    seed_line = user_seed or f"seed-{random.randint(1000, 9999)}"
    strands = _placement_syllabus_shuffle(module_name, level, seed_line)
    strand_lines = ''
    if strands:
        strand_lines = (
            '\nSyllabus strands for THIS attempt (distribute questions across them; prefer unused strands first):\n- '
            + '\n- '.join(strands[:16])
        )
    style_inst = _placement_style_instruction(module_name, level)
    aptitude_topic_lock = ''
    if module_name == 'aptitude':
        canon = PLACEMENT_SYLLABUS_TOPICS['aptitude'].get(level) or PLACEMENT_SYLLABUS_TOPICS['aptitude']['l1']
        aptitude_topic_lock = (
            '\nMANDATORY APTITUDE STRANDS (this level only — every question MUST clearly fit exactly one line below; '
            'do not use any other topic family):\n'
            + '\n'.join(f'- {s}' for s in canon)
            + '\nWhen generating multiple items, prioritize distinct strands before repeating any strand.\n'
            'GLOBAL UNIQUENESS: excluded fingerprints include questions already assigned to ANY user; '
            'you must not reproduce or trivially paraphrase those stems.'
        )
    if module_name in ('aptitude', 'fundamentals'):
        level_guidance = (
            "L1 Bloom 1-2: single-step recall, short numeric/logic stems (aim under 45 words). "
            "Never use the word 'scenario' (any case). No workplace stories."
            if level == 'l1' else
            "L2 Bloom 3-4: applied judgment. For EVERY item, the question string MUST begin with "
            "the exact prefix \"Scenario: \" then at most TWO short sentences of context, then one clear MCQ line. "
            "CRITICAL: Keep each option under 18 words and solutions brief so the JSON stays valid and complete."
        )
    else:
        level_guidance = (
            "L1 Bloom 1-2: compact stems, definition/small-step reasoning, no heavy optimization proofs."
            if level == 'l1' else
            "L2 Bloom 3-4: multi-step reasoning, explicit constraints, trade-offs, and algorithmic nuance."
        )

    is_mcq_module = module_name in ('aptitude', 'fundamentals')
    long_mcq_l2 = is_mcq_module and level == 'l2'
    # Small batches + accumulation: avoids truncated JSON and yields enough items after level/dedupe filters.
    batch_cap = 6 if long_mcq_l2 else (7 if is_mcq_module else 6)
    overshoot = 2 if long_mcq_l2 else 3
    collected = []
    sig_seen = set(excluded_signatures)
    last_error = None
    temp = 0.52 if level == 'l1' else 0.82
    if module_name == 'aptitude':
        temp = 0.58 if level == 'l1' else 0.86
    if max_runtime_sec is None:
        # Keep aptitude L1 snappy; fall back quickly instead of hanging.
        if module_name == 'aptitude' and level == 'l1':
            max_runtime_sec = float(os.environ.get('PLACEMENT_APT_L1_MAX_RUNTIME_SEC', '20'))
        else:
            max_runtime_sec = float(os.environ.get('PLACEMENT_GEN_MAX_RUNTIME_SEC', '40'))
    started_at = time.time()
    outer_cap = 25
    batch_idx = 0

    while len(collected) < num_questions and batch_idx < outer_cap:
        if (time.time() - started_at) > max_runtime_sec:
            last_error = Exception('Placement question generation timed out; using fallback where available.')
            break
        batch_idx += 1
        need = num_questions - len(collected)
        ask_count = min(need + overshoot, batch_cap)
        ask_count = max(ask_count, min(4, need + 1))
        batch_seed = f"{seed_line}-b{batch_idx}"

        uniq = ''
        if user_id is not None or attempt_nonce:
            uniq = (
                f"\nUNIQUE_RUN_ID: user-{user_id or 0}-nonce-{attempt_nonce or 'na'}-batch-{batch_idx}-"
                f"t{int(time.time() * 1000) % 1_000_000_000}\n"
                "Anti-repeat: vary all numeric constants, names, and contexts vs any default template; "
                "do not copy famous textbook/diagnostic stems verbatim.\n"
            )
        prompt = f"""
Assessment theme: {topic}
{uniq}{style_inst}
{aptitude_topic_lock}
{strand_lines}

Generate exactly {ask_count} high-quality {module_prompt} questions for placement preparation.
{level_guidance}
Randomization seed: {batch_seed}
Avoid repeated/ambiguous questions.
Do not repeat items matching these excluded SHA-256 fingerprints (question text + options):
{json.dumps(list(sig_seen)[:200])}

{schema_hint}
Return JSON only, no markdown fences. Close all strings and brackets; output must be complete valid JSON.
"""

        if is_mcq_module:
            max_tokens = 8192
        else:
            max_tokens = min(8192, 1200 + ask_count * 400)

        inner_ok = False
        max_groq_attempts = 10
        for groq_attempt in range(max_groq_attempts):
            if (time.time() - started_at) > max_runtime_sec:
                last_error = Exception('Placement question generation timed out; using fallback where available.')
                break
            try:
                call_timeout = 30 if (module_name == 'aptitude' and level == 'l1') else 70
                payload = _placement_chat_completion(
                    messages=[
                        {
                            "role": "system",
                            "content": (
                                "You are an expert placement test setter. Output strictly valid JSON: "
                                "one array only, properly closed, no trailing commas."
                            ),
                        },
                        {"role": "user", "content": prompt},
                    ],
                    temperature=temp,
                    max_tokens=max_tokens,
                    timeout=call_timeout,
                )
            except PlacementRateLimitError as exc:
                last_error = Exception('OpenRouter/Groq rate-limited. Backing off and retrying.')
                wait_s = min(25.0, (exc.retry_after or (2.0 ** min(groq_attempt, 5))) + random.uniform(0.1, 0.7))
                time.sleep(wait_s)
                continue
            except Exception as exc:
                last_error = exc
                continue

            finish = (payload.get('finish_reason') or '') or ''
            raw = payload.get('content', '').strip()
            cleaned = raw.replace('```json', '').replace('```', '').strip()
            try:
                parsed = json.loads(cleaned)
            except json.JSONDecodeError as exc:
                last_error = Exception(f'Invalid JSON from question generator ({exc}).')
                if finish == 'length':
                    last_error = Exception('Model output was truncated; retrying with a smaller batch.')
                continue
            if not isinstance(parsed, list) or not parsed:
                last_error = Exception('Invalid response format from Groq question generator.')
                continue

            added_this = 0
            for q in parsed:
                if len(collected) >= num_questions:
                    break
                if not isinstance(q, dict):
                    continue
                sig = _placement_question_signature(q)
                if sig and sig in sig_seen:
                    continue
                if not _placement_level_match(q, level, module_name):
                    continue
                if sig:
                    sig_seen.add(sig)
                collected.append(q)
                added_this += 1
            inner_ok = True
            if len(collected) >= num_questions:
                return collected[:num_questions]
            if added_this == 0:
                last_error = Exception('Generated set did not satisfy strict level constraints.')
            break

        if not inner_ok:
            continue
        # Space out multi-batch placement calls to stay under Groq RPM/TPM (esp. fundamentals L2).
        if is_mcq_module and len(collected) < num_questions:
            try:
                batch_delay = float(os.environ.get('PLACEMENT_GROQ_BATCH_DELAY_SEC', '1.4'))
            except ValueError:
                batch_delay = 1.4
            time.sleep(max(0.0, batch_delay))

    if len(collected) >= num_questions:
        return collected[:num_questions]
    raise last_error or Exception('Could not generate level-aligned placement questions.')

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

def _hash_otp(value):
    return hashlib.sha256(f"{value}|{app.config['SECRET_KEY']}".encode('utf-8')).hexdigest()

def _generate_otp():
    return f"{secrets.randbelow(1000000):06d}"

def _send_email_otp(email, otp):
    subject = "UniTest verification code"
    body = f"Your UniTest email OTP is: {otp}. It is valid for 10 minutes."
    mail_server = os.environ.get('MAIL_SERVER')
    mail_port = int(os.environ.get('MAIL_PORT', '587'))
    mail_username = os.environ.get('MAIL_USERNAME')
    mail_password = os.environ.get('MAIL_PASSWORD')
    mail_from = (os.environ.get('MAIL_FROM') or mail_username or '').strip()
    mail_use_tls = str(os.environ.get('MAIL_USE_TLS', 'true')).lower() == 'true'

    # Dev fallback when SMTP is not configured.
    if not (mail_server and mail_username and mail_password):
        print(f"[DEV EMAIL OTP] {email} => {otp}")
        return True

    try:
        msg = MIMEMultipart()
        msg['From'] = mail_from or mail_username
        msg['To'] = email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))

        with smtplib.SMTP(mail_server, mail_port, timeout=15) as server:
            if mail_use_tls:
                server.starttls()
            server.login(mail_username, mail_password)
            server.sendmail(mail_username, [email], msg.as_string())
        return True
    except Exception as e:
        print(f"Email OTP send failed: {e}")
        return False

def _send_password_reset_email(email, reset_link):
    subject = "UniTest password reset link"
    body = (
        "You requested a password reset for your UniTest account.\n\n"
        f"Reset link: {reset_link}\n\n"
        "This link is valid for 1 hour. If you did not request this, you can ignore this email."
    )
    mail_server = os.environ.get('MAIL_SERVER')
    mail_port = int(os.environ.get('MAIL_PORT', '587'))
    mail_username = os.environ.get('MAIL_USERNAME')
    mail_password = os.environ.get('MAIL_PASSWORD')
    mail_from = (os.environ.get('MAIL_FROM') or mail_username or '').strip()
    mail_use_tls = str(os.environ.get('MAIL_USE_TLS', 'true')).lower() == 'true'

    if not (mail_server and mail_username and mail_password):
        print(f"[DEV RESET EMAIL] {email} => {reset_link}")
        return False

    try:
        msg = MIMEMultipart()
        msg['From'] = mail_from or mail_username
        msg['To'] = email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))

        with smtplib.SMTP(mail_server, mail_port, timeout=15) as server:
            if mail_use_tls:
                server.starttls()
            server.login(mail_username, mail_password)
            server.sendmail(mail_username, [email], msg.as_string())
        return True
    except Exception as e:
        print(f"Password reset email send failed: {e}")
        return False

def _make_unique_username(base_username):
    base = re.sub(r'[^a-zA-Z0-9_]', '', (base_username or 'user').strip())[:24] or 'user'
    candidate = base
    idx = 1
    while db.session.query(User).filter_by(username=candidate).first():
        candidate = f"{base}{idx}"
        idx += 1
    return candidate

def _record_login_event(user):
    login_time = datetime.utcnow()
    user.last_login = login_time
    ip_address = request.headers.get('X-Forwarded-For', request.remote_addr)
    if ip_address and ',' in ip_address:
        ip_address = ip_address.split(',')[0].strip()
    geo_data = get_geolocation_from_ip(ip_address)
    login_history = LoginHistory(
        user_id=user.id,
        login_time=login_time,
        ip_address=ip_address,
        user_agent=request.headers.get('User-Agent', 'Unknown'),
        latitude=geo_data.get('latitude') if geo_data else None,
        longitude=geo_data.get('longitude') if geo_data else None,
        city=geo_data.get('city') if geo_data else None,
        country=geo_data.get('country') if geo_data else None,
        region=geo_data.get('region') if geo_data else None
    )
    db.session.add(login_history)

def _record_site_activity():
    # Log only meaningful page visits (GET HTML pages), skip assets/internal calls.
    if request.method != 'GET':
        return
    if request.path.startswith('/static'):
        return
    if request.endpoint in {
        'static', 'admin_metrics', 'favicon', 'internal_error', 'not_found_error'
    }:
        return
    if request.path.startswith('/api'):
        return
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return

    ip_address = request.headers.get('X-Forwarded-For', request.remote_addr)
    if ip_address and ',' in ip_address:
        ip_address = ip_address.split(',')[0].strip()

    activity = SiteActivity(
        user_id=current_user.id if current_user.is_authenticated else None,
        path=request.path[:255],
        endpoint=(request.endpoint or '')[:120] if request.endpoint else None,
        method=request.method,
        ip_address=ip_address,
        user_agent=(request.headers.get('User-Agent') or 'Unknown')[:255]
    )
    db.session.add(activity)
    db.session.commit()

def handle_gemini_api_error(e, context="API call"):
    """Handle Gemini API errors and provide user-friendly messages"""
    error_str = str(e).lower()
    error_msg = str(e)
    
    # Check for quota exceeded
    if 'quota' in error_str or '429' in error_msg:
        print(f"⚠️ GEMINI API QUOTA EXCEEDED in {context}")
        print(f"   Error: {error_msg}")
        print("   → Check Google Cloud Console for usage: https://console.cloud.google.com/")
        print("   → Quota resets at midnight UTC")
        return {
            'error': 'quota_exceeded',
            'message': 'API quota exceeded. Please try again later or check your usage limits.',
            'user_message': 'The AI service is temporarily unavailable due to quota limits. Please try again in a few minutes.'
        }
    
    # Check for rate limit
    elif 'rate limit' in error_str or 'rate_limit' in error_str:
        print(f"⚠️ GEMINI API RATE LIMIT in {context}")
        print(f"   Error: {error_msg}")
        print("   → Too many requests per minute. Please wait and retry.")
        return {
            'error': 'rate_limit',
            'message': 'Rate limit exceeded. Please wait a moment and try again.',
            'user_message': 'Too many requests. Please wait a moment and try again.'
        }
    
    # Check for invalid API key
    elif 'api key' in error_str or 'invalid' in error_str or '401' in error_msg or '403' in error_msg:
        print(f"❌ GEMINI API AUTHENTICATION ERROR in {context}")
        print(f"   Error: {error_msg}")
        print("   → Check your GOOGLE_AI_API_KEY environment variable")
        return {
            'error': 'auth_error',
            'message': 'API authentication failed. Please check your API key configuration.',
            'user_message': 'API configuration error. Please contact support.'
        }
    
    # Check for permission denied
    elif 'permission' in error_str or 'forbidden' in error_str:
        print(f"❌ GEMINI API PERMISSION DENIED in {context}")
        print(f"   Error: {error_msg}")
        print("   → Enable 'Generative Language API' in Google Cloud Console")
        return {
            'error': 'permission_denied',
            'message': 'API permission denied. Please enable Generative Language API.',
            'user_message': 'API service not available. Please contact support.'
        }
    
    # Generic error
    else:
        print(f"❌ GEMINI API ERROR in {context}: {error_msg}")
        return {
            'error': 'api_error',
            'message': f'API error: {error_msg}',
            'user_message': 'An error occurred while processing your request. Please try again.'
        }

def evaluate_subjective_answer(question, student_answer, model_answer):
    """Use AI to evaluate subjective answers"""
    if not genai or not student_answer.strip():
        return 0.0

    try:
        # Use gemini-2.5-flash (current free tier) with fallback to gemini-2.5-flash-lite
        try:
            model = genai.GenerativeModel("gemini-2.5-flash")
        except:
            try:
                model = genai.GenerativeModel("gemini-2.5-flash-lite")
            except:
                try:
                    model = genai.GenerativeModel("gemini-1.5-flash")
                except:
                    raise Exception("No working Gemini model found. Check API key and quota.")
        prompt = f"""
        Evaluate this student's answer for the given question:

        Question: {question}
        Student Answer: {student_answer}
        Model Answer: {model_answer}

        Rate the student's answer on a scale of 0.0 to 1.0 based on:
        - Accuracy and correctness
        - Completeness
        - Understanding demonstrated
        - Relevance to the question

        Return only a number between 0.0 and 1.0 (e.g., 0.8 for 80% correct)
        """

        response = model.generate_content(prompt)
        score_text = response.text.strip()

        # Extract number from response
        score_match = re.search(r'(\d*\.?\d+)', score_text)
        if score_match:
            score = float(score_match.group(1))
            return min(max(score, 0.0), 1.0)  # Clamp between 0 and 1

        return 0.5  # Default if can't parse
    except Exception as e:
        error_info = handle_gemini_api_error(e, "evaluate_subjective_answer")
        print(f"Error in evaluate_subjective_answer: {error_info['message']}")
        return 0.5  # Default on error

def execute_code(code, language, test_input, time_limit=2, memory_limit=256):
    """Execute code using Piston API (free, no API key needed)"""
    piston_url = "https://emkc.org/api/v2/piston/execute"
    
    piston_languages = {
        'python': 'python',
        'python3': 'python',
        'java': 'java',
        'cpp': 'cpp',
        'c': 'c'
    }
    
    piston_lang = piston_languages.get(language.lower(), 'python3')
    
    try:
        payload = {
            "language": piston_lang,
            "version": "*",
            "files": [{"content": code}],
            "stdin": test_input,
            "args": [],
            "compile_timeout": 10000,
            "run_timeout": time_limit * 1000,
            "compile_memory_limit": memory_limit * 1024 * 1024,
            "run_memory_limit": memory_limit * 1024 * 1024
        }
        
        response = requests.post(piston_url, json=payload, timeout=15)
        
        if response.status_code == 200:
            try:
                result = response.json()
                if not result:
                    return {
                        'status': 'error',
                        'message': 'Empty response from execution API',
                        'output': '',
                        'stderr': ''
                    }
                
                if result.get('run'):
                    run_result = result['run']
                    if not run_result:
                        return {
                            'status': 'error',
                            'message': 'Invalid run result from API',
                            'output': '',
                            'stderr': ''
                        }
                    
                    if run_result.get('code') == 0:
                        return {
                            'status': 'success',
                            'output': run_result.get('stdout', '').strip(),
                            'stderr': run_result.get('stderr', '').strip()
                        }
                    else:
                        return {
                            'status': 'error',
                            'message': 'Runtime Error',
                            'output': run_result.get('stdout', '').strip(),
                            'stderr': run_result.get('stderr', '').strip() or run_result.get('stdout', '')
                        }
                else:
                    # No 'run' key in response - return error
                    return {
                        'status': 'error',
                        'message': 'Unexpected API response format',
                        'output': '',
                        'stderr': str(result) if result else 'No response data'
                    }
            except (ValueError, KeyError) as e:
                return {
                    'status': 'error',
                    'message': f'Failed to parse API response: {str(e)}',
                    'output': '',
                    'stderr': ''
                }
        else:
            return {
                'status': 'error',
                'message': f'API returned status {response.status_code}',
                'output': '',
                'stderr': response.text[:200] if hasattr(response, 'text') else ''
            }
    except requests.exceptions.RequestException as e:
        return {
            'status': 'error',
            'message': f'Network error: {str(e)}',
            'output': '',
            'stderr': ''
        }
    except Exception as e:
        return {
            'status': 'error',
            'message': f'Execution error: {str(e)}',
            'output': '',
            'stderr': ''
        }

def run_test_cases(code, language, test_cases, time_limit=2, memory_limit=256):
    """Run multiple test cases and return results"""
    results = []
    passed = 0
    executor_unavailable = False
    executor_message = ''
    
    for test_case in test_cases:
        test_input = test_case.get('input', '') if isinstance(test_case, dict) else ''
        expected_output = test_case.get('expected_output', '').strip() if isinstance(test_case, dict) else ''
        is_hidden = test_case.get('is_hidden', False) if isinstance(test_case, dict) else False
        
        exec_result = execute_code(code, language, test_input, time_limit, memory_limit)
        
        # Ensure exec_result is not None and is a dict
        if not exec_result or not isinstance(exec_result, dict):
            exec_result = {
                'status': 'error',
                'message': 'Execution returned invalid result',
                'output': '',
                'stderr': ''
            }
        
        error_blob = (
            f"{exec_result.get('message', '')} {exec_result.get('stderr', '')} {exec_result.get('output', '')}"
        ).lower()
        if (
            'public piston api is now whitelist only' in error_blob or
            'whitelist only' in error_blob or
            'contact engineermon on discord' in error_blob
        ):
            executor_unavailable = True
            executor_message = (
                'Code execution service is temporarily unavailable (provider whitelist mode). '
                'Please retry later or deploy with your own runner key.'
            )
            break

        if exec_result.get('status') == 'success':
            actual_output = exec_result.get('output', '').strip()
            is_correct = actual_output == expected_output
            if is_correct:
                passed += 1
        else:
            actual_output = exec_result.get('stderr', '') or exec_result.get('message', 'Error')
            is_correct = False
        
        results.append({
            'input': test_input,
            'expected_output': expected_output,
            'actual_output': actual_output,
            'is_correct': is_correct,
            'is_hidden': is_hidden
        })
    
    total = len(test_cases)
    percentage = (passed / total * 100) if total > 0 else 0
    
    return {
        'results': results,
        'passed': passed,
        'total': total,
        'percentage': percentage,
        'executor_unavailable': executor_unavailable,
        'executor_message': executor_message
    }

def evaluate_coding_answer_ai(question, user_code, language, sample_input='', sample_output='', solution=''):
    """Fallback AI evaluator for coding answers when runtime service is unavailable."""
    groq_key = os.environ.get('GROQ_API_KEY')
    if not groq_key or not (user_code or '').strip():
        return 0.7, "AI fallback used due to runner outage. Baseline credit applied."

    model = os.environ.get('GROQ_MODEL', 'llama-3.3-70b-versatile')
    prompt = f"""
You are evaluating a student's coding interview answer.
Return ONLY strict JSON with keys:
{{
  "score": <float between 0 and 1>,
  "feedback": "<1-2 line concise feedback>"
}}

Question:
{question}

Language:
{language}

Student code:
{user_code}

Sample input:
{sample_input}

Expected sample output:
{sample_output}

Reference approach:
{solution}

Scoring rubric:
- Correctness and logical validity: 50%
- Handling edge cases: 20%
- Time/space complexity quality: 20%
- Code clarity/structure: 10%
"""
    try:
        response = requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {groq_key}",
                "Content-Type": "application/json"
            },
            json={
                "model": model,
                "temperature": 0.1,
                "max_tokens": 220,
                "messages": [
                    {"role": "system", "content": "You are a strict coding evaluator. Output valid JSON only."},
                    {"role": "user", "content": prompt}
                ]
            },
            timeout=25
        )
        if response.status_code >= 400:
            return 0.7, "AI fallback partially unavailable; baseline credit applied."

        raw = (
            response.json()
            .get('choices', [{}])[0]
            .get('message', {})
            .get('content', '')
            .strip()
        )
        cleaned = raw.replace('```json', '').replace('```', '').strip()
        parsed = json.loads(cleaned)
        score = float(parsed.get('score', 0.7))
        feedback = str(parsed.get('feedback', 'AI fallback used for evaluation.')).strip()
        score = min(max(score, 0.0), 1.0)
        return score, feedback[:400]
    except Exception:
        return 0.7, "AI fallback evaluation failed; baseline credit applied."

def get_difficulty_from_bloom_level(bloom_level):
    """Map Bloom's taxonomy level to difficulty level"""
    if bloom_level <= 2:
        return "beginner"
    elif bloom_level <= 4:
        return "intermediate"
    else:
        return "difficult"

def extract_pdf_content_with_ocr(file_path):
    """Extract text from PDF using OCR (for scanned/image-based PDFs)"""
    if not OCR_AVAILABLE:
        return None
    
    try:
        import tempfile
        import os
        
        # Convert PDF pages to images
        # Note: Requires Poppler to be installed (pdf2image dependency)
        try:
            images = convert_from_path(file_path, dpi=300)
            print(f"Converted PDF to {len(images)} images for OCR processing")
        except Exception as convert_error:
            print(f"Error converting PDF to images (Poppler may not be installed): {convert_error}")
            print("Install Poppler: Windows - download from poppler.freedesktop.org, Linux - sudo apt-get install poppler-utils, Mac - brew install poppler")
            return None
        
        ocr_text = ""
        for i, image in enumerate(images):
            try:
                # Perform OCR on each page
                page_text = pytesseract.image_to_string(image, lang='eng')
                if page_text.strip():
                    ocr_text += f"\n--- Page {i+1} ---\n{page_text}\n"
                    print(f"OCR extracted text from page {i+1} ({len(page_text)} characters)")
            except Exception as ocr_error:
                print(f"Error performing OCR on page {i+1}: {ocr_error}")
                # Check if Tesseract is installed
                if "tesseract" in str(ocr_error).lower() or "not found" in str(ocr_error).lower():
                    print("Tesseract OCR engine not found. Please install Tesseract.")
                continue
        
        return ocr_text.strip() if ocr_text.strip() else None
    except Exception as e:
        print(f"Error in OCR processing: {e}")
        import traceback
        traceback.print_exc()
        return None

def extract_pdf_content(file_paths):
    """Extract text content from one or multiple PDF files (supports both text and scanned PDFs)"""
    all_content = []
    
    for file_path in file_paths:
        try:
            content = ""
            text_extracted = False
            
            if file_path.lower().endswith('.pdf'):
                # First, try to extract text directly (for text-based PDFs)
                try:
                    with open(file_path, 'rb') as pdf_file:
                        pdf_reader = PyPDF2.PdfReader(pdf_file)
                        
                        # Check if PDF is encrypted
                        if pdf_reader.is_encrypted:
                            try:
                                pdf_reader.decrypt('')
                            except Exception:
                                print(f"Could not decrypt PDF: {file_path}. Trying OCR...")
                                # Try local OCR first, then cloud OCR
                                ocr_content = None
                                if OCR_AVAILABLE:
                                    ocr_content = extract_pdf_content_with_ocr(file_path)
                                if not ocr_content:
                                    ocr_content = extract_pdf_content_with_cloud_ocr(file_path)
                                if ocr_content:
                                    all_content.append(ocr_content)
                                    print(f"OCR extracted content from encrypted PDF ({len(ocr_content)} characters)")
                                continue
                        
                        # Extract text from all pages
                        for page in pdf_reader.pages:
                            try:
                                page_text = page.extract_text()
                                if page_text and page_text.strip():
                                    content += page_text + "\n"
                                    text_extracted = True
                            except Exception as page_error:
                                print(f"Error extracting text from page: {page_error}")
                                continue
                except Exception as pdf_error:
                    print(f"Error reading PDF: {pdf_error}")
                    text_extracted = False
                
                # If no text extracted or very little text, try OCR (for scanned PDFs)
                if not text_extracted or (content.strip() and len(content.strip()) < 100):
                    print(f"Little or no text extracted from PDF. Attempting OCR...")
                    
                    # First try local OCR if available
                    ocr_content = None
                    if OCR_AVAILABLE:
                        ocr_content = extract_pdf_content_with_ocr(file_path)
                        if ocr_content:
                            print(f"Local OCR extracted {len(ocr_content)} characters")
                    
                    # If local OCR failed or not available, try cloud OCR (works in serverless)
                    if not ocr_content:
                        print("Local OCR not available or failed. Trying cloud OCR API...")
                        ocr_content = extract_pdf_content_with_cloud_ocr(file_path)
                        if ocr_content:
                            print(f"Cloud OCR extracted {len(ocr_content)} characters")
                    
                    if ocr_content:
                        # Use OCR content if it's better than extracted text
                        if len(ocr_content) > len(content):
                            content = ocr_content
                            print(f"Using OCR content ({len(ocr_content)} characters)")
                        elif content.strip():
                            # Combine both if we have some text
                            content = content + "\n\n[Additional content from OCR:]\n" + ocr_content
                    else:
                        print("Both local and cloud OCR failed to extract content")
            
            if content.strip():
                all_content.append(content)
            else:
                print(f"Warning: No content extracted from {file_path}")
                
        except Exception as e:
            print(f"Error processing PDF {file_path}: {e}")
            import traceback
            traceback.print_exc()
            continue
    
    # Combine all content
    combined_content = "\n\n".join(all_content)
    return combined_content if combined_content.strip() else None

def generate_quiz_openrouter(topic, difficulty_level, question_type="mcq", num_questions=5, pdf_content=None):
    """Generate quiz using OpenRouter API (primary)"""
    import requests
    
    openrouter_key = os.environ.get('OPENROUTER_API_KEY')
    if not openrouter_key:
        raise Exception("OPENROUTER_API_KEY not set. Set it in Vercel environment variables.")
    
    # Validate API key format
    if not openrouter_key.startswith('sk-or-'):
        print(f"⚠️ WARNING: OpenRouter API key format may be incorrect. Expected 'sk-or-...', got: {openrouter_key[:15]}...")
    else:
        print(f"✅ OpenRouter API key format validated: {openrouter_key[:15]}...")
    
    # Debug: Check API key length (should be around 100+ characters)
    print(f"🔍 Debug: API key length: {len(openrouter_key)} characters")
    
    # Map difficulty levels
    difficulty_mapping = {
        "beginner": {"bloom_level": 1, "description": "Remembering and Understanding level - basic facts, definitions, and simple concepts"},
        "intermediate": {"bloom_level": 3, "description": "Applying and Analyzing level - practical application and analysis of concepts"},
        "advanced": {"bloom_level": 5, "description": "Evaluating and Creating level - critical thinking, evaluation, and synthesis"},
        "difficult": {"bloom_level": 5, "description": "Evaluating and Creating level - critical thinking, evaluation, and synthesis"}
    }
    difficulty_info = difficulty_mapping.get(difficulty_level, difficulty_mapping["beginner"])
    level_description = difficulty_info["description"]
    
    import random
    random_seed = random.randint(1000, 9999)
    
    # Build prompt
    pdf_context = ""
    if pdf_content:
        if len(pdf_content) > 15000:
            pdf_content = pdf_content[:5000] + "\n\n[... content truncated ...]\n\n" + pdf_content[-10000:]
        pdf_context = f"""
            
IMPORTANT: Use the following PDF content as the PRIMARY SOURCE for generating questions. All questions MUST be based on this content:

PDF CONTENT:
{pdf_content}

Generate questions based ONLY on the information provided in the PDF content above. If the topic "{topic}" is mentioned, use it as context, but prioritize the PDF content.
"""
    
    is_advanced = difficulty_level in ("advanced", "difficult")
    is_graph_topic = ("graph" in (topic or "").lower()) or ("dsa" in (topic or "").lower())

    adv_level_block = (
        "- At least 70% questions must be scenario/case based and require algorithm choice, complexity reasoning, or edge-case analysis.\n"
        "- Include topics like shortest paths, MST, topological sort, SCC/bridges/articulation points, flows/matching, DAG DP, and proof-style properties.\n"
        "- Do NOT ask only recall questions like 'what is graph' or 'what is cycle'."
        if is_advanced
        else "- Keep question depth aligned to selected difficulty."
    )
    image_req_block = (
        "- Include image_url in at least 2 questions (graph diagrams/adjacency visuals) when possible.\n"
        "- image_url must be a direct public image link."
        if (is_advanced and is_graph_topic and num_questions >= 3)
        else "- image_url is optional."
    )

    if question_type == "mcq":
        prompt = f"""CRITICAL: You MUST generate questions ONLY on the topic: "{topic}"

Generate exactly {num_questions} multiple-choice questions on the topic: "{topic}" at {difficulty_level.upper()} level ({level_description}).

IMPORTANT REQUIREMENTS:
1. ALL questions MUST be about "{topic}" - do NOT generate questions on other topics
2. Include exactly {num_questions} questions
3. Each question must have exactly 4 answer choices (A, B, C, D)
4. Make questions diverse and varied - avoid repetitive patterns
5. Use randomization seed {random_seed} to ensure variety
6. Include a "level" key specifying the Bloom's Taxonomy level (Remembering, Understanding, Applying, etc.)
7. Return each item with keys: "question", "options", "answer", "type", and optional "image_url"
8. type MUST be exactly "mcq"
9. "answer" MUST be one of: A, B, C, D
10. Avoid trivial definition-only questions unless necessary for context

{pdf_context if pdf_content else ''}

ADVANCED-LEVEL REQUIREMENTS:
{adv_level_block}

IMAGE REQUIREMENT:
{image_req_block}

Return output in valid JSON format ONLY (no explanations, no markdown):
[
    {{"question": "What is AI?", "options": ["A. option1", "B. option2", "C. option3", "D. option4"], "answer": "A", "type": "mcq", "image_url": "https://example.com/diagram.png"}},
    ...
]"""
    elif question_type == "coding":
        prompt = f"""CRITICAL: Generate {num_questions} PROGRAMMING/CODING problems on the topic: "{topic}"

Difficulty Level: {difficulty_level.upper()} ({level_description})
{pdf_context if pdf_content else ''}

IMPORTANT: These must be ACTUAL CODING PROBLEMS that require writing code, NOT theoretical/subjective questions.

REQUIREMENTS:
1. Generate EXACTLY {num_questions} coding problems
2. Each problem must require students to write actual code (Python, Java, C++, or C)
3. Include clear input/output specifications
4. Include test cases with expected outputs
5. Problems should be solvable in 15-30 minutes

EXAMPLES OF GOOD CODING PROBLEMS:
- "Write a function to find the sum of all elements in an array"
- "Implement a function to check if a string is a palindrome"
- "Write a program to find the factorial of a number"
- "Create a function to reverse a linked list"

EXAMPLES OF BAD (DO NOT GENERATE THESE):
- "Explain the difference between arrays and linked lists" (This is subjective, NOT coding)
- "What is recursion?" (This is theoretical, NOT coding)

Return ONLY valid JSON in this exact format:
[
    {{
        "question": "Write a function that takes an array of integers and returns the sum of all elements.\\n\\nInput: First line contains n (size of array). Second line contains n space-separated integers.\\nOutput: Print the sum of all elements.",
        "type": "coding",
        "sample_input": "5\\n1 2 3 4 5",
        "sample_output": "15",
        "test_cases": [
            {{"input": "5\\n1 2 3 4 5", "expected_output": "15", "is_hidden": false}},
            {{"input": "3\\n10 20 30", "expected_output": "60", "is_hidden": false}},
            {{"input": "4\\n-1 -2 3 4", "expected_output": "4", "is_hidden": true}}
        ],
        "time_limit_seconds": 2,
        "memory_limit_mb": 256,
        "starter_code": {{
            "python": "def solve(arr):\\n    # Write your code here\\n    pass\\n\\n# Read input\\nn = int(input())\\narr = list(map(int, input().split()))\\nprint(solve(arr))",
            "java": "import java.util.*;\\npublic class Solution {{\\n    public static void main(String[] args) {{\\n        Scanner sc = new Scanner(System.in);\\n        int n = sc.nextInt();\\n        int[] arr = new int[n];\\n        for(int i=0; i<n; i++) arr[i] = sc.nextInt();\\n        // Write your code here\\n    }}\\n}}",
            "cpp": "#include <iostream>\\nusing namespace std;\\nint main() {{\\n    int n;\\n    cin >> n;\\n    int arr[n];\\n    for(int i=0; i<n; i++) cin >> arr[i];\\n    // Write your code here\\n    return 0;\\n}}",
            "c": "#include <stdio.h>\\nint main() {{\\n    int n;\\n    scanf(\\"%d\\", &n);\\n    int arr[n];\\n    for(int i=0; i<n; i++) scanf(\\"%d\\", &arr[i]);\\n    // Write your code here\\n    return 0;\\n}}"
        }}
    }}
]"""
    else:  # subjective
        prompt = f"""CRITICAL: You MUST generate questions ONLY on the topic: "{topic}"

Generate exactly {num_questions} subjective questions on the topic: "{topic}" at {difficulty_level.upper()} level ({level_description}).

IMPORTANT REQUIREMENTS:
1. ALL questions MUST be about "{topic}" - do NOT generate questions on other topics
2. Include exactly {num_questions} questions
3. Questions should be open-ended and require detailed answers
4. Make questions diverse and varied - avoid repetitive patterns
5. Use randomization seed {random_seed} to ensure variety
6. Include a "level" key specifying the Bloom's Taxonomy level
7. Vary the marks between 5, 10, 15, and 20 marks for different questions

{pdf_context if pdf_content else ''}

Return output in valid JSON format ONLY (no explanations, no markdown):
[
    {{"question": "Explain the concept of AI and its applications", "answer": "Sample answer explaining AI...", "type": "subjective", "marks": 10}},
    ...
]"""
    
    # Try free models first, then paid
    models_to_try = [
        "meta-llama/llama-3.1-8b-instruct:free",  # Completely free
        "mistralai/mistral-7b-instruct:free",     # Free
        "openai/gpt-3.5-turbo",                   # Free tier available
    ]
    
    for model_name in models_to_try:
        try:
            print(f"🔄 Trying OpenRouter model: {model_name}")
            # Ensure API key is properly formatted
            auth_header = f"Bearer {openrouter_key.strip()}"
            print(f"🔍 Debug: Authorization header starts with: {auth_header[:25]}...")
            
            response = requests.post(
                "https://openrouter.ai/api/v1/chat/completions",
                headers={
                    "Authorization": auth_header,
                    "Content-Type": "application/json",
                    "HTTP-Referer": "https://www.unitest.in",
                    "X-Title": "UniTest Quiz Generator"
                },
                json={
                    "model": model_name,
                    "messages": [
                        {"role": "system", "content": "You are an expert quiz generator. Always return valid JSON arrays."},
                        {"role": "user", "content": prompt}
                    ],
                    "temperature": 0.7,
                    "max_tokens": 4000
                },
                timeout=60
            )
            
            if response.status_code == 200:
                result = response.json()
                content = result['choices'][0]['message']['content']
                
                # Parse JSON from response
                json_match = re.search(r"```json\n(.*)\n```", content, re.DOTALL)
                if json_match:
                    questions = json.loads(json_match.group(1))
                else:
                    questions = json.loads(content)
                
                # Validate questions
                for q in questions:
                    if 'type' not in q:
                        q['type'] = question_type
                    
                    # Ensure coding questions have required fields
                    if question_type == 'coding':
                        q['type'] = 'coding'  # Force correct type
                        # Ensure required fields exist
                        if 'sample_input' not in q:
                            q['sample_input'] = ''
                        if 'sample_output' not in q:
                            q['sample_output'] = ''
                        if 'test_cases' not in q or not q['test_cases']:
                            # Generate default test cases from sample if available
                            q['test_cases'] = [
                                {"input": q.get('sample_input', ''), "expected_output": q.get('sample_output', ''), "is_hidden": False}
                            ]
                        if 'time_limit_seconds' not in q:
                            q['time_limit_seconds'] = 2
                        if 'memory_limit_mb' not in q:
                            q['memory_limit_mb'] = 256
                        if 'starter_code' not in q or not q['starter_code']:
                            q['starter_code'] = {
                                "python": "# Write your code here\n",
                                "java": "// Write your code here\n",
                                "cpp": "// Write your code here\n",
                                "c": "// Write your code here\n"
                            }
                
                print(f"✅ OpenRouter SUCCESS: Model '{model_name}' generated {len(questions)} {question_type} questions")
                return questions
            else:
                error_detail = response.text[:200] if response.text else "No error message"
                print(f"  Model {model_name} failed: {response.status_code} - {error_detail}")
                if response.status_code == 401:
                    print(f"  ⚠️ Authentication error - check if API key is correct: {openrouter_key[:15]}...")
                continue
        except Exception as model_error:
            print(f"  Model {model_name} error: {str(model_error)[:100]}")
            continue
    
    raise Exception("All OpenRouter models failed. Please check your API key or try again later.")

def generate_quiz_gemini(topic, difficulty_level, question_type="mcq", num_questions=5, pdf_content=None):
    """Generate quiz using Gemini API (fallback when OpenRouter fails)"""
    if not genai:
        raise Exception("Google Generative AI library not available")
    
    # Check if API key is set
    api_key = os.environ.get('GOOGLE_AI_API_KEY')
    if not api_key:
        raise Exception("GOOGLE_AI_API_KEY environment variable is not set. Please configure it in Vercel: Settings → Environment Variables")

    try:
        # List available models and use the first one that works
        genai.configure(api_key=api_key)
        model = None
        
        # First, try to list available models
        try:
            available_models = genai.list_models()
            model_names = []
            for m in available_models:
                if 'generateContent' in m.supported_generation_methods:
                    # Extract model name (remove 'models/' prefix if present)
                    model_name = m.name.replace('models/', '') if m.name.startswith('models/') else m.name
                    model_names.append(model_name)
                    print(f"  Available model: {model_name}")
            
            # Try available models in order of preference (gemini-2.5-flash is the current free tier model)
            preferred_order = ['gemini-2.5-flash', 'gemini-2.5-flash-lite', 'gemini-1.5-flash-latest', 'gemini-1.5-flash', 'gemini-1.5-pro-latest', 'gemini-1.5-pro', 'gemini-pro']
            for preferred in preferred_order:
                if preferred in model_names:
                    try:
                        model = genai.GenerativeModel(preferred)
                        print(f"✓ Gemini: Using available model: {preferred}")
                        break
                    except Exception as e:
                        print(f"  Failed to use {preferred}: {str(e)[:50]}")
                        continue
            
            # If no preferred model worked, try first available
            if not model and model_names:
                try:
                    first_model = model_names[0]
                    model = genai.GenerativeModel(first_model)
                    print(f"✓ Using first available model: {first_model}")
                except Exception as e:
                    print(f"  Failed to use {first_model}: {str(e)[:50]}")
        except Exception as list_error:
            print(f"⚠️ Could not list models: {list_error}")
        
        # Fallback: try current free tier models directly (gemini-2.5-flash is the primary free tier model)
        if not model:
            for fallback_name in ['gemini-2.5-flash', 'gemini-2.5-flash-lite', 'gemini-1.5-flash-latest', 'gemini-1.5-flash', 'gemini-1.5-pro-latest', 'gemini-1.5-pro']:
                try:
                    model = genai.GenerativeModel(fallback_name)
                    print(f"✓ Using fallback model: {fallback_name}")
                    break
                except Exception as e:
                    print(f"  Fallback {fallback_name} failed: {str(e)[:50]}")
                    continue
        
        if not model:
            raise Exception("No working Gemini model found. Check API key and quota. Available models may have changed.")
        
        # Map difficulty levels to Bloom's taxonomy levels and descriptions
        difficulty_mapping = {
            "beginner": {
                "bloom_level": 1,
                "description": "Remembering and Understanding level - basic facts, definitions, and simple concepts"
            },
            "intermediate": {
                "bloom_level": 3,
                "description": "Applying and Analyzing level - practical application and analysis of concepts"
            },
            "advanced": {
                "bloom_level": 5,
                "description": "Evaluating and Creating level - critical thinking, evaluation, and synthesis"
            },
            "difficult": {
                "bloom_level": 5,
                "description": "Evaluating and Creating level - critical thinking, evaluation, and synthesis"
            }
        }
        
        difficulty_info = difficulty_mapping.get(difficulty_level, difficulty_mapping["beginner"])
        bloom_level = difficulty_info["bloom_level"]
        level_description = difficulty_info["description"]
        
        # Add randomization seed to prompt for variety
        import random
        random_seed = random.randint(1000, 9999)

        # Build prompt with PDF content if provided
        pdf_context = ""
        if pdf_content:
            # Truncate PDF content if too long (keep last 15000 characters to stay within token limits)
            # For multiple PDFs, we want to preserve as much content as possible
            if len(pdf_content) > 15000:
                # Keep first 5000 and last 10000 to preserve both intro and conclusion
                pdf_content = pdf_content[:5000] + "\n\n[... content truncated ...]\n\n" + pdf_content[-10000:]
            pdf_context = f"""
            
IMPORTANT: Use the following PDF content as the PRIMARY SOURCE for generating questions. All questions MUST be based on this content:

PDF CONTENT:
{pdf_content}

Generate questions based ONLY on the information provided in the PDF content above. If the topic "{topic}" is mentioned, use it as context, but prioritize the PDF content.
"""
        
        is_advanced = difficulty_level in ("advanced", "difficult")
        is_graph_topic = ("graph" in (topic or "").lower()) or ("dsa" in (topic or "").lower())

        adv_level_block_gemini = (
            "- At least 70% questions must be scenario/case based and require algorithm choice, complexity reasoning, or edge-case analysis.\n"
            "- Include topics like shortest paths, MST, topological sort, SCC/bridges/articulation points, flows/matching, DAG DP, and proof-style properties.\n"
            "- Do NOT ask only recall questions like 'what is graph' or 'what is cycle'."
            if is_advanced
            else "- Keep question depth aligned to selected difficulty."
        )
        image_req_block_gemini = (
            "- Include image_url in at least 2 questions (graph diagrams/adjacency visuals) when possible.\n"
            "- image_url must be a direct public image link."
            if (is_advanced and is_graph_topic and num_questions >= 3)
            else "- image_url is optional."
        )

        if question_type == "mcq":
            prompt = f"""
CRITICAL: You MUST generate questions ONLY on the topic: "{topic}"

Generate exactly {num_questions} multiple-choice questions on the topic: "{topic}" at {difficulty_level.upper()} level ({level_description}).

IMPORTANT REQUIREMENTS:
1. ALL questions MUST be about "{topic}" - do NOT generate questions on other topics
2. Include exactly {num_questions} questions
3. Each question must have exactly 4 answer choices (A, B, C, D)
4. Make questions diverse and varied - avoid repetitive patterns
5. Use randomization seed {random_seed} to ensure variety
6. Include a "level" key specifying the Bloom's Taxonomy level (Remembering, Understanding, Applying, etc.)
7. Return each item with keys: "question", "options", "answer", "type", and optional "image_url"
8. type MUST be exactly "mcq"
9. "answer" MUST be one of: A, B, C, D
10. Avoid trivial definition-only questions unless necessary for context

{pdf_context if pdf_content else ''}

ADVANCED-LEVEL REQUIREMENTS:
{adv_level_block_gemini}

IMAGE REQUIREMENT:
{image_req_block_gemini}

Return output in valid JSON format ONLY (no explanations, no markdown):
[
    {{"question": "What is AI?", "options": ["A. option1", "B. option2", "C. option3", "D. option4"], "answer": "A", "type": "mcq", "image_url": "https://example.com/diagram.png"}},
    ...
]
            """
        elif question_type == "coding":
            prompt = f"""
CRITICAL: You MUST generate exactly {num_questions} coding programming problems{f' based on the provided PDF content' if pdf_content else f' on the topic: {topic}'}

Difficulty Level: {difficulty_level.upper()} ({level_description})
{pdf_context if pdf_content else ''}

IMPORTANT REQUIREMENTS:
1. Generate EXACTLY {num_questions} coding problems (not MCQ, not subjective, but actual programming problems)
2. Each problem MUST have the following structure:
   - "question": A clear problem statement describing what the student needs to code
   - "type": MUST be exactly "coding" (not "mcq" or anything else)
   - "sample_input": Sample input that demonstrates the problem
   - "sample_output": Expected output for the sample input
   - "test_cases": An array with at least 3-5 test cases, each with:
     * "input": The test input as a string
     * "expected_output": The expected output as a string
     * "is_hidden": boolean (false for visible test cases, true for hidden ones)
   - "time_limit_seconds": 2 (default)
   - "memory_limit_mb": 256 (default)
   - "starter_code": An object with starter code templates for each language:
     * "python": Python starter code
     * "java": Java starter code  
     * "cpp": C++ starter code
     * "c": C starter code

3. Problems should be diverse and test different programming concepts related to {topic}
4. Use randomization seed {random_seed} to ensure variety

EXAMPLE FORMAT (follow this EXACT structure):
[
    {{
        "question": "Write a function to find the maximum element in an array. The function should take an array of integers and return the maximum value.",
        "type": "coding",
        "sample_input": "5\\n1 3 5 2 4",
        "sample_output": "5",
        "test_cases": [
            {{"input": "3\\n1 2 3", "expected_output": "3", "is_hidden": false}},
            {{"input": "4\\n10 5 8 12", "expected_output": "12", "is_hidden": false}},
            {{"input": "5\\n-1 -5 -3 -2 -4", "expected_output": "-1", "is_hidden": true}},
            {{"input": "1\\n42", "expected_output": "42", "is_hidden": true}}
        ],
        "time_limit_seconds": 2,
        "memory_limit_mb": 256,
        "starter_code": {{
            "python": "def find_max(arr):\\n    # Your code here\\n    pass",
            "java": "public class Solution {{\\n    public static int findMax(int[] arr) {{\\n        // Your code here\\n        return 0;\\n    }}\\n}}",
            "cpp": "#include <iostream>\\n#include <vector>\\nusing namespace std;\\n\\nint findMax(vector<int>& arr) {{\\n    // Your code here\\n    return 0;\\n}}",
            "c": "#include <stdio.h>\\n\\nint findMax(int arr[], int n) {{\\n    // Your code here\\n    return 0;\\n}}"
        }}
    }}
]

Return ONLY valid JSON array. Do NOT include any markdown code blocks, explanations, or text outside the JSON array.
            """
        else:  # subjective
            prompt = f"""
CRITICAL: You MUST generate questions ONLY on the topic: "{topic}"

Generate exactly {num_questions} subjective questions on the topic: "{topic}" at {difficulty_level.upper()} level ({level_description}).

IMPORTANT REQUIREMENTS:
1. ALL questions MUST be about "{topic}" - do NOT generate questions on other topics
2. Include exactly {num_questions} questions
3. Questions should be open-ended and require detailed answers
4. Make questions diverse and varied - avoid repetitive patterns
5. Use randomization seed {random_seed} to ensure variety
6. Include a "level" key specifying the Bloom's Taxonomy level
7. Vary the marks between 5, 10, 15, and 20 marks for different questions

{pdf_context if pdf_content else ''}

Return output in valid JSON format ONLY (no explanations, no markdown):
[
    {{"question": "Explain the concept of AI and its applications", "answer": "Sample answer explaining AI...", "type": "subjective", "marks": 10}},
    ...
]
            """

        response = model.generate_content(prompt)

        if not response.text:
            raise ValueError("Empty response from AI")

        json_match = re.search(r"```json\n(.*)\n```", response.text, re.DOTALL)
        if json_match:
            questions = json.loads(json_match.group(1))
        else:
            try:
                questions = json.loads(response.text)
            except:
                raise ValueError("Invalid response format from AI")

        # Validate that questions match the requested type
        if questions:
            # Ensure all questions have the correct type
            for q in questions:
                if 'type' not in q:
                    q['type'] = question_type
                elif q.get('type') != question_type:
                    print(f"WARNING: Question type mismatch. Expected {question_type}, got {q.get('type')}")
                    q['type'] = question_type  # Force correct type
                
                # Ensure coding questions have required fields
                if question_type == 'coding':
                    q['type'] = 'coding'  # Force correct type
                    # Ensure required fields exist
                    if 'sample_input' not in q:
                        q['sample_input'] = ''
                    if 'sample_output' not in q:
                        q['sample_output'] = ''
                    if 'test_cases' not in q or not q['test_cases']:
                        # Generate default test cases from sample if available
                        q['test_cases'] = [
                            {"input": q.get('sample_input', ''), "expected_output": q.get('sample_output', ''), "is_hidden": False}
                        ]
                    if 'time_limit_seconds' not in q:
                        q['time_limit_seconds'] = 2
                    if 'memory_limit_mb' not in q:
                        q['memory_limit_mb'] = 256
                    if 'starter_code' not in q or not q['starter_code']:
                        q['starter_code'] = {
                            "python": "# Write your code here\n",
                            "java": "// Write your code here\n",
                            "cpp": "// Write your code here\n",
                            "c": "// Write your code here\n"
                        }
            
            print(f"DEBUG: Validated {len(questions)} questions, all have type: {question_type}")

        return questions

    except Exception as e:
        error_info = handle_gemini_api_error(e, "generate_quiz_gemini")
        error_msg = error_info.get('user_message', str(e))
        print(f"Error in generate_quiz_gemini: {str(e)}")
        print(f"Error details: {error_info}")
        import traceback
        traceback.print_exc()
        raise Exception(error_msg)

def generate_quiz(topic, difficulty_level, question_type="mcq", num_questions=5, pdf_content=None):
    """Generate quiz - tries OpenRouter first, falls back to Gemini if OpenRouter fails"""
    
    print("=" * 60)
    print("📝 QUIZ GENERATION STARTED")
    print(f"   Topic: {topic}")
    print(f"   Type: {question_type}, Difficulty: {difficulty_level}, Count: {num_questions}")
    print("=" * 60)
    
    openrouter_error = None
    # Try OpenRouter first (primary)
    openrouter_key = os.environ.get('OPENROUTER_API_KEY')
    if openrouter_key:
        print("✅ OpenRouter API key found")
        try:
            print("🚀 PRIMARY: Attempting OpenRouter API...")
            result = generate_quiz_openrouter(topic, difficulty_level, question_type, num_questions, pdf_content)
            print("✅ SUCCESS: Quiz generated using OpenRouter API (primary)")
            print("=" * 60)
            return result
        except Exception as e:
            openrouter_error = e
            print(f"⚠️ OpenRouter failed: {str(e)[:100]}")
            print("🔄 FALLBACK: Switching to Gemini API...")
    else:
        print("⚠️ OpenRouter API key not found - using Gemini as primary")
    
    # Fallback to Gemini if OpenRouter fails or not configured
    try:
        print("🔄 FALLBACK: Attempting Gemini API...")
        result = generate_quiz_gemini(topic, difficulty_level, question_type, num_questions, pdf_content)
        print("✅ SUCCESS: Quiz generated using Gemini API (fallback)")
        print("=" * 60)
        return result
    except Exception as gemini_error:
        # If both fail, raise error
        openrouter_msg = str(openrouter_error)[:100] if openrouter_error else 'not configured'
        error_msg = f"Both OpenRouter and Gemini failed. OpenRouter: {openrouter_msg}, Gemini: {str(gemini_error)[:100]}"
        print(f"❌ FAILED: {error_msg}")
        print("=" * 60)
        raise Exception("Failed to generate quiz. Please check your API keys configuration.")

def process_document(file_path):
    """Process uploaded document to extract content and topic"""
    try:
        # Ensure NLTK data is available
        ensure_nltk_data()
        
        content = ""
        if file_path.lower().endswith('.pdf'):
            try:
                with open(file_path, 'rb') as pdf_file:
                    pdf_reader = PyPDF2.PdfReader(pdf_file)
                    
                    # Check if PDF is encrypted
                    if pdf_reader.is_encrypted:
                        print("PDF is encrypted, attempting to decrypt...")
                        try:
                            pdf_reader.decrypt('')
                        except Exception as decrypt_error:
                            print(f"Could not decrypt PDF: {decrypt_error}")
                            # Try using AI to extract topic from filename or metadata
                            return extract_topic_from_pdf_metadata(file_path)
                    
                    # Extract text from all pages
                    num_pages = len(pdf_reader.pages)
                    print(f"Processing PDF with {num_pages} pages...")
                    
                    for i, page in enumerate(pdf_reader.pages):
                        try:
                            page_text = page.extract_text()
                            if page_text:
                                content += page_text + " "
                        except Exception as page_error:
                            print(f"Error extracting text from page {i+1}: {page_error}")
                            continue
                    
                    # If no content extracted, try metadata
                    if not content.strip():
                        print("No text extracted from PDF pages, trying metadata...")
                        metadata_topic = extract_topic_from_pdf_metadata(file_path)
                        if metadata_topic:
                            return metadata_topic
                        
                        # If still no content, try using AI with filename
                        print("Attempting AI-based topic extraction from filename...")
                        return extract_topic_from_filename(file_path)
                        
            except Exception as pdf_error:
                print(f"Error reading PDF file: {pdf_error}")
                # Fallback to filename-based extraction
                return extract_topic_from_filename(file_path)
        else:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as text_file:
                content = text_file.read()

        # If still no content, return None
        if not content or not content.strip():
            print("No content extracted from document")
            return extract_topic_from_filename(file_path)

        # Extract topic using word frequency analysis
        try:
            tokens = word_tokenize(content.lower())
            stop_words = set(stopwords.words('english'))
            meaningful_words = [word for word in tokens if word.isalnum() and len(word) > 2 and word not in stop_words]

            if not meaningful_words:
                print("No meaningful words found after filtering")
                return extract_topic_from_filename(file_path)

            word_freq = Counter(meaningful_words)
            # Get top 3 most common words to better identify topic
            top_words = word_freq.most_common(5)
            print(f"Top words found: {top_words}")
            
            # Filter out common generic words
            generic_words = {'chapter', 'page', 'section', 'question', 'answer', 'example', 'figure', 'table'}
            for word, count in top_words:
                if word not in generic_words and len(word) > 3:
                    main_topic = word.capitalize()
                    print(f"Extracted topic: {main_topic}")
                    return main_topic
            
            # If all top words are generic, use the first one
            if top_words:
                main_topic = top_words[0][0].capitalize()
                print(f"Extracted topic (fallback): {main_topic}")
                return main_topic
                
        except Exception as token_error:
            print(f"Error in tokenization: {token_error}")
            return extract_topic_from_filename(file_path)

        return None

    except Exception as e:
        print(f"Error processing document: {str(e)}")
        import traceback
        traceback.print_exc()
        # Final fallback: try to extract from filename
        return extract_topic_from_filename(file_path)

def extract_topic_from_pdf_metadata(file_path):
    """Extract topic from PDF metadata"""
    try:
        with open(file_path, 'rb') as pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            metadata = pdf_reader.metadata
            
            if metadata:
                # Try to extract topic from title or subject
                title = metadata.get('/Title', '')
                subject = metadata.get('/Subject', '')
                
                if title:
                    # Extract first meaningful word from title
                    words = title.split()
                    for word in words:
                        if len(word) > 3 and word.isalnum():
                            return word.capitalize()
                
                if subject:
                    words = subject.split()
                    for word in words:
                        if len(word) > 3 and word.isalnum():
                            return word.capitalize()
    except Exception as e:
        print(f"Error extracting from metadata: {e}")
    
    return None

def extract_topic_from_filename(file_path):
    """Extract topic from filename as last resort"""
    try:
        import os
        # Get filename from path
        filename = os.path.basename(file_path)
        # Remove extension
        filename_without_ext = os.path.splitext(filename)[0]
        
        print(f"Attempting to extract topic from filename: {filename_without_ext}")
        
        # Try to extract meaningful words from filename
        # Remove common prefixes/suffixes and clean up
        filename_clean = filename_without_ext.replace('_', ' ').replace('-', ' ').replace('.', ' ')
        # Remove multiple spaces
        filename_clean = ' '.join(filename_clean.split())
        words = filename_clean.split()
        
        print(f"Words extracted from filename: {words}")
        
        # Filter out common words (expanded list)
        common_words = {'pdf', 'document', 'file', 'quiz', 'question', 'bank', 'qb', 'notes', 'paper', 'exam', 'test', 'assignment', 'hw', 'homework'}
        
        # First pass: look for meaningful words (longer than 3 chars, not common)
        for word in words:
            word_clean = word.lower().strip('.,!?()[]{}')
            # Remove any non-alphanumeric characters except numbers
            word_clean = ''.join(c for c in word_clean if c.isalnum() or c.isdigit())
            if len(word_clean) > 3 and word_clean not in common_words:
                topic = word_clean.capitalize()
                print(f"Extracted topic from filename (first pass): {topic}")
                return topic
        
        # Second pass: look for any word longer than 2 characters (includes codes like "BCS702")
        for word in words:
            word_clean = word.upper().strip('.,!?()[]{}')
            # Keep alphanumeric characters (including numbers)
            word_clean = ''.join(c for c in word_clean if c.isalnum() or c.isdigit())
            if len(word_clean) >= 3 and word_clean not in common_words:
                # Check if it looks like a course code (e.g., BCS702, CS101)
                if any(c.isdigit() for c in word_clean) and any(c.isalpha() for c in word_clean):
                    topic = word_clean
                    print(f"Extracted topic from filename (course code): {topic}")
                    return topic
                elif len(word_clean) > 3:
                    topic = word_clean.capitalize()
                    print(f"Extracted topic from filename (second pass): {topic}")
                    return topic
        
        # Third pass: use first meaningful word (minimum 2 chars)
        for word in words:
            word_clean = word.upper().strip('.,!?()[]{}')
            word_clean = ''.join(c for c in word_clean if c.isalnum() or c.isdigit())
            if len(word_clean) >= 2 and word_clean not in common_words:
                topic = word_clean
                print(f"Extracted topic from filename (third pass): {topic}")
                return topic
        
        # Last resort: use first 10 characters of filename (cleaned)
        if filename_without_ext:
            # Remove special characters
            topic_clean = ''.join(c for c in filename_without_ext[:15] if c.isalnum() or c.isspace())
            topic_clean = topic_clean.strip()
            if topic_clean:
                topic = topic_clean.capitalize()
                print(f"Using filename as topic (fallback): {topic}")
                return topic
            
    except Exception as e:
        print(f"Error extracting from filename: {e}")
        import traceback
        traceback.print_exc()
    
    return None

# Routes
@app.route('/')
def home():
    """Home page - PUBLIC ROUTE (no auth required)"""
    # Don't initialize database here - it can cause 401 errors if DB fails
    # Database will be initialized when needed (lazy initialization)
    try:
        return render_template('home.html')
    except Exception as e:
        print(f"Error rendering home template: {e}")
        # Fallback HTML if template fails - ensures home page is always accessible
        return '''
        <!DOCTYPE html>
        <html>
        <head>
            <title>UNITEST - AI-Powered Learning Platform</title>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
        </head>
        <body>
            <h1>Welcome to UNITEST</h1>
            <p>AI-Powered Quiz Generator & Learning Platform</p>
            <a href="/login">Login</a> | <a href="/signup">Sign Up</a>
        </body>
        </html>
        ''', 200

@app.route('/favicon.ico')
def favicon():
    """Serve favicon for better Google search results"""
    return send_file('static/favicon.ico', mimetype='image/x-icon')

@app.route('/favicon.png')
def favicon_png():
    """Serve favicon PNG if needed"""
    return send_file('static/favicon.ico', mimetype='image/x-icon')

@app.route('/test')
def test():
    """Simple test route"""
    return jsonify({
        'message': 'Flask app is working!',
        'environment': 'production' if os.environ.get('DATABASE_URL') else 'development',
        'secret_key_set': bool(os.environ.get('SECRET_KEY')),
        'database_url_set': bool(os.environ.get('DATABASE_URL')),
        'api_key_set': bool(os.environ.get('GOOGLE_AI_API_KEY'))
    })

@app.route('/health')
def health_check():
    """Health check endpoint for debugging"""
    try:
        # Test database connection
        db.session.execute('SELECT 1')
        return jsonify({
            'status': 'healthy',
            'database': 'connected',
            'environment': 'production' if os.environ.get('DATABASE_URL') else 'development'
        })
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'error': str(e),
            'database': 'disconnected'
        }), 500

@app.route('/test-sitemap')
def test_sitemap():
    """Test endpoint to verify sitemap is accessible"""
    try:
        from flask import Response
        from datetime import datetime, timezone
        current_date = datetime.now(timezone.utc).strftime('%Y-%m-%d')
        return jsonify({
            'status': 'success',
            'sitemap_url': 'https://unitest.in/sitemap.xml',
            'current_date': current_date,
            'message': 'Sitemap should be accessible at /sitemap.xml'
        })
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/sitemap.xml', methods=['GET', 'HEAD'])
def sitemap():
    """Serve sitemap.xml for Google Search Console - PUBLIC ROUTE (no auth required)"""
    from flask import Response
    from datetime import datetime, timezone
    
    # Get current date in UTC to avoid timezone issues - format: YYYY-MM-DD
    current_date = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    
    # Always return inline sitemap with current date - most reliable for Vercel serverless
    sitemap_content = f'''<?xml version="1.0" encoding="UTF-8"?>
<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
  <url>
    <loc>https://unitest.in/</loc>
    <lastmod>{current_date}</lastmod>
    <changefreq>weekly</changefreq>
    <priority>1.0</priority>
  </url>
  <url>
    <loc>https://unitest.in/login</loc>
    <lastmod>{current_date}</lastmod>
    <changefreq>monthly</changefreq>
    <priority>0.8</priority>
  </url>
  <url>
    <loc>https://unitest.in/signup</loc>
    <lastmod>{current_date}</lastmod>
    <changefreq>monthly</changefreq>
    <priority>0.8</priority>
  </url>
  <url>
    <loc>https://unitest.in/dashboard</loc>
    <lastmod>{current_date}</lastmod>
    <changefreq>weekly</changefreq>
    <priority>0.9</priority>
  </url>
  <url>
    <loc>https://unitest.in/quiz</loc>
    <lastmod>{current_date}</lastmod>
    <changefreq>weekly</changefreq>
    <priority>0.9</priority>
  </url>
</urlset>'''
    
    response = Response(sitemap_content, mimetype='application/xml')
    response.headers['Content-Type'] = 'application/xml; charset=utf-8'
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    # Allow all user agents including Googlebot
    response.headers['Access-Control-Allow-Origin'] = '*'
    return response

@app.route('/robots.txt')
def robots():
    return send_file('static/robots.txt', mimetype='text/plain')

@app.route('/google77cd707098d48f23.html')
def google_verification():
    """Google Search Console verification file"""
    return send_file('static/google77cd707098d48f23.html', mimetype='text/html')

@app.route('/static/<path:filename>')
def static_files(filename):
    return send_file(f'static/{filename}')

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        try:
            username = request.form['username']
            email = request.form['email']
            password = request.form['password']
            confirm_password = request.form['confirm_password']
            role = request.form.get('role', 'student')

            if not all([username, email, password, confirm_password]):
                flash('Please fill in all fields', 'error')
                return redirect(url_for('signup'))

            if password != confirm_password:
                flash('Passwords do not match', 'error')
                return redirect(url_for('signup'))

            # Check if user exists
            existing_user = db.session.query(User).filter_by(username=username).first()
            if existing_user:
                flash('Username already exists', 'error')
                return redirect(url_for('signup'))

            existing_email = db.session.query(User).filter_by(email=email).first()
            if existing_email:
                flash('Email already exists', 'error')
                return redirect(url_for('signup'))

            email_otp = _generate_otp()
            expires_at = (datetime.utcnow() + timedelta(minutes=10)).isoformat()

            email_sent = _send_email_otp(email, email_otp)
            if not email_sent:
                flash('Unable to send email OTP right now. Please try again.', 'error')
                return redirect(url_for('signup'))

            session['signup_otp_pending'] = {
                'username': username,
                'email': email.lower(),
                'password_hash': generate_password_hash(password),
                'role': role if role in ['student', 'teacher'] else 'student',
                'email_otp_hash': _hash_otp(email_otp),
                'otp_expires_at': expires_at,
                'email_verified': False
            }
            flash('OTP sent to your email. Verify it to complete signup.', 'success')
            return redirect(url_for('verify_signup_otp'))
        except Exception as e:
            print(f"Error in signup: {str(e)}")
            db.session.rollback()
            flash(f'Signup failed: {str(e)}', 'error')
            return redirect(url_for('signup'))

    return render_template('signup.html')

@app.route('/verify-signup-otp', methods=['GET', 'POST'])
def verify_signup_otp():
    pending = session.get('signup_otp_pending')
    if not pending:
        flash('No pending signup verification found.', 'error')
        return redirect(url_for('signup'))

    if request.method == 'POST':
        action = request.form.get('action', 'verify')
        now = datetime.utcnow()
        expires_at = datetime.fromisoformat(pending['otp_expires_at'])

        if now > expires_at:
            session.pop('signup_otp_pending', None)
            flash('OTP expired. Please sign up again.', 'error')
            return redirect(url_for('signup'))

        if action == 'resend_email':
            otp = _generate_otp()
            if _send_email_otp(pending['email'], otp):
                pending['email_otp_hash'] = _hash_otp(otp)
                pending['otp_expires_at'] = (datetime.utcnow() + timedelta(minutes=10)).isoformat()
                session['signup_otp_pending'] = pending
                flash('Email OTP resent.', 'success')
            else:
                flash('Failed to resend email OTP.', 'error')
            return redirect(url_for('verify_signup_otp'))

        email_otp = request.form.get('email_otp', '').strip()
        if _hash_otp(email_otp) != pending['email_otp_hash']:
            flash('Invalid email OTP.', 'error')
            return redirect(url_for('verify_signup_otp'))

        try:
            user = User(
                username=pending['username'],
                email=pending['email'],
                password_hash=pending['password_hash'],
                role=pending['role'],
                email_verified=True,
                auth_provider='local'
            )
            db.session.add(user)
            db.session.commit()
            session.pop('signup_otp_pending', None)
            flash('Account created and verified successfully! Please login.', 'success')
            return redirect(url_for('login'))
        except Exception as e:
            print(f"Error in verify_signup_otp: {e}")
            db.session.rollback()
            flash('Verification failed. Please try signup again.', 'error')
            return redirect(url_for('signup'))

    return render_template('verify_signup_otp.html', pending=pending)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        try:
            username = request.form['username']
            password = request.form['password']

            user = db.session.query(User).filter_by(username=username).first()
            if user and check_password_hash(user.password_hash, password):
                # Backward compatibility: users created before email-OTP rollout can
                # verify themselves by successfully logging in with existing credentials.
                if not user.email_verified and user.auth_provider == 'local':
                    user.email_verified = True
                    db.session.commit()
                    flash('Legacy account verified. Login successful.', 'success')
                elif not user.email_verified:
                    flash('Please complete email verification before login.', 'error')
                    return redirect(url_for('login'))
                _record_login_event(user)
                db.session.commit()
                login_user(user)
                flash('Logged in successfully!', 'success')
                return redirect(url_for('dashboard'))
            else:
                flash('Invalid username or password', 'error')
                return redirect(url_for('login'))
        except Exception as e:
            print(f"Error in login: {str(e)}")
            db.session.rollback()
            flash(f'Login error: {str(e)}', 'error')
            return redirect(url_for('login'))

    return render_template('login.html')

@app.route('/auth/google/start')
def google_auth_start():
    client_id = os.environ.get('GOOGLE_OAUTH_CLIENT_ID')
    if not client_id:
        flash('Google login is not configured yet. Set GOOGLE_OAUTH_CLIENT_ID.', 'error')
        return redirect(url_for('login'))

    state = secrets.token_urlsafe(24)
    session['google_oauth_state'] = state
    redirect_uri = url_for('google_auth_callback', _external=True)
    query = urlencode({
        'client_id': client_id,
        'redirect_uri': redirect_uri,
        'response_type': 'code',
        'scope': 'openid email profile',
        'state': state,
        'access_type': 'offline',
        'prompt': 'select_account'
    })
    return redirect(f"https://accounts.google.com/o/oauth2/v2/auth?{query}")

@app.route('/auth/google/callback')
def google_auth_callback():
    try:
        state = request.args.get('state', '')
        if not state or state != session.get('google_oauth_state'):
            flash('Google login state mismatch. Try again.', 'error')
            return redirect(url_for('login'))

        code = request.args.get('code')
        if not code:
            flash('Google login failed: missing authorization code.', 'error')
            return redirect(url_for('login'))

        client_id = os.environ.get('GOOGLE_OAUTH_CLIENT_ID')
        client_secret = os.environ.get('GOOGLE_OAUTH_CLIENT_SECRET')
        if not (client_id and client_secret):
            flash('Google login is not fully configured.', 'error')
            return redirect(url_for('login'))

        redirect_uri = url_for('google_auth_callback', _external=True)
        token_resp = requests.post(
            "https://oauth2.googleapis.com/token",
            data={
                'code': code,
                'client_id': client_id,
                'client_secret': client_secret,
                'redirect_uri': redirect_uri,
                'grant_type': 'authorization_code'
            },
            timeout=20
        )
        if token_resp.status_code != 200:
            print(f"Google token exchange failed: {token_resp.text[:500]}")
            flash('Google login failed during token exchange.', 'error')
            return redirect(url_for('login'))

        access_token = token_resp.json().get('access_token')
        if not access_token:
            flash('Google login failed: missing access token.', 'error')
            return redirect(url_for('login'))

        userinfo_resp = requests.get(
            "https://openidconnect.googleapis.com/v1/userinfo",
            headers={'Authorization': f'Bearer {access_token}'},
            timeout=20
        )
        if userinfo_resp.status_code != 200:
            print(f"Google userinfo failed: {userinfo_resp.text[:500]}")
            flash('Google login failed while fetching profile.', 'error')
            return redirect(url_for('login'))

        profile = userinfo_resp.json()
        email = (profile.get('email') or '').lower()
        google_sub = profile.get('sub')
        if not email or not google_sub:
            flash('Google profile is missing required fields.', 'error')
            return redirect(url_for('login'))

        user = db.session.query(User).filter_by(google_id=google_sub).first()
        if not user:
            user = db.session.query(User).filter_by(email=email).first()

        if not user:
            username_seed = profile.get('name') or email.split('@')[0]
            user = User(
                username=_make_unique_username(username_seed),
                email=email,
                password_hash=generate_password_hash(secrets.token_urlsafe(24)),
                role='student',
                email_verified=True,
                google_id=google_sub,
                auth_provider='google'
            )
            db.session.add(user)
            db.session.commit()
        else:
            user.google_id = user.google_id or google_sub
            user.auth_provider = 'google'
            user.email_verified = True
            db.session.commit()

        _record_login_event(user)
        db.session.commit()
        login_user(user)
        flash('Logged in successfully with Google!', 'success')
        return redirect(url_for('dashboard'))
    except Exception as e:
        print(f"Google callback error: {e}")
        db.session.rollback()
        flash('Google login failed. Please try again.', 'error')
        return redirect(url_for('login'))

@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        try:
            email = request.form.get('email', '').strip()
            username = request.form.get('username', '').strip()
            
            if not email and not username:
                flash('Please enter either your email or username', 'error')
                return redirect(url_for('forgot_password'))
            
            # Find user by email or username
            if email:
                user = db.session.query(User).filter_by(email=email).first()
            else:
                user = db.session.query(User).filter_by(username=username).first()
            
            if user:
                # Generate reset token
                reset_token = secrets.token_urlsafe(32)
                user.reset_token = reset_token
                user.reset_token_expiry = datetime.utcnow() + timedelta(hours=1)  # Token valid for 1 hour
                db.session.commit()
                
                # Generate reset link
                reset_link = url_for('reset_password', token=reset_token, _external=True)

                sent = _send_password_reset_email(user.email, reset_link)
                if not sent:
                    flash('Could not send reset email right now. Please verify mail settings and try again.', 'error')
                    return redirect(url_for('forgot_password'))
                flash('If an account exists with that email/username, a password reset link has been sent.', 'info')
                return redirect(url_for('forgot_password'))
            else:
                # Don't reveal if user exists for security
                flash('If an account exists with that email/username, a password reset link has been sent.', 'info')
                return redirect(url_for('forgot_password'))
        except Exception as e:
            print(f"Error in forgot_password: {str(e)}")
            db.session.rollback()
            flash(f'Error processing request: {str(e)}', 'error')
            return redirect(url_for('forgot_password'))
    
    return render_template('forgot_password.html')

@app.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    if request.method == 'POST':
        try:
            new_password = request.form.get('password', '').strip()
            confirm_password = request.form.get('confirm_password', '').strip()
            
            if not new_password or not confirm_password:
                flash('Please fill in all fields', 'error')
                return redirect(url_for('reset_password', token=token))
            
            if new_password != confirm_password:
                flash('Passwords do not match', 'error')
                return redirect(url_for('reset_password', token=token))
            
            if len(new_password) < 6:
                flash('Password must be at least 6 characters long', 'error')
                return redirect(url_for('reset_password', token=token))
            
            # Find user by token
            user = db.session.query(User).filter_by(reset_token=token).first()
            
            if not user:
                flash('Invalid or expired reset token', 'error')
                return redirect(url_for('forgot_password'))
            
            # Check if token is expired
            if user.reset_token_expiry and user.reset_token_expiry < datetime.utcnow():
                flash('Reset token has expired. Please request a new one.', 'error')
                user.reset_token = None
                user.reset_token_expiry = None
                db.session.commit()
                return redirect(url_for('forgot_password'))
            
            # Update password
            user.password_hash = generate_password_hash(new_password)
            user.reset_token = None
            user.reset_token_expiry = None
            db.session.commit()
            
            flash('Password reset successfully! You can now login with your new password.', 'success')
            return redirect(url_for('login'))
        except Exception as e:
            print(f"Error in reset_password: {str(e)}")
            db.session.rollback()
            flash(f'Error resetting password: {str(e)}', 'error')
            return redirect(url_for('reset_password', token=token))
    
    # GET request - show reset form
    user = db.session.query(User).filter_by(reset_token=token).first()
    if not user:
        flash('Invalid or expired reset token', 'error')
        return redirect(url_for('forgot_password'))
    
    if user.reset_token_expiry and user.reset_token_expiry < datetime.utcnow():
        flash('Reset token has expired. Please request a new one.', 'error')
        user.reset_token = None
        user.reset_token_expiry = None
        db.session.commit()
        return redirect(url_for('forgot_password'))
    
    return render_template('reset_password.html', token=token)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('home'))

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    if request.method == 'POST':
        action = request.form.get('action', 'update_profile')

        if action == 'update_profile':
            try:
                new_username = request.form.get('username', '').strip()
                new_email = request.form.get('email', '').strip().lower()

                if not new_username or not new_email:
                    flash('Username and email are required.', 'error')
                    return redirect(url_for('profile'))

                if new_username != current_user.username:
                    existing_username = db.session.query(User).filter(User.username == new_username, User.id != current_user.id).first()
                    if existing_username:
                        flash('Username already taken.', 'error')
                        return redirect(url_for('profile'))
                    current_user.username = new_username

                if new_email != current_user.email:
                    existing_email = db.session.query(User).filter(User.email == new_email, User.id != current_user.id).first()
                    if existing_email:
                        flash('Email already in use by another account.', 'error')
                        return redirect(url_for('profile'))

                    otp = _generate_otp()
                    if not _send_email_otp(new_email, otp):
                        flash('Could not send verification OTP to new email.', 'error')
                        return redirect(url_for('profile'))

                    session['profile_email_change_pending'] = {
                        'new_email': new_email,
                        'otp_hash': _hash_otp(otp),
                        'expires_at': (datetime.utcnow() + timedelta(minutes=10)).isoformat()
                    }
                    db.session.commit()
                    flash('Profile updated. Verify OTP sent to your new email to finalize email change.', 'success')
                    return redirect(url_for('profile'))

                db.session.commit()
                flash('Profile updated successfully.', 'success')
                return redirect(url_for('profile'))
            except Exception as e:
                print(f"Profile update error: {e}")
                db.session.rollback()
                flash('Could not update profile. Please try again.', 'error')
                return redirect(url_for('profile'))

        if action == 'send_current_email_otp':
            otp = _generate_otp()
            if _send_email_otp(current_user.email, otp):
                session['profile_email_verify_pending'] = {
                    'email': current_user.email,
                    'otp_hash': _hash_otp(otp),
                    'expires_at': (datetime.utcnow() + timedelta(minutes=10)).isoformat()
                }
                flash('Verification OTP sent to your current email.', 'success')
            else:
                flash('Could not send verification OTP right now.', 'error')
            return redirect(url_for('profile'))

        if action == 'verify_current_email_otp':
            pending = session.get('profile_email_verify_pending')
            if not pending:
                flash('No active email verification request. Send OTP first.', 'error')
                return redirect(url_for('profile'))
            if pending.get('email') != current_user.email:
                session.pop('profile_email_verify_pending', None)
                flash('Email changed. Please request OTP again.', 'error')
                return redirect(url_for('profile'))
            if datetime.utcnow() > datetime.fromisoformat(pending['expires_at']):
                session.pop('profile_email_verify_pending', None)
                flash('OTP expired. Please request a new one.', 'error')
                return redirect(url_for('profile'))
            entered = request.form.get('current_email_otp', '').strip()
            if _hash_otp(entered) != pending['otp_hash']:
                flash('Invalid OTP.', 'error')
                return redirect(url_for('profile'))
            current_user.email_verified = True
            db.session.commit()
            session.pop('profile_email_verify_pending', None)
            flash('Email verified successfully.', 'success')
            return redirect(url_for('profile'))

        if action == 'verify_new_email_otp':
            pending = session.get('profile_email_change_pending')
            if not pending:
                flash('No pending email change found.', 'error')
                return redirect(url_for('profile'))
            if datetime.utcnow() > datetime.fromisoformat(pending['expires_at']):
                session.pop('profile_email_change_pending', None)
                flash('OTP expired. Update email again to get a fresh OTP.', 'error')
                return redirect(url_for('profile'))
            entered = request.form.get('new_email_otp', '').strip()
            if _hash_otp(entered) != pending['otp_hash']:
                flash('Invalid OTP for new email.', 'error')
                return redirect(url_for('profile'))
            current_user.email = pending['new_email']
            current_user.email_verified = True
            db.session.commit()
            session.pop('profile_email_change_pending', None)
            flash('New email verified and updated successfully.', 'success')
            return redirect(url_for('profile'))

        if action == 'resend_new_email_otp':
            pending = session.get('profile_email_change_pending')
            if not pending:
                flash('No pending email change found.', 'error')
                return redirect(url_for('profile'))
            otp = _generate_otp()
            if _send_email_otp(pending['new_email'], otp):
                pending['otp_hash'] = _hash_otp(otp)
                pending['expires_at'] = (datetime.utcnow() + timedelta(minutes=10)).isoformat()
                session['profile_email_change_pending'] = pending
                flash('OTP resent to new email.', 'success')
            else:
                flash('Failed to resend OTP.', 'error')
            return redirect(url_for('profile'))

        if action == 'change_password':
            current_password = request.form.get('current_password', '')
            new_password = request.form.get('new_password', '')
            confirm_new_password = request.form.get('confirm_new_password', '')

            if not current_password or not new_password or not confirm_new_password:
                flash('Please fill all password fields.', 'error')
                return redirect(url_for('profile'))

            if not check_password_hash(current_user.password_hash, current_password):
                flash('Current password is incorrect.', 'error')
                return redirect(url_for('profile'))

            if new_password != confirm_new_password:
                flash('New password and confirm password do not match.', 'error')
                return redirect(url_for('profile'))

            if len(new_password) < 6:
                flash('New password must be at least 6 characters long.', 'error')
                return redirect(url_for('profile'))

            if current_password == new_password:
                flash('New password must be different from current password.', 'error')
                return redirect(url_for('profile'))

            try:
                current_user.password_hash = generate_password_hash(new_password)
                db.session.commit()
                flash('Password updated successfully.', 'success')
            except Exception as e:
                print(f"Password update error: {e}")
                db.session.rollback()
                flash('Could not update password. Please try again.', 'error')
            return redirect(url_for('profile'))

    return render_template(
        'profile.html',
        email_change_pending=session.get('profile_email_change_pending'),
        email_verify_pending=session.get('profile_email_verify_pending')
    )

@app.route('/dashboard')
@login_required
def dashboard():
    progress_records = db.session.query(Progress).filter_by(user_id=current_user.id).all()
    # Teacher's quizzes
    my_quizzes = []
    show_archived = request.args.get('show_archived', '0') == '1'
    if getattr(current_user, 'role', 'student') == 'teacher':
        teacher_quiz_query = db.session.query(Quiz).filter_by(created_by=current_user.id)
        if not show_archived:
            teacher_quiz_query = teacher_quiz_query.filter(Quiz.is_archived == False)  # noqa: E712
        my_quizzes = teacher_quiz_query.order_by(Quiz.created_at.desc()).all()
    # Student shared quiz history
    my_submissions = []
    if getattr(current_user, 'role', 'student') == 'student':
        my_submissions = db.session.query(QuizSubmission).filter_by(student_id=current_user.id).order_by(QuizSubmission.submitted_at.desc()).all()
    current_time = datetime.utcnow()  # For 15-minute review unlock
    return render_template(
        'dashboard.html',
        progress_records=progress_records,
        my_quizzes=my_quizzes,
        my_submissions=my_submissions,
        current_time=current_time,
        show_archived=show_archived
    )

@app.route('/placement_track')
@login_required
def placement_track():
    state = _get_placement_state()
    scores = state.get('scores', {})
    completed = state.get('completed', {})
    levels = state.get('levels', {})

    # Group discussion/interview track.
    gd_attempts = db.session.query(PlacementInterview).filter_by(user_id=current_user.id).order_by(
        PlacementInterview.created_at.desc()
    ).limit(10).all()
    total_attempts = sum(1 for k in PLACEMENT_SEQUENCE if completed.get(k, False))

    # Weighted readiness index.
    readiness_index = (
        float(scores.get('aptitude', 0.0)) * 0.25 +
        float(scores.get('group_discussion', 0.0)) * 0.20 +
        float(scores.get('fundamentals', 0.0)) * 0.20 +
        float(scores.get('basic_coding', 0.0)) * 0.15 +
        float(scores.get('coding', 0.0)) * 0.20
    )

    module_scores = {
        key: round(float(scores.get(key, 0.0)), 1)
        for key in PLACEMENT_SEQUENCE
    }
    module_levels = {}
    for key in PLACEMENT_SEQUENCE:
        lv = levels.get(key, {})
        module_levels[key] = {
            'l1_score': round(float(lv.get('l1_score', 0.0)), 1),
            'l2_score': round(float(lv.get('l2_score', 0.0)), 1),
            'l1_passed': bool(lv.get('l1_passed', False)),
            'l2_passed': bool(lv.get('l2_passed', False)),
            'active_level': lv.get('active_level', 'l1'),
        }

    next_locked = {
        'group_discussion': not completed.get('aptitude', False),
        'fundamentals': not completed.get('group_discussion', False),
        'basic_coding': not completed.get('fundamentals', False),
        'coding': not completed.get('basic_coding', False),
    }

    return render_template(
        'placement_track.html',
        readiness_index=round(readiness_index, 1),
        module_scores=module_scores,
        total_attempts=total_attempts,
        gd_attempts=gd_attempts,
        completed=completed,
        module_levels=module_levels,
        current_stage=state.get('current_stage', 'aptitude'),
        next_locked=next_locked,
        aptitude_l1_pass=PLACEMENT_APTITUDE_L1_PASS,
        aptitude_l2_pass=PLACEMENT_APTITUDE_L2_PASS,
        gd_l1_threshold=PLACEMENT_L1_PASS,
        gd_l2_threshold=PLACEMENT_L2_PASS,
    )


def _placement_redirect_take_quiz():
    """Placement quizzes must never be cached at the edge or browser (prevents 'same quiz for everyone')."""
    resp = redirect(url_for('take_quiz'))
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp


@app.route('/placement_track/start/<module>')
@login_required
def placement_start_module(module):
    module = (module or '').strip().lower()
    requested_level = (request.args.get('level') or 'l1').strip().lower()
    if requested_level not in ('l1', 'l2'):
        requested_level = 'l1'
    if module not in PLACEMENT_SEQUENCE:
        flash('Invalid placement module.', 'error')
        return redirect(url_for('placement_track'))

    state = _get_placement_state()
    completed = state.get('completed', {})
    if module == 'group_discussion' and not completed.get('aptitude', False):
        flash('Complete Aptitude first to unlock Group Discussion.', 'error')
        return redirect(url_for('placement_track'))
    if module == 'fundamentals' and not completed.get('group_discussion', False):
        flash('Complete Group Discussion first to unlock CS Fundamentals.', 'error')
        return redirect(url_for('placement_track'))
    if module == 'basic_coding' and not completed.get('fundamentals', False):
        flash('Complete CS Fundamentals first to unlock Basic Coding.', 'error')
        return redirect(url_for('placement_track'))
    if module == 'coding' and not completed.get('basic_coding', False):
        flash('Complete Basic Coding first to unlock Technical Coding.', 'error')
        return redirect(url_for('placement_track'))

    levels = state.get('levels', {})
    recent_signatures = _get_recent_placement_signatures(state, module, requested_level)
    level_data = levels.get(module, {})
    l1_passed = bool(level_data.get('l1_passed', False))
    if requested_level == 'l2' and not l1_passed:
        flash(
            f'Pass L1 (>= {_placement_pass_threshold(module, "l1"):.0f}%) first to unlock L2.',
            'error',
        )
        return redirect(url_for('placement_track'))
    if requested_level == 'l1' and l1_passed:
        requested_level = 'l2'

    attempt_seed = (
        f"u{current_user.id}-{module}-{requested_level}-"
        f"{int(time.time())}-{secrets.randbelow(1_000_000)}-{secrets.token_hex(8)}"
    )
    attempt_nonce = secrets.token_hex(12)

    state['current_stage'] = module
    state['levels'][module]['active_level'] = requested_level
    _save_placement_state(state)

    if module == 'aptitude':
        is_l2 = requested_level == 'l2'
        try:
            generated = _generate_placement_questions_groq(
                _placement_quiz_display_topic('aptitude', requested_level),
                'aptitude',
                10,
                level=requested_level,
                user_seed=attempt_seed,
                excluded_signatures=recent_signatures,
                user_id=current_user.id,
                attempt_nonce=attempt_nonce,
            )
            questions = generated[:10]
        except Exception as e:
            print(f"Groq aptitude generation failed, using fallback bank: {e}")
            randomized = list(PLACEMENT_APTITUDE_BANK)
            rng_fb = random.Random(int(hashlib.sha256(attempt_seed.encode('utf-8')).hexdigest()[:14], 16))
            rng_fb.shuffle(randomized)
            questions = []
            used = set(recent_signatures)
            for q in randomized:
                if not _placement_level_match(q, requested_level, 'aptitude', relax_mcq_scenario=True):
                    continue
                sig = _placement_question_signature(q)
                if sig and sig in used:
                    continue
                if sig:
                    used.add(sig)
                questions.append(q)
                if len(questions) >= 10:
                    break
            if len(questions) < 10:
                for q in randomized:
                    if q in questions:
                        continue
                    if not _placement_level_match(q, requested_level, 'aptitude', relax_mcq_scenario=True):
                        continue
                    questions.append(q)
                    if len(questions) >= 10:
                        break
            questions = questions[:10]
        _remember_placement_signatures(state, module, requested_level, questions)
        _save_placement_state(state)
        session['current_quiz'] = {
            'questions': questions,
            'topic': _placement_quiz_display_topic('aptitude', requested_level),
            'bloom_level': 3 if is_l2 else 1,
            'difficulty_level': 'advanced' if is_l2 else 'intermediate',
            'placement_module': 'aptitude',
            'placement_level': requested_level,
        }
        return _placement_redirect_take_quiz()

    if module == 'fundamentals':
        is_l2 = requested_level == 'l2'
        topic = _placement_quiz_display_topic('fundamentals', requested_level)
        questions = _generate_placement_questions_groq(
            topic, 'fundamentals', 10,
            level=requested_level,
            user_seed=attempt_seed,
            excluded_signatures=recent_signatures,
            user_id=current_user.id,
            attempt_nonce=attempt_nonce,
        )
        if not questions:
            flash('Could not generate fundamentals questions. Please try again.', 'error')
            return redirect(url_for('placement_track'))
        _remember_placement_signatures(state, module, requested_level, questions)
        _save_placement_state(state)
        session['current_quiz'] = {
            'questions': questions,
            'topic': topic,
            'bloom_level': 3 if is_l2 else 1,
            'difficulty_level': 'advanced' if is_l2 else 'intermediate',
            'placement_module': 'fundamentals',
            'placement_level': requested_level,
        }
        return _placement_redirect_take_quiz()

    if module == 'coding':
        is_l2 = requested_level == 'l2'
        topic = _placement_quiz_display_topic('coding', requested_level)
        prefer = ('medium', 'hard') if is_l2 else ('easy',)
        questions = _pick_leetcode_mix(
            excluded_signatures=recent_signatures,
            prefer=prefer,
            rng_seed=f"{attempt_seed}|placement-coding",
        )
        if is_l2:
            questions = [q for q in questions if str(q.get('difficulty', '')).lower() in ('medium', 'hard')] or questions
        else:
            questions = [q for q in questions if str(q.get('difficulty', '')).lower() == 'easy'] or questions
        questions = questions[:5]
        practice_recommendations = _pick_practice_recommendations(10)
        if not questions:
            questions = _generate_placement_questions_groq(
                topic, 'coding', 5,
                level=requested_level,
                user_seed=attempt_seed,
                excluded_signatures=recent_signatures,
                user_id=current_user.id,
                attempt_nonce=attempt_nonce,
            ) or []
            questions = [_normalize_coding_question(q) for q in questions if isinstance(q, dict)]
            practice_recommendations = _pick_practice_recommendations(10)
        if not questions:
            flash('Could not generate coding questions. Please try again.', 'error')
            return redirect(url_for('placement_track'))
        _remember_placement_signatures(state, module, requested_level, questions)
        _save_placement_state(state)
        session['current_quiz'] = {
            'questions': questions,
            'topic': topic,
            'bloom_level': 3 if is_l2 else 1,
            'difficulty_level': 'advanced' if is_l2 else 'intermediate',
            'placement_module': 'coding',
            'placement_level': requested_level,
            'practice_recommendations': practice_recommendations
        }
        return _placement_redirect_take_quiz()

    if module == 'basic_coding':
        is_l2 = requested_level == 'l2'
        topic = _placement_quiz_display_topic('basic_coding', requested_level)
        try:
            questions = _generate_basic_coding_questions_groq(
                5,
                level=requested_level,
                user_seed=attempt_seed,
                excluded_signatures=recent_signatures,
                user_id=current_user.id,
                attempt_nonce=attempt_nonce,
            )
        except Exception as e:
            print(f"Groq basic coding generation failed, fallback to easy technical mix: {e}")
            questions = _pick_leetcode_mix(
                excluded_signatures=recent_signatures,
                prefer=('easy', 'medium'),
                rng_seed=f"{attempt_seed}|placement-basic-fallback",
            )
            questions = [q for q in questions if str(q.get('difficulty', '')).lower() == 'easy'][:5] or questions[:5]
            questions = [_normalize_coding_question(q) for q in questions]
        if not questions:
            flash('Could not generate basic coding questions. Please try again.', 'error')
            return redirect(url_for('placement_track'))
        _remember_placement_signatures(state, module, requested_level, questions)
        _save_placement_state(state)
        session['current_quiz'] = {
            'questions': questions,
            'topic': topic,
            'bloom_level': 3 if is_l2 else 1,
            'difficulty_level': 'intermediate' if is_l2 else 'beginner',
            'placement_module': 'basic_coding',
            'placement_level': requested_level,
        }
        return _placement_redirect_take_quiz()

    return redirect(url_for('placement_track'))

@app.route('/placement_track/ai_interview', methods=['POST'])
@login_required
def placement_ai_interview():
    try:
        data = request.get_json(silent=True) or {}
        topic = (data.get('topic') or 'General').strip()[:150]
        level = (data.get('level') or 'l1').strip().lower()
        if level not in ('l1', 'l2'):
            level = 'l1'
        response_text = (data.get('response') or '').strip()
        if not response_text:
            return jsonify({'success': False, 'error': 'Response is required.'}), 400

        if level == 'l1':
            level_rubric = (
                "Level L1 (beginner GD): Expect clear structure, simple framing (agree/disagree), relevance to the topic, "
                "basic introduction and intelligible points. Reward communication clarity and confidence; "
                "do not heavily penalize absence of data citations or deep policy detail."
            )
        else:
            level_rubric = (
                "Level L2 (advanced GD): Expect nuanced stance, awareness of counterarguments, policy or stakeholder depth "
                "where appropriate, leadership or conflict-aware reasoning, and examples or data when natural."
            )

        prompt = f"""
You are an interview evaluator for Group Discussion and communication rounds.
{level_rubric}
Evaluate the candidate response for topic "{topic}".

Candidate response:
\"\"\"{response_text}\"\"\"

Return STRICT JSON only in this exact schema:
{{
  "score": 0-100 number,
  "feedback": "2-4 lines concise feedback",
  "speech_clarity": 0-100 number,
  "grammar": 0-100 number,
  "fluency": 0-100 number,
  "confidence": 0-100 number,
  "strengths": ["point1", "point2"],
  "improvements": ["point1", "point2"],
  "follow_up_question": "one realistic follow-up GD/interview question"
}}
"""
        try:
            payload = _placement_chat_completion(
                messages=[
                    {"role": "system", "content": "You are an interview evaluator. Output strict JSON only."},
                    {"role": "user", "content": prompt},
                ],
                temperature=0.2,
                max_tokens=1200,
                timeout=35,
            )
        except PlacementRateLimitError:
            return jsonify({'success': False, 'error': 'OpenRouter/Groq rate-limited. Retry in a few seconds.'}), 429
        raw = payload.get('content', '').strip()
        cleaned = raw.replace("```json", "").replace("```", "").strip()
        parsed = json.loads(cleaned)

        score = float(parsed.get('score', 0.0))
        score = max(0.0, min(100.0, score))
        speech_clarity = max(0.0, min(100.0, float(parsed.get('speech_clarity', score))))
        grammar = max(0.0, min(100.0, float(parsed.get('grammar', score))))
        fluency = max(0.0, min(100.0, float(parsed.get('fluency', score))))
        confidence = max(0.0, min(100.0, float(parsed.get('confidence', score))))
        feedback = str(parsed.get('feedback', '')).strip()
        follow_up_question = str(parsed.get('follow_up_question', '')).strip()
        strengths = parsed.get('strengths') or []
        improvements = parsed.get('improvements') or []

        interview = PlacementInterview(
            user_id=current_user.id,
            topic=topic or 'General',
            response_text=response_text,
            score=score,
            feedback=feedback,
            follow_up_question=follow_up_question
        )
        db.session.add(interview)
        db.session.commit()

        state = _get_placement_state()
        lv = state['levels'].setdefault('group_discussion', _default_placement_levels()['group_discussion'])
        if level == 'l1':
            lv['l1_score'] = round(score, 1)
            lv['l1_passed'] = score >= PLACEMENT_L1_PASS
            if lv['l1_passed']:
                lv['active_level'] = 'l2'
        else:
            if not lv.get('l1_passed', False):
                return jsonify({'success': False, 'error': f'Pass L1 first (>= {PLACEMENT_L1_PASS:.0f}%).'}), 400
            lv['l2_score'] = round(score, 1)
            lv['l2_passed'] = score >= PLACEMENT_L2_PASS
            lv['active_level'] = 'l2'
        state['scores']['group_discussion'] = round((float(lv.get('l1_score', 0.0)) + float(lv.get('l2_score', 0.0))) / 2.0, 1)
        state['completed']['group_discussion'] = bool(lv.get('l1_passed') and lv.get('l2_passed'))
        if state['completed']['group_discussion']:
            state['current_stage'] = 'fundamentals'
        _save_placement_state(state)

        return jsonify({
            'success': True,
            'score': round(score, 1),
            'speech_clarity': round(speech_clarity, 1),
            'grammar': round(grammar, 1),
            'fluency': round(fluency, 1),
            'confidence': round(confidence, 1),
            'feedback': feedback,
            'strengths': strengths,
            'improvements': improvements,
            'follow_up_question': follow_up_question,
            'level': level,
            'level_passed': lv['l1_passed'] if level == 'l1' else lv['l2_passed'],
            'l1_passed': lv.get('l1_passed', False),
            'l2_passed': lv.get('l2_passed', False),
            'module_cleared': state['completed']['group_discussion'],
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': f'AI interview evaluation failed: {str(e)}'}), 500

@app.route('/placement_track/suggest_topic')
@login_required
def placement_suggest_topic():
    try:
        level = (request.args.get('level') or 'l1').strip().lower()
        if level not in ('l1', 'l2'):
            level = 'l1'
        pool = PLACEMENT_GD_TOPIC_POOL.get(level) or PLACEMENT_GD_TOPIC_POOL['l1']
        fallback = random.choice(pool) if pool else "AI in education: opportunity or overdependence?"

        if level == 'l1':
            user_prompt = (
                'Give GD content points for beginners on ONE topic. Output ONLY a short topic title '
                '(max 12 words, no quotes) suitable for campus placement from themes like: simple current affairs, '
                'social issues (education, poverty), basic abstract topics, agree/disagree, communication clarity.'
            )
        else:
            user_prompt = (
                'Simulate a group discussion with arguments and counterarguments: output ONLY one challenging '
                'topic title (max 12 words, no quotes) for advanced GD: abstract debates, case-based GD, '
                'policy (AI, economy), leadership, conflict handling, data-backed argument angles.'
            )
        try:
            payload = _placement_chat_completion(
                messages=[
                    {"role": "system", "content": "Return only a short plain-text topic title, nothing else."},
                    {"role": "user", "content": user_prompt},
                ],
                temperature=0.88,
                max_tokens=48,
                timeout=20,
            )
        except Exception:
            return jsonify({'success': True, 'topic': fallback[:140], 'level': level})
        topic = payload.get('content', '').strip()
        topic = topic.replace('"', '').replace('\n', ' ').strip()
        if not topic:
            topic = fallback
        return jsonify({'success': True, 'topic': topic[:140], 'level': level})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/placement_track/transcribe_voice', methods=['POST'])
@login_required
def placement_transcribe_voice():
    try:
        data = request.get_json(silent=True) or {}
        audio_data_url = (data.get('audio_data_url') or '').strip()
        if not audio_data_url or ',' not in audio_data_url:
            return jsonify({'success': False, 'error': 'Audio data is required.'}), 400

        groq_key = os.environ.get('GROQ_API_KEY')
        if not groq_key:
            return jsonify({'success': False, 'error': 'GROQ_API_KEY is not configured.'}), 500
        transcribe_model = os.environ.get('GROQ_TRANSCRIBE_MODEL', 'whisper-large-v3-turbo')

        header, b64 = audio_data_url.split(',', 1)
        mime = 'audio/webm'
        if 'audio/wav' in header:
            mime = 'audio/wav'
        elif 'audio/mp4' in header or 'audio/m4a' in header:
            mime = 'audio/mp4'
        audio_bytes = base64.b64decode(b64)

        files = {
            'file': ('speech_input.webm', audio_bytes, mime),
        }
        payload = {
            'model': transcribe_model,
            'temperature': '0',
            'response_format': 'json'
        }
        response = requests.post(
            "https://api.groq.com/openai/v1/audio/transcriptions",
            headers={"Authorization": f"Bearer {groq_key}"},
            data=payload,
            files=files,
            timeout=45
        )
        if response.status_code == 429:
            return jsonify({'success': False, 'error': 'Groq transcription rate limit reached. Retry in a few seconds.'}), 429
        if response.status_code >= 400:
            return jsonify({'success': False, 'error': f"Transcription failed ({response.status_code})."}), 500
        transcript = (response.json().get('text') or '').strip()
        return jsonify({'success': True, 'transcript': transcript})
    except Exception as e:
        return jsonify({'success': False, 'error': f'Transcription failed: {str(e)}'}), 500


def _mock_interview_extract_resume_text(upload_file):
    if not upload_file or not upload_file.filename:
        return ''
    name = (upload_file.filename or '').lower()
    tmp = None
    try:
        suffix = '.pdf' if name.endswith('.pdf') else '.txt'
        fd, tmp = tempfile.mkstemp(suffix=suffix)
        os.close(fd)
        upload_file.save(tmp)
        if name.endswith('.pdf'):
            parts = extract_pdf_content([tmp])
            text = '\n'.join(parts) if parts else ''
        else:
            with open(tmp, 'r', encoding='utf-8', errors='ignore') as f:
                text = f.read()
        text = (text or '').strip()
        if len(text) > MOCK_INTERVIEW_RESUME_CHARS:
            text = text[:MOCK_INTERVIEW_RESUME_CHARS] + '\n[truncated]'
        return text
    finally:
        if tmp and os.path.isfile(tmp):
            try:
                os.remove(tmp)
            except OSError:
                pass


@app.route('/mock_interview')
@login_required
def mock_interview_page():
    is_teacher = getattr(current_user, 'role', 'student') == 'teacher'
    recent = (
        MockInterviewSession.query.filter_by(user_id=current_user.id)
        .order_by(MockInterviewSession.created_at.desc())
        .limit(8)
        .all()
    )
    return render_template(
        'mock_interview.html',
        recent_sessions=recent,
        is_teacher=is_teacher,
    )


@app.route('/mock_interview/start', methods=['POST'])
@login_required
def mock_interview_start():
    try:
        persona = (request.form.get('persona') or 'alex').strip().lower()
        if persona not in ('alex', 'carie'):
            persona = 'alex'
        role = (request.form.get('role') or '').strip()
        if not role:
            return jsonify({'success': False, 'error': 'Please enter the role you are practicing for.'}), 400
        jd = (request.form.get('job_description') or '').strip()
        total_questions = _mock_interview_parse_total_questions(request.form.get('question_count'))
        round_plan = _mock_interview_round_plan(total_questions)
        resume_file = request.files.get('resume')
        resume_text = _mock_interview_extract_resume_text(resume_file)
        if resume_file and resume_file.filename and not resume_text:
            return jsonify({'success': False, 'error': 'Could not read the resume file. Try PDF or plain text.'}), 400

        first_kind = round_plan[0]
        rt = MOCK_INTERVIEW_ROUND_SPECS.get(first_kind, ('Round 1', ''))[0]

        system = (
            _mock_interview_persona_voice(persona)
            + ' You output ONLY valid JSON, no markdown. Keys must match exactly.'
        )
        ctx = (
            f"Candidate target role: {role}\n"
            f"Optional company/job description:\n{jd or '(none provided)'}\n\n"
            f"Resume excerpt (may be empty):\n{resume_text or '(no resume uploaded)'}\n"
        )
        user_prompt = (
            ctx
            + f"This mock interview has exactly {total_questions} verbal questions in a structured arc "
            + "(warm-up, then technical and behavioral rounds as specified for each question).\n"
            + _mock_interview_round_instruction(first_kind)
            + '\nReturn JSON with keys: '
            '"intro" (2 short sentences you would say aloud as welcome—briefly mention warm-up / technical / behavioral flow), '
            '"question" (one clear interview question for THIS first round only; suitable to ask verbally; align with role, resume, and JD).'
        )
        data = _groq_chat_json(system, user_prompt, max_tokens=500)
        intro = (data.get('intro') or '').strip()
        question = (data.get('question') or '').strip()
        if not question:
            return jsonify({'success': False, 'error': 'AI did not return a question. Try again.'}), 500

        # Store full state in DB — large resume/JD + rounds exceed browser cookie limits for Flask sessions.
        try:
            MockInterviewState.query.filter_by(user_id=current_user.id).delete()
            db.session.commit()
        except Exception:
            db.session.rollback()
        sid = str(uuid.uuid4())
        full_state = {
            'persona': persona,
            'role': role[:255],
            'job_description': jd[:8000] if jd else '',
            'resume_excerpt': resume_text[:MOCK_INTERVIEW_RESUME_CHARS],
            'pending_question': question,
            'pending_round_kind': first_kind,
            'intro': intro,
            'rounds': [],
            'total_questions': total_questions,
            'round_plan': round_plan,
        }
        db.session.add(
            MockInterviewState(id=sid, user_id=current_user.id, state_json=json.dumps(full_state))
        )
        db.session.commit()
        session[MOCK_INTERVIEW_SESSION_KEY] = {'state_id': sid}
        session.modified = True
        return jsonify(
            {
                'success': True,
                'intro': intro,
                'question': question,
                'persona': persona,
                'total_questions': total_questions,
                'round_kind': first_kind,
                'round_title': rt,
                'teacher_preview': getattr(current_user, 'role', 'student') == 'teacher',
            }
        )
    except json.JSONDecodeError as e:
        return jsonify({'success': False, 'error': f'AI response parse error: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/mock_interview/submit_answer', methods=['POST'])
@login_required
def mock_interview_submit_answer():
    payload = request.get_json(silent=True) or {}
    answer = (payload.get('answer') or '').strip()
    if not answer:
        return jsonify({'success': False, 'error': 'Please provide your answer (type or transcribe).'}), 400

    meta = session.get(MOCK_INTERVIEW_SESSION_KEY) or {}
    sid = meta.get('state_id')
    row = (
        MockInterviewState.query.filter_by(id=sid, user_id=current_user.id).first()
        if sid
        else None
    )
    if not row:
        return jsonify({'success': False, 'error': 'No active interview. Start again from the setup page.'}), 400
    try:
        st = json.loads(row.state_json) if row.state_json else {}
    except json.JSONDecodeError:
        st = {}
    if not st.get('pending_question'):
        return jsonify({'success': False, 'error': 'No active interview. Start again from the setup page.'}), 400

    persona = st.get('persona', 'alex')
    role = st.get('role', '')
    jd = st.get('job_description', '')
    resume_excerpt = st.get('resume_excerpt', '')
    pending_q = st['pending_question']
    rounds = list(st.get('rounds') or [])
    total_questions = int(st.get('total_questions') or MOCK_INTERVIEW_DEFAULT_QUESTIONS)
    round_plan = list(st.get('round_plan') or _mock_interview_round_plan(total_questions))
    pending_kind = st.get('pending_round_kind') or (round_plan[0] if round_plan else 'warm_up')
    rt_cur, _rd_cur = MOCK_INTERVIEW_ROUND_SPECS.get(pending_kind, ('Interview', ''))

    system_eval = (
        _mock_interview_persona_voice(persona)
        + ' Evaluate interview answers fairly. Output ONLY valid JSON with keys: '
        '"articulation" (0-100 int), "domain_knowledge" (0-100 int), "relevance" (0-100 int), '
        '"communication" (0-100 int), "step_score" (0-100 int average of the four), '
        '"strengths" (array of 2-4 short strings), "improvements" (array of 2-4 short strings), '
        '"brief_feedback" (2-3 sentences for the candidate).'
    )
    user_eval = (
        f"Role: {role}\nJob description context:\n{jd or '(none)'}\n\n"
        f"Resume context:\n{resume_excerpt or '(none)'}\n\n"
        + _mock_interview_round_instruction(pending_kind)
        + f"\n\nQuestion:\n{pending_q}\n\nCandidate answer (may be from speech-to-text):\n{answer}\n"
    )
    try:
        ev = _groq_chat_json(system_eval, user_eval, max_tokens=700)
    except Exception as e:
        return jsonify({'success': False, 'error': f'Evaluation failed: {str(e)}'}), 500

    def _num(x, default=70):
        try:
            v = int(float(x))
            return max(0, min(100, v))
        except (TypeError, ValueError):
            return default

    evaluation = {
        'articulation': _num(ev.get('articulation')),
        'domain_knowledge': _num(ev.get('domain_knowledge')),
        'relevance': _num(ev.get('relevance')),
        'communication': _num(ev.get('communication')),
        'step_score': _num(ev.get('step_score')),
        'strengths': ev.get('strengths') or [],
        'improvements': ev.get('improvements') or [],
        'brief_feedback': (ev.get('brief_feedback') or '').strip(),
    }
    if not evaluation['step_score']:
        evaluation['step_score'] = (
            evaluation['articulation']
            + evaluation['domain_knowledge']
            + evaluation['relevance']
            + evaluation['communication']
        ) // 4

    rounds.append(
        {
            'question': pending_q,
            'answer': answer,
            'round_kind': pending_kind,
            'round_title': rt_cur,
            'evaluation': evaluation,
        }
    )

    done = len(rounds) >= total_questions
    next_question = None
    final_payload = None
    is_teacher = getattr(current_user, 'role', 'student') == 'teacher'

    if not done:
        next_idx = len(rounds)
        next_kind = round_plan[next_idx] if next_idx < len(round_plan) else 'technical'
        system_q = (
            _mock_interview_persona_voice(persona)
            + ' Continue the interview. Output ONLY valid JSON with key "question" (one next question, verbal, concise). '
            + 'The question MUST match the round type described in the user message.'
        )
        hist_lines = []
        for i, r in enumerate(rounds, start=1):
            rk = r.get('round_kind', '')
            rt = r.get('round_title', rk)
            hist_lines.append(
                f"Q{i} ({rt}): {r['question']}\nA{i}: {r['answer'][:1200]}"
            )
        user_q = (
            f"Role: {role}\nJD:\n{jd or '(none)'}\nResume:\n{resume_excerpt or '(none)'}\n\n"
            "Prior Q&A:\n" + '\n\n'.join(hist_lines) + '\n\n'
            + _mock_interview_round_instruction(next_kind)
            + f"\nAsk the next question only (question {next_idx + 1} of {total_questions}). "
            + 'Avoid repeating earlier questions.'
        )
        try:
            nxt = _groq_chat_json(system_q, user_q, max_tokens=400)
            next_question = (nxt.get('question') or '').strip()
        except Exception:
            next_question = ''
        if not next_question:
            done = True

    if done:
        system_final = (
            _mock_interview_persona_voice(persona)
            + ' Summarize the full mock interview. Output ONLY valid JSON with keys: '
            '"final_score" (0-100 int), "hiring_readiness" (one of: not_ready, ready, strong), '
            '"summary" (3-5 sentences), "top_strengths" (array of strings), "top_gaps" (array of strings).'
        )
        recap = []
        for i, r in enumerate(rounds, start=1):
            rt = r.get('round_title', r.get('round_kind', ''))
            recap.append(
                f"Q{i} ({rt}): {r['question']}\nA{i}: {r['answer'][:1500]}\nScores: {r['evaluation']}\n"
            )
        user_final = 'Interview recap:\n\n' + '\n'.join(recap)
        try:
            final_payload = _groq_chat_json(system_final, user_final, max_tokens=900)
        except Exception as e:
            final_payload = {
                'final_score': evaluation['step_score'],
                'hiring_readiness': 'ready',
                'summary': f'Could not generate full summary ({str(e)}). Average step score used.',
                'top_strengths': evaluation.get('strengths') or [],
                'top_gaps': evaluation.get('improvements') or [],
            }
        fs = _num(final_payload.get('final_score'), evaluation['step_score'])
        transcript = {
            'persona': persona,
            'role': role,
            'job_description': jd,
            'resume_excerpt': resume_excerpt[:2000],
            'total_questions': total_questions,
            'round_plan': round_plan,
            'rounds': rounds,
            'final': final_payload,
            'teacher_preview': is_teacher,
        }
        if not is_teacher:
            row = MockInterviewSession(
                user_id=current_user.id,
                persona=persona,
                role=role[:255],
                job_description=jd[:12000] if jd else None,
                resume_excerpt=resume_excerpt[:12000] if resume_excerpt else None,
                transcript_json=json.dumps(transcript),
                final_score=float(fs),
                final_report_json=json.dumps(final_payload),
                is_preview=False,
                question_count=total_questions,
            )
            db.session.add(row)
            db.session.commit()
        try:
            MockInterviewState.query.filter_by(id=sid).delete()
            db.session.commit()
        except Exception:
            db.session.rollback()
        session.pop(MOCK_INTERVIEW_SESSION_KEY, None)
        return jsonify(
            {
                'success': True,
                'complete': True,
                'evaluation': evaluation,
                'final_score': fs,
                'final_report': final_payload,
                'saved': not is_teacher,
                'teacher_preview': is_teacher,
            }
        )

    next_kind = round_plan[len(rounds)] if len(rounds) < len(round_plan) else 'technical'
    rt_next = MOCK_INTERVIEW_ROUND_SPECS.get(next_kind, ('Interview', ''))[0]

    st['rounds'] = rounds
    st['pending_question'] = next_question
    st['pending_round_kind'] = next_kind
    row.state_json = json.dumps(st)
    db.session.commit()
    return jsonify(
        {
            'success': True,
            'complete': False,
            'evaluation': evaluation,
            'question_index': len(rounds),
            'total_questions': total_questions,
            'next_question': next_question,
            'round_kind': next_kind,
            'round_title': rt_next,
        }
    )


def generate_quiz_code(length=6):
    import random, string
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=length))

def require_teacher():
    if getattr(current_user, 'role', 'student') != 'teacher':
        flash('Teacher access required', 'error')
        return redirect(url_for('dashboard'))
    return None

# Teacher: create quiz (form + post)
@app.route('/teacher/quiz/new', methods=['GET', 'POST'])
@login_required
def teacher_create_quiz():
    guard = require_teacher()
    if guard:
        return guard

    if request.method == 'POST':
        try:
            title = request.form.get('title', '').strip()
            questions_raw = request.form.get('questions_json', '').strip()
            if not title or not questions_raw:
                flash('Title and questions are required', 'error')
                return redirect(url_for('teacher_create_quiz'))

            # parse questions json
            questions = json.loads(questions_raw)
            # generate unique code
            code = generate_quiz_code()
            while db.session.query(Quiz).filter_by(code=code).first() is not None:
                code = generate_quiz_code()

            quiz = Quiz(title=title, code=code, created_by=current_user.id)
            db.session.add(quiz)
            db.session.flush()  # get quiz.id

            for q in questions:
                qtype = q.get('type', 'mcq')
                opts = q.get('options', []) if qtype == 'mcq' else []
                qq = QuizQuestion(
                    quiz_id=quiz.id,
                    question=q.get('question', ''),
                    options_json=json.dumps(opts) if opts else None,
                    answer=q.get('answer', ''),
                    qtype=qtype,
                    marks=int(q.get('marks', 1)),
                    image_url=q.get('image_url')
                )
                db.session.add(qq)

            db.session.commit()
            flash(f'Quiz created! Share code: {code}', 'success')
            return redirect(url_for('dashboard'))
        except Exception as e:
            db.session.rollback()
            print(f"Error creating quiz: {str(e)}")
            flash('Error creating quiz. Ensure valid JSON.', 'error')
            return redirect(url_for('teacher_create_quiz'))

    return render_template('create_quiz.html')

@app.route('/teacher/quiz/<code>/archive', methods=['POST'])
@login_required
def teacher_archive_quiz(code):
    guard = require_teacher()
    if guard:
        return guard

    quiz = db.session.query(Quiz).filter_by(code=code.upper(), created_by=current_user.id).first()
    if not quiz:
        flash('Quiz not found or permission denied.', 'error')
        return redirect(url_for('dashboard'))

    try:
        action = request.form.get('action', 'archive').strip().lower()
        quiz.is_archived = (action != 'unarchive')
        db.session.commit()
        flash('Quiz archived successfully.' if quiz.is_archived else 'Quiz moved back to active list.', 'success')
    except Exception as e:
        print(f"Archive toggle error for quiz {code}: {e}")
        db.session.rollback()
        flash('Could not update archive status.', 'error')

    # Keep current list mode after action.
    show_archived = '1' if request.form.get('show_archived') == '1' else '0'
    return redirect(url_for('dashboard', show_archived=show_archived))

# Teacher: simple create by topic and number of questions
@app.route('/teacher/quiz/new_simple', methods=['GET', 'POST'])
@login_required
def teacher_create_quiz_simple():
    guard = require_teacher()
    if guard:
        return guard

    if request.method == 'POST':
        try:
            topic = request.form.get('topic', '').strip()
            count = int(request.form.get('count', '0') or 0)
            title = request.form.get('title', '').strip()
            difficulty = request.form.get('difficulty', 'beginner').strip()
            marks = int(request.form.get('marks', '1') or 1)
            duration = request.form.get('duration', '').strip()
            duration_minutes = int(duration) if duration else None
            question_type = request.form.get('question_type', 'mcq').strip()  # Get question type

            if count <= 0:
                flash('Please provide number of questions (>0).', 'error')
                return redirect(url_for('teacher_create_quiz_simple'))

            # Handle PDF file uploads (supports OCR for scanned PDFs)
            pdf_content = None
            pdf_file_paths = []
            
            # Check for PDF file uploads (supports multiple files)
            if 'notes_pdf' in request.files:
                files = request.files.getlist('notes_pdf')
                for file in files:
                    if file and file.filename and file.filename.lower().endswith('.pdf'):
                        try:
                            import tempfile
                            import os
                            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
                                file.save(tmp_file.name)
                                pdf_file_paths.append(tmp_file.name)
                                print(f'Saved PDF: {file.filename} to {tmp_file.name}')
                        except Exception as e:
                            print(f'Error saving PDF {file.filename}: {str(e)}')
                            continue
            
            # Extract content from PDFs (with OCR support for scanned PDFs)
            if pdf_file_paths:
                try:
                    pdf_content = extract_pdf_content(pdf_file_paths)
                    if pdf_content:
                        # Extract a topic from PDF for tracking purposes (optional)
                        if not topic:
                            # Try to extract topic from first PDF filename
                            import os
                            first_pdf_name = os.path.basename(pdf_file_paths[0])
                            topic = extract_topic_from_filename(pdf_file_paths[0]) or "PDF Content"
                        
                        pdf_info = f'Successfully processed {len(pdf_file_paths)} PDF file(s). Questions will be generated from PDF content.'
                        if OCR_AVAILABLE:
                            pdf_info += ' (Local OCR + Cloud OCR enabled for scanned PDFs)'
                        else:
                            pdf_info += ' (Cloud OCR enabled for scanned PDFs)'
                        flash(pdf_info, 'success')
                    else:
                        flash('Could not extract content from PDF(s). Questions will be generated from topic only.', 'warning')
                        # Clean up temp files
                        import os
                        for path in pdf_file_paths:
                            try:
                                os.unlink(path)
                            except:
                                pass
                except Exception as e:
                    flash(f'Error processing PDF(s): {str(e)}. Questions will be generated from topic only.', 'warning')
                    # Clean up temp files
                    import os
                    for path in pdf_file_paths:
                        try:
                            os.unlink(path)
                        except:
                            pass
            
            # If no PDFs and no topic, require topic
            if not pdf_content and not topic:
                flash('Please either enter a topic OR upload PDF file(s) to generate questions from.', 'error')
                return redirect(url_for('teacher_create_quiz_simple'))

            questions = generate_quiz(topic or "PDF Content", difficulty if difficulty in ['beginner','intermediate','advanced'] else 'beginner', question_type, count, pdf_content) or []
            
            # Clean up temporary PDF files after question generation
            if pdf_file_paths:
                import os
                for path in pdf_file_paths:
                    try:
                        os.unlink(path)
                    except Exception as cleanup_error:
                        print(f"Warning: Could not delete temp file {path}: {cleanup_error}")
            if not questions:
                flash('Failed to generate questions. Try again.', 'error')
                return redirect(url_for('teacher_create_quiz_simple'))

            if not title:
                title = f"{topic} Quiz"

            # Default marks fill
            for q in questions:
                q['marks'] = marks

            # Store in session for preview
            session['preview_quiz'] = {
                'title': title,
                'topic': topic,
                'difficulty': difficulty,
                'duration_minutes': duration_minutes,
                'question_type': question_type,
                'questions': questions
            }
            return render_template('preview_quiz.html', data=session['preview_quiz'])
        except Exception as e:
            db.session.rollback()
            print(f"Error creating simple quiz: {str(e)}")
            flash('Error creating quiz. Please try again.', 'error')
            return redirect(url_for('teacher_create_quiz_simple'))

    return render_template('create_quiz_simple.html')

# Extract questions from PDFs based on keywords
@app.route('/teacher/quiz/extract_questions', methods=['POST'])
@login_required
def extract_questions_from_pdf():
    guard = require_teacher()
    if guard:
        return jsonify({'success': False, 'error': 'Teacher access required'}), 403
    
    try:
        # Get PDF files and keywords
        pdf_files = request.files.getlist('pdf_files')
        keywords = request.form.get('keywords', '').strip()
        
        if not pdf_files or not any(f.filename for f in pdf_files):
            return jsonify({'success': False, 'error': 'Please select at least one PDF file'})
        
        if not keywords:
            return jsonify({'success': False, 'error': 'Please enter keywords or a sentence'})
        
        # Save PDF files temporarily
        pdf_file_paths = []
        for file in pdf_files:
            if file and file.filename and file.filename.lower().endswith('.pdf'):
                try:
                    import tempfile
                    import os
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
                        file.save(tmp_file.name)
                        pdf_file_paths.append(tmp_file.name)
                except Exception as e:
                    print(f'Error saving PDF {file.filename}: {str(e)}')
                    continue
        
        if not pdf_file_paths:
            return jsonify({'success': False, 'error': 'No valid PDF files were uploaded'})
        
        # Extract content from PDFs
        pdf_content = extract_pdf_content(pdf_file_paths)
        
        if not pdf_content:
            # Clean up temp files
            import os
            for path in pdf_file_paths:
                try:
                    os.unlink(path)
                except:
                    pass
            return jsonify({'success': False, 'error': 'Could not extract content from PDF(s). Please ensure the PDFs contain readable text.'})
        
        # Use AI to extract questions based on keywords
        response_text = ""
        try:
            # Use gemini-2.5-flash (current free tier) with fallback to gemini-2.5-flash-lite
            try:
                model = genai.GenerativeModel("gemini-2.5-flash")
            except:
                try:
                    model = genai.GenerativeModel("gemini-2.5-flash-lite")
                except:
                    try:
                        model = genai.GenerativeModel("gemini-1.5-flash")
                    except:
                        raise Exception("No working Gemini model found. Check API key and quota.")
            
            # Truncate PDF content if too long
            if len(pdf_content) > 15000:
                pdf_content = pdf_content[:5000] + "\n\n[... content truncated ...]\n\n" + pdf_content[-10000:]
            
            prompt = f"""You are an expert at extracting relevant questions from educational content.

PDF CONTENT:
{pdf_content}

KEYWORDS/TOPICS TO FOCUS ON:
{keywords}

Based on the PDF content above, extract questions that are directly related to the keywords/topics provided. 

IMPORTANT REQUIREMENTS:
1. Extract questions that are SPECIFICALLY related to the keywords: "{keywords}"
2. Questions must be based ONLY on the information present in the PDF content
3. Extract a variety of question types (both multiple choice and subjective)
4. For each question, provide:
   - The question text
   - For MCQ: 4 options (A, B, C, D) and the correct answer
   - For subjective: A sample answer or key points
   - Question type (mcq or subjective)

Return the questions in the following JSON format:
{{
  "questions": [
    {{
      "question": "Question text here",
      "type": "mcq",
      "options": ["Option A", "Option B", "Option C", "Option D"],
      "answer": "A"
    }},
    {{
      "question": "Question text here",
      "type": "subjective",
      "answer": "Sample answer or key points"
    }}
  ]
}}

Extract 5-10 questions that are most relevant to the keywords provided. Focus on quality over quantity."""

            response = model.generate_content(prompt)
            response_text = response.text.strip()
            
            # Parse JSON from response (handle markdown code blocks)
            import json
            import re
            
            # Remove markdown code blocks if present
            json_match = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', response_text, re.DOTALL)
            if json_match:
                response_text = json_match.group(1)
            else:
                # Try to find JSON object directly
                json_match = re.search(r'\{.*\}', response_text, re.DOTALL)
                if json_match:
                    response_text = json_match.group(0)
            
            result = json.loads(response_text)
            questions = result.get('questions', [])
            
            # Clean up temp files
            import os
            for path in pdf_file_paths:
                try:
                    os.unlink(path)
                except:
                    pass
            
            if questions:
                return jsonify({
                    'success': True,
                    'questions': questions,
                    'keywords': keywords,
                    'pdf_count': len(pdf_file_paths)
                })
            else:
                return jsonify({'success': False, 'error': 'No questions could be extracted. Please try different keywords.'})
                
        except json.JSONDecodeError as e:
            print(f"JSON parsing error: {e}")
            if response_text:
                print(f"Response text: {response_text[:500]}")
            # Clean up temp files
            import os
            for path in pdf_file_paths:
                try:
                    os.unlink(path)
                except:
                    pass
            return jsonify({'success': False, 'error': 'Failed to parse AI response. Please try again.'})
        except Exception as e:
            print(f"Error extracting questions: {e}")
            import traceback
            traceback.print_exc()
            # Clean up temp files
            import os
            for path in pdf_file_paths:
                try:
                    os.unlink(path)
                except:
                    pass
            return jsonify({'success': False, 'error': f'Error extracting questions: {str(e)}'})
            
    except Exception as e:
        print(f"Error in extract_questions_from_pdf: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'Unexpected error: {str(e)}'})

# Preview step for teacher to adjust marks before finalizing
@app.route('/teacher/quiz/preview', methods=['POST'])
@login_required
def teacher_quiz_preview():
    guard = require_teacher()
    if guard:
        return guard
    try:
        # Rebuild inputs
        title = request.form.get('title', '').strip()
        topic = request.form.get('topic', '').strip()
        count = int(request.form.get('count', '0') or 0)
        difficulty = request.form.get('difficulty', 'beginner').strip()
        marks = int(request.form.get('marks', '1') or 1)
        duration = request.form.get('duration', '').strip()
        duration_minutes = int(duration) if duration else None

        if not topic or count <= 0:
            flash('Provide topic and number of questions.', 'error')
            return redirect(url_for('teacher_create_quiz_simple'))

        questions = generate_quiz(topic, difficulty if difficulty in ['beginner','intermediate','advanced'] else 'beginner', 'mcq', count) or []
        if not questions:
            flash('Failed to generate questions.', 'error')
            return redirect(url_for('teacher_create_quiz_simple'))

        # Default marks fill
        for q in questions:
            q['marks'] = marks

        # Store in session for finalize
        session['preview_quiz'] = {
            'title': title or f"{topic} Quiz",
            'topic': topic,
            'difficulty': difficulty,
            'duration_minutes': duration_minutes,
            'questions': questions
        }
        return render_template('preview_quiz.html', data=session['preview_quiz'])
    except Exception as e:
        print(f"Preview error: {e}")
        flash('Error preparing preview.', 'error')
        return redirect(url_for('teacher_create_quiz_simple'))

@app.route('/teacher/quiz/finalize', methods=['POST'])
@login_required
def teacher_quiz_finalize():
    guard = require_teacher()
    if guard:
        return guard
    data = session.get('preview_quiz')
    if not data:
        flash('No quiz in preview.', 'error')
        return redirect(url_for('teacher_create_quiz_simple'))
    try:
        # Read marks overrides
        q_overrides = []
        for i, q in enumerate(data['questions']):
            new_marks = request.form.get(f'marks_{i}')
            try:
                q['marks'] = int(new_marks)
            except Exception:
                pass
            q_overrides.append(q)

        code = generate_quiz_code()
        while db.session.query(Quiz).filter_by(code=code).first() is not None:
            code = generate_quiz_code()

        quiz = Quiz(title=data['title'], code=code, created_by=current_user.id, difficulty=data['difficulty'], duration_minutes=data['duration_minutes'])
        db.session.add(quiz)
        db.session.flush()

        for q in q_overrides:
            qtype = q.get('type', data.get('question_type', 'mcq'))
            opts = q.get('options', [])
            
            qq = QuizQuestion(
                quiz_id=quiz.id,
                question=q.get('question', ''),
                options_json=json.dumps(opts) if opts else None,
                answer=q.get('answer', ''),
                qtype=qtype,
                marks=int(q.get('marks', 1)),
                image_url=q.get('image_url')
            )
            
            # Handle coding questions
            if qtype == 'coding':
                qq.test_cases_json = json.dumps(q.get('test_cases', []))
                qq.language_constraints = json.dumps(q.get('language_constraints', ['python', 'java', 'cpp', 'c']))
                qq.time_limit_seconds = q.get('time_limit_seconds', 2)
                qq.memory_limit_mb = q.get('memory_limit_mb', 256)
                qq.sample_input = q.get('sample_input', '')
                qq.sample_output = q.get('sample_output', '')
                qq.starter_code = json.dumps(q.get('starter_code', {}))
            
            db.session.add(qq)

        db.session.commit()
        session.pop('preview_quiz', None)
        flash(f'Quiz created! Share code: {code}', 'success')
        return redirect(url_for('dashboard'))
    except Exception as e:
        db.session.rollback()
        import traceback
        error_details = str(e)
        traceback.print_exc()
        print(f"Finalize error: {error_details}")
        # Check if it's a database column error
        if 'no column named' in error_details.lower() or 'column' in error_details.lower():
            flash(f'Database migration needed! Error: {error_details}. Please run: python migrate_new_features.py', 'error')
        else:
            flash(f'Error finalizing quiz: {error_details}', 'error')
        return redirect(url_for('teacher_create_quiz_simple'))

# Student: join quiz by code
@app.route('/quiz/join', methods=['GET', 'POST'])
@login_required
def join_quiz():
    if request.method == 'POST':
        code = request.form.get('code', '').strip().upper()
        quiz = db.session.query(Quiz).filter_by(code=code).first()
        if not quiz:
            flash('Invalid quiz code', 'error')
            return redirect(url_for('join_quiz'))
        return redirect(url_for('take_shared_quiz', code=code))
    return render_template('join_quiz.html')

# Take shared quiz
@app.route('/quiz/take/<code>')
@login_required
def take_shared_quiz(code):
    quiz = db.session.query(Quiz).filter_by(code=code.upper()).first()
    if not quiz:
        flash('Quiz not found', 'error')
        return redirect(url_for('join_quiz'))
    
    # Check if student has already completed this quiz
    existing_completed = db.session.query(QuizSubmission).filter_by(
        quiz_id=quiz.id, 
        student_id=current_user.id, 
        completed=True
    ).first()
    
    if existing_completed:
        flash('You have already attempted this quiz. You can only take it once.', 'error')
        return redirect(url_for('dashboard'))
    
    q_rows = db.session.query(QuizQuestion).filter_by(quiz_id=quiz.id).all()
    # Parse options JSON server-side to avoid template errors
    parsed_questions = []
    for q in q_rows:
        try:
            options = json.loads(q.options_json) if q.options_json else []
        except Exception:
            options = []
        
        question_data = {
            'id': q.id,
            'question': q.question,
            'qtype': q.qtype,
            'marks': q.marks,
            'image_url': getattr(q, 'image_url', None),
            'options': options,
        }
        
        # Add coding question data
        if q.qtype == 'coding':
            try:
                question_data['test_cases'] = json.loads(q.test_cases_json) if q.test_cases_json else []
                question_data['language_constraints'] = json.loads(q.language_constraints) if q.language_constraints else ['python', 'java', 'cpp', 'c']
                question_data['time_limit_seconds'] = q.time_limit_seconds or 2
                question_data['memory_limit_mb'] = q.memory_limit_mb or 256
                question_data['sample_input'] = q.sample_input or ''
                question_data['sample_output'] = q.sample_output or ''
                question_data['starter_code'] = json.loads(q.starter_code) if q.starter_code else {}
            except Exception as e:
                print(f"Error parsing coding question data: {e}")
                question_data['test_cases'] = []
                question_data['language_constraints'] = ['python', 'java', 'cpp', 'c']
        
        parsed_questions.append(question_data)
    
    # Ensure a started submission exists (one per student/quiz if not completed)
    existing = db.session.query(QuizSubmission).filter_by(quiz_id=quiz.id, student_id=current_user.id, completed=False).first()
    if not existing:
        existing = QuizSubmission(quiz_id=quiz.id, student_id=current_user.id, question_count=len(q_rows))
        db.session.add(existing)
        db.session.commit()
    return render_template('take_shared_quiz.html', quiz=quiz, questions=parsed_questions)

# Proctoring: record snapshot
@app.route('/api/proctoring/snapshot/<code>', methods=['POST'])
@login_required
def api_proctoring_snapshot(code):
    quiz = db.session.query(Quiz).filter_by(code=code.upper()).first()
    if not quiz:
        return jsonify({'error': 'Quiz not found'}), 404
    submission = db.session.query(QuizSubmission).filter_by(quiz_id=quiz.id, student_id=current_user.id, completed=False).first()
    if not submission:
        return jsonify({'error': 'No active attempt'}), 400
    data = request.get_json(silent=True) or {}
    snapshot_type = (data.get('type') or 'webcam').lower()
    if snapshot_type not in ('screen', 'webcam'):
        snapshot_type = 'webcam'
    image_data = data.get('image')
    if not image_data:
        return jsonify({'ok': False, 'error': 'missing_image'}), 400
    captured_at = datetime.utcnow()
    if data.get('captured_at'):
        try:
            captured_at = datetime.fromisoformat(str(data['captured_at']).replace('Z', '+00:00'))
        except Exception:
            pass
    device_fp = data.get('device_fingerprint')
    if device_fp and not submission.device_fingerprint:
        submission.device_fingerprint = device_fp[:512]
    elif device_fp and submission.device_fingerprint and submission.device_fingerprint != device_fp[:512]:
        db.session.add(ProctoringBreach(submission_id=submission.id, breach_type='DEVICE_ID_CHANGE_DETECTED', occurred_at=captured_at))
        submission.device_fingerprint = device_fp[:512]
    # Prevent very large payloads from overwhelming DB/storage.
    if image_data and len(image_data) > 2500000:
        image_data = None
    stored_image_ref = persist_proctoring_image(image_data, submission.id, snapshot_type, captured_at)
    try:
        db.session.add(ProctoringSnapshot(submission_id=submission.id, snapshot_type=snapshot_type, image_data=stored_image_ref, captured_at=captured_at))
        db.session.commit()
        return jsonify({'ok': True, 'stored': bool(stored_image_ref)})
    except Exception as e:
        db.session.rollback()
        print(f"Snapshot persist failed for submission {submission.id}: {e}")
        return jsonify({'ok': False, 'error': 'snapshot_store_failed'}), 500

# Proctoring: record breach
@app.route('/api/proctoring/breach/<code>', methods=['POST'])
@login_required
def api_proctoring_breach(code):
    quiz = db.session.query(Quiz).filter_by(code=code.upper()).first()
    if not quiz:
        return jsonify({'error': 'Quiz not found'}), 404
    submission = db.session.query(QuizSubmission).filter_by(quiz_id=quiz.id, student_id=current_user.id, completed=False).first()
    if not submission:
        return jsonify({'error': 'No active attempt'}), 400
    data = request.get_json(silent=True) or {}
    breach_type = (data.get('breach_type') or data.get('type') or 'UNKNOWN').upper().replace(' ', '_')[:80]
    occurred_at = datetime.utcnow()
    if data.get('occurred_at'):
        try:
            occurred_at = datetime.fromisoformat(str(data['occurred_at']).replace('Z', '+00:00'))
        except Exception:
            pass
    db.session.add(ProctoringBreach(submission_id=submission.id, breach_type=breach_type, occurred_at=occurred_at))
    db.session.commit()
    return jsonify({'ok': True})

# Submit shared quiz
@app.route('/quiz/submit/<code>', methods=['POST'])
@login_required
def submit_shared_quiz(code):
    quiz = db.session.query(Quiz).filter_by(code=code.upper()).first()
    if not quiz:
        flash('Quiz not found', 'error')
        return redirect(url_for('join_quiz'))
    
    # Check if student has already completed this quiz
    existing_completed = db.session.query(QuizSubmission).filter_by(
        quiz_id=quiz.id, 
        student_id=current_user.id, 
        completed=True
    ).first()
    
    if existing_completed:
        flash('You have already attempted this quiz. You can only take it once.', 'error')
        return redirect(url_for('dashboard'))
    
    questions = db.session.query(QuizQuestion).filter_by(quiz_id=quiz.id).all()

    total_marks = 0.0
    scored_marks = 0.0
    submission = db.session.query(QuizSubmission).filter_by(quiz_id=quiz.id, student_id=current_user.id, completed=False).first()
    if not submission:
        submission = QuizSubmission(quiz_id=quiz.id, student_id=current_user.id)
        db.session.add(submission)
        db.session.flush()

    answered_count = 0
    for q in questions:
        total_marks += float(q.marks or 1)
        key = f'q_{q.id}'
        user_ans = request.form.get(key, '').strip()
        is_correct = None
        ai_score = None
        gained = 0.0
        code_language = None
        test_results_json = None
        passed_test_cases = 0
        total_test_cases = 0

        if q.qtype == 'mcq':
            is_correct = (user_ans.split('. ')[0] == (q.answer or '')) if user_ans else False
            gained = float(q.marks or 1) if is_correct else 0.0
        elif q.qtype == 'coding':
            # Handle coding question submission
            code_data = request.form.get(f'code_{q.id}', '').strip()
            language = request.form.get(f'language_{q.id}', 'python')
            
            if code_data:
                try:
                    test_cases = json.loads(q.test_cases_json) if q.test_cases_json else []
                    time_limit = q.time_limit_seconds or 2
                    memory_limit = q.memory_limit_mb or 256
                    
                    test_results = run_test_cases(code_data, language, test_cases, time_limit, memory_limit)
                    
                    passed_test_cases = test_results['passed']
                    total_test_cases = test_results['total']
                    percentage_passed = test_results['percentage'] / 100.0
                    
                    gained = float(q.marks or 1) * percentage_passed
                    is_correct = percentage_passed == 1.0
                    code_language = language
                    test_results_json = json.dumps(test_results['results'])
                    user_ans = code_data
                except Exception as e:
                    print(f"Error evaluating coding question: {e}")
                    gained = 0.0
                    is_correct = False
                    code_language = language
                    test_results_json = json.dumps([])
                    user_ans = code_data
        else:
            # subjective via AI
            if user_ans:
                ai_score = evaluate_subjective_answer(q.question, user_ans, q.answer or '')
                gained = float(q.marks or 1) * float(ai_score or 0.0)
                is_correct = (ai_score or 0.0) >= 0.6
            else:
                ai_score = 0.0
                is_correct = False

        scored_marks += gained
        ans = QuizAnswer(
            submission_id=submission.id,
            question_id=q.id,
            user_answer=user_ans,
            is_correct=is_correct,
            ai_score=ai_score,
            scored_marks=gained,
            code_language=code_language,
            test_results_json=test_results_json,
            passed_test_cases=passed_test_cases,
            total_test_cases=total_test_cases
        )
        db.session.add(ans)
        if user_ans:
            answered_count += 1

    percentage = (scored_marks / total_marks) * 100 if total_marks > 0 else 0
    passed = percentage >= 60
    submission.score = scored_marks
    submission.total = total_marks
    submission.percentage = percentage
    submission.passed = passed
    # set review unlock time 15 minutes after submission
    submission.review_unlocked_at = datetime.utcnow() + timedelta(minutes=15)
    # Get violation flags from form (if submitted normally)
    submission.fullscreen_exit_flag = request.form.get('fullscreen_exit') == 'true'
    submission.alt_tab_flag = request.form.get('alt_tab_flag') == 'true'
    submission.win_shift_s_flag = request.form.get('win_shift_s_flag') == 'true'
    submission.win_prtscn_flag = request.form.get('win_prtscn_flag') == 'true'
    submission.prtscn_flag = request.form.get('prtscn_flag') == 'true'
    # Tab switch also counts as alt_tab
    if request.form.get('tab_switch_flag') == 'true':
        submission.alt_tab_flag = True
    
    # Check if any violation occurred
    has_violation = (submission.fullscreen_exit_flag or submission.alt_tab_flag or 
                     submission.win_shift_s_flag or submission.win_prtscn_flag or 
                     submission.prtscn_flag)
    
    submission.answered_count = answered_count
    submission.question_count = len(questions)
    submission.is_full_completion = (answered_count == len(questions)) and (not has_violation)
    submission.completed = True
    db.session.commit()

    flash(f'Submitted. Score: {scored_marks:.1f}/{total_marks} ({percentage:.0f}%).', 'success')
    return redirect(url_for('dashboard'))

# Auto-submit partial answers on fullscreen exit or tab close
@app.route('/quiz/auto_submit/<code>', methods=['POST'])
@login_required
def auto_submit_partial(code):
    try:
        quiz = db.session.query(Quiz).filter_by(code=code.upper()).first()
        if not quiz:
            return ('', 204)
        questions = db.session.query(QuizQuestion).filter_by(quiz_id=quiz.id).all()
        submission = db.session.query(QuizSubmission).filter_by(quiz_id=quiz.id, student_id=current_user.id, completed=False).first()
        if not submission:
            submission = QuizSubmission(quiz_id=quiz.id, student_id=current_user.id)
            db.session.add(submission)
            db.session.flush()

        total_marks = 0.0
        scored_marks = 0.0
        answered_count = 0
        data = request.get_json(silent=True) or {}
        
        # Fallback: manually parse JSON if get_json didn't work (for sendBeacon)
        if not data and request.data:
            try:
                import json
                data = json.loads(request.data.decode('utf-8'))
                print(f"DEBUG: Manually parsed JSON from request.data")
            except Exception as e:
                print(f"DEBUG: Failed to parse JSON manually: {e}")
                data = {}
        
        print(f"DEBUG: Received data in auto_submit_partial: {data}")
        print(f"DEBUG: Data type: {type(data)}")
        print(f"DEBUG: Request content type: {request.content_type}")
        if request.data:
            print(f"DEBUG: Request data (first 200 chars): {request.data[:200]}")

        for q in questions:
            total_marks += float(q.marks or 1)
            key = f'q_{q.id}'
            user_ans = (data.get(key) or '').strip()
            is_correct = None
            ai_score = None
            gained = 0.0
            if q.qtype == 'mcq':
                is_correct = (user_ans.split('. ')[0] == (q.answer or '')) if user_ans else False
                gained = float(q.marks or 1) if is_correct else 0.0
            else:
                if user_ans:
                    ai_score = evaluate_subjective_answer(q.question, user_ans, q.answer or '')
                    gained = float(q.marks or 1) * float(ai_score or 0.0)
                    is_correct = (ai_score or 0.0) >= 0.6
                else:
                    ai_score = 0.0
                    is_correct = False
            if user_ans:
                answered_count += 1
            # Upsert answer
            existing_ans = db.session.query(QuizAnswer).filter_by(submission_id=submission.id, question_id=q.id).first()
            if existing_ans:
                existing_ans.user_answer = user_ans
                existing_ans.is_correct = is_correct
                existing_ans.ai_score = ai_score
                existing_ans.scored_marks = gained
            else:
                db.session.add(QuizAnswer(
                    submission_id=submission.id,
                    question_id=q.id,
                    user_answer=user_ans,
                    is_correct=is_correct,
                    ai_score=ai_score,
                    scored_marks=gained
                ))

        submission.score = scored_marks
        submission.total = total_marks
        submission.percentage = (scored_marks / total_marks) * 100 if total_marks > 0 else 0
        submission.passed = submission.percentage >= 60
        from datetime import timedelta
        submission.review_unlocked_at = datetime.utcnow() + timedelta(minutes=15)
        # Set violation flags from request data
        violation_flags = data.get('violation_flags', {})
        print(f"DEBUG: Received violation_flags: {violation_flags}")
        print(f"DEBUG: violation_flags type: {type(violation_flags)}")
        if violation_flags:
            print(f"DEBUG: alt_tab value: {violation_flags.get('alt_tab')}, tab_switch: {violation_flags.get('tab_switch')}")
        submission.fullscreen_exit_flag = True  # Auto-submit always means fullscreen exit
        # Convert to boolean explicitly - tab_switch also counts as alt_tab
        alt_tab_violation = bool(violation_flags.get('alt_tab', False)) or bool(violation_flags.get('tab_switch', False)) if violation_flags else False
        submission.alt_tab_flag = alt_tab_violation
        submission.win_shift_s_flag = bool(violation_flags.get('win_shift_s', False)) if violation_flags else False
        submission.win_prtscn_flag = bool(violation_flags.get('win_prtscn', False)) if violation_flags else False
        submission.prtscn_flag = bool(violation_flags.get('prtscn', False)) if violation_flags else False
        print(f"DEBUG: Set flags - alt_tab: {submission.alt_tab_flag}, win_shift_s: {submission.win_shift_s_flag}, win_prtscn: {submission.win_prtscn_flag}, prtscn: {submission.prtscn_flag}")
        
        submission.answered_count = answered_count
        submission.question_count = len(questions)
        submission.is_full_completion = False
        submission.completed = True
        db.session.commit()
        return ('', 204)
    except Exception as e:
        db.session.rollback()
        return ('', 204)

# Teacher: view results
@app.route('/teacher/quiz/<code>/results')
@login_required
def teacher_quiz_results(code):
    guard = require_teacher()
    if guard:
        return guard
    quiz = db.session.query(Quiz).filter_by(code=code.upper(), created_by=current_user.id).first()
    if not quiz:
        flash('Quiz not found', 'error')
        return redirect(url_for('dashboard'))
    submissions = db.session.query(QuizSubmission).filter_by(quiz_id=quiz.id).order_by(QuizSubmission.submitted_at.desc()).all()
    # Join with users
    student_map = {u.id: u for u in db.session.query(User).filter(User.id.in_([s.student_id for s in submissions])).all()}
    return render_template('teacher_results.html', quiz=quiz, submissions=submissions, student_map=student_map)

# Teacher: download results as CSV or Excel
@app.route('/teacher/quiz/<code>/results/download/<format>')
@login_required
def download_quiz_results(code, format):
    guard = require_teacher()
    if guard:
        return guard
    quiz = db.session.query(Quiz).filter_by(code=code.upper(), created_by=current_user.id).first()
    if not quiz:
        flash('Quiz not found', 'error')
        return redirect(url_for('dashboard'))
    submissions = db.session.query(QuizSubmission).filter_by(quiz_id=quiz.id).order_by(QuizSubmission.submitted_at.desc()).all()
    student_map = {u.id: u for u in db.session.query(User).filter(User.id.in_([s.student_id for s in submissions])).all()}
    header = [
        'Student', 'Score', 'Total', 'Percentage', 'Passed',
        'Integrity', 'Violation Count', 'Submitted At'
    ]
    rows = []
    for s in submissions:
        student = student_map.get(s.student_id)
        name = student.username if student else f'ID {s.student_id}'
        violation_count = int(bool(s.fullscreen_exit_flag)) + int(bool(s.alt_tab_flag)) + int(bool(s.win_shift_s_flag)) + int(bool(s.win_prtscn_flag)) + int(bool(s.prtscn_flag))
        integrity_status = 'Flagged' if (s.marked_as_cheating or violation_count > 0 or not s.is_full_completion) else 'Clean'
        rows.append([
            name,
            f'{s.score:.1f}',
            f'{s.total:.1f}',
            f'{s.percentage:.1f}',
            'Yes' if s.passed else 'No',
            integrity_status,
            violation_count,
            s.submitted_at.strftime('%Y-%m-%d %H:%M') if s.submitted_at else ''
        ])
    fn = f'{quiz.code}_results'
    fmt = (format or 'csv').lower()
    if fmt == 'xlsx':
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Results'
        for c, h in enumerate(header, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        for r, row in enumerate(rows, 2):
            for c, v in enumerate(row, 1):
                ws.cell(row=r, column=c, value=v)
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name=f'{fn}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(header)
    writer.writerows(rows)
    return send_file(io.BytesIO(output.getvalue().encode('utf-8-sig')), as_attachment=True, download_name=f'{fn}.csv', mimetype='text/csv')

# Teacher: proctor report
@app.route('/teacher/quiz/<code>/proctor-report/<int:submission_id>')
@login_required
def teacher_proctor_report(code, submission_id):
    guard = require_teacher()
    if guard:
        return guard
    quiz = db.session.query(Quiz).filter_by(code=code.upper(), created_by=current_user.id).first()
    if not quiz:
        flash('Quiz not found', 'error')
        return redirect(url_for('dashboard'))
    submission = db.session.query(QuizSubmission).filter_by(id=submission_id, quiz_id=quiz.id).first()
    if not submission:
        flash('Submission not found', 'error')
        return redirect(url_for('teacher_quiz_results', code=code))
    student = db.session.query(User).filter_by(id=submission.student_id).first()
    snapshots = db.session.query(ProctoringSnapshot).filter_by(submission_id=submission.id).order_by(ProctoringSnapshot.captured_at).all()
    breaches = db.session.query(ProctoringBreach).filter_by(submission_id=submission.id).order_by(ProctoringBreach.occurred_at).all()
    breach_counts = {}
    for b in breaches:
        breach_counts[b.breach_type] = breach_counts.get(b.breach_type, 0) + 1
    return render_template('proctor_report.html', quiz=quiz, submission=submission, student=student, snapshots=snapshots, breaches=breaches, breach_counts=breach_counts, total_breaches=len(breaches), default_thresholds={'TOTAL_BREACHES': 30, 'TAB_SWITCH_COUNT': 20, 'FULLSCREEN_EXIT_COUNT': 20, 'CAMERA_OFF_DETECTED': 5})

# Teacher: proctor report actions
@app.route('/teacher/quiz/<code>/proctor-report/<int:submission_id>/action', methods=['POST'])
@login_required
def teacher_proctor_report_action(code, submission_id):
    guard = require_teacher()
    if guard:
        return guard
    quiz = db.session.query(Quiz).filter_by(code=code.upper(), created_by=current_user.id).first()
    if not quiz:
        return redirect(url_for('dashboard'))
    submission = db.session.query(QuizSubmission).filter_by(id=submission_id, quiz_id=quiz.id).first()
    if not submission:
        return redirect(url_for('teacher_quiz_results', code=code))
    action = request.form.get('action')
    if action == 'mark_cheating':
        submission.marked_as_cheating = True
    elif action == 'unmark_cheating':
        submission.marked_as_cheating = False
    elif action == 'add_note':
        note = request.form.get('proctor_note', '').strip()
        if note:
            submission.proctor_notes = (submission.proctor_notes or '') + '\n' + note
    db.session.commit()
    return redirect(url_for('teacher_proctor_report', code=code, submission_id=submission_id))

# Student: view quiz result (after 15 minutes)
@app.route('/quiz/result/<int:submission_id>')
@login_required
def view_quiz_result(submission_id):
    submission = db.session.query(QuizSubmission).filter_by(
        id=submission_id, 
        student_id=current_user.id
    ).first()
    if not submission:
        flash('Submission not found', 'error')
        return redirect(url_for('dashboard'))
    
    # Check if 15 minutes have passed
    current_time = datetime.utcnow()
    if submission.review_unlocked_at and submission.review_unlocked_at > current_time:
        flash('Results will be available 15 minutes after submission', 'info')
        return redirect(url_for('dashboard'))
    
    # Get quiz and questions
    quiz = db.session.query(Quiz).filter_by(id=submission.quiz_id).first()
    if not quiz:
        flash('Quiz not found', 'error')
        return redirect(url_for('dashboard'))
    
    questions = db.session.query(QuizQuestion).filter_by(quiz_id=quiz.id).order_by(QuizQuestion.id).all()
    answers = db.session.query(QuizAnswer).filter_by(submission_id=submission.id).all()
    
    answer_map = {ans.question_id: ans for ans in answers}
    
    # Format results for template
    results = []
    for q in questions:
        ans = answer_map.get(q.id)
        if q.qtype == 'mcq':
            options = json.loads(q.options_json) if q.options_json else []
            user_answer = ans.user_answer if ans else ''
            correct_option = next((opt for opt in options if opt.startswith(f"{q.answer}.")), '') if q.answer else ''
            
            results.append({
                'question': q.question,
                'user_answer': user_answer,
                'correct_answer': correct_option,
                'is_correct': ans.is_correct if ans else False,
                'type': 'mcq',
                'marks': q.marks
            })
        elif q.qtype == 'subjective':
            results.append({
                'question': q.question,
                'user_answer': ans.user_answer if ans else '',
                'sample_answer': q.answer or 'N/A',
                'ai_score': ans.ai_score if ans else 0.0,
                'scored_marks': ans.scored_marks if ans else 0.0,
                'marks': q.marks,
                'type': 'subjective'
            })
        elif q.qtype == 'coding':
            test_results = json.loads(ans.test_results_json) if ans and ans.test_results_json else []
            results.append({
                'question': q.question,
                'user_answer': ans.user_answer if ans else '',
                'code_language': ans.code_language if ans else '',
                'passed_test_cases': ans.passed_test_cases if ans else 0,
                'total_test_cases': ans.total_test_cases if ans else 0,
                'test_results': test_results,
                'scored_marks': ans.scored_marks if ans else 0.0,
                'marks': q.marks,
                'type': 'coding',
                'sample_input': q.sample_input,
                'sample_output': q.sample_output
            })
    
    final_score = f"{submission.score:.1f}/{submission.total:.1f}"
    percentage = submission.percentage
    passed = submission.passed
    
    return render_template('shared_quiz_results.html', 
                         quiz=quiz,
                         submission=submission,
                         results=results,
                         final_score=final_score,
                         percentage=percentage,
                         passed=passed)

# Teacher: allow student to retake quiz
@app.route('/teacher/quiz/<code>/allow-retake/<int:submission_id>', methods=['POST'])
@login_required
def allow_student_retake(code, submission_id):
    guard = require_teacher()
    if guard:
        return guard
    
    # Verify quiz ownership
    quiz = db.session.query(Quiz).filter_by(code=code.upper(), created_by=current_user.id).first()
    if not quiz:
        flash('Quiz not found or you do not have permission', 'error')
        return redirect(url_for('dashboard'))
    
    # Get the submission and verify it belongs to this quiz
    submission = db.session.query(QuizSubmission).filter_by(
        id=submission_id,
        quiz_id=quiz.id
    ).first()
    
    if not submission:
        flash('Submission not found', 'error')
        return redirect(url_for('teacher_quiz_results', code=code))
    
    # Get student name
    student = db.session.query(User).filter_by(id=submission.student_id).first()
    student_name = student.username if student else f"Student {submission.student_id}"
    
    try:
        # Delete dependent child rows first to satisfy foreign key constraints.
        db.session.query(QuizAnswer).filter_by(submission_id=submission.id).delete()
        db.session.query(ProctoringSnapshot).filter_by(submission_id=submission.id).delete()
        db.session.query(ProctoringBreach).filter_by(submission_id=submission.id).delete()

        # Delete submission
        db.session.delete(submission)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Error allowing retake for submission {submission.id}: {str(e)}")
        flash('Could not reset this submission. Please try again.', 'error')
        return redirect(url_for('teacher_quiz_results', code=code))

    flash(f'Successfully allowed {student_name} to retake the quiz. Their previous submission has been reset.', 'success')
    return redirect(url_for('teacher_quiz_results', code=code))

# API endpoints for code testing
@app.route('/api/test_code', methods=['POST'])
@login_required
def test_code():
    """Test code execution for coding questions"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
            
        code = data.get('code', '')
        language = data.get('language', 'python')
        test_input = data.get('test_input', '')
        time_limit = int(data.get('time_limit', 2))
        memory_limit = int(data.get('memory_limit', 256))
        
        if not code:
            return jsonify({'success': False, 'error': 'No code provided'}), 400
        
        result = execute_code(code, language, test_input, time_limit, memory_limit)
        return jsonify({'success': True, 'result': result})
        
    except json.JSONDecodeError as e:
        return jsonify({'success': False, 'error': f'Invalid JSON: {str(e)}'}), 400
    except Exception as e:
        print(f"Error in test_code API: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/run_test_cases', methods=['POST'])
@login_required
def api_run_test_cases():
    """Run test cases for coding questions"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
            
        code = data.get('code', '')
        language = data.get('language', 'python')
        test_cases = data.get('test_cases', [])
        time_limit = int(data.get('time_limit', 2))
        memory_limit = int(data.get('memory_limit', 256))
        
        if not code:
            return jsonify({'success': False, 'error': 'Code is required'}), 400
        if not test_cases or not isinstance(test_cases, list):
            return jsonify({'success': False, 'error': 'Test cases must be a non-empty list'}), 400
        
        result = run_test_cases(code, language, test_cases, time_limit, memory_limit)
        return jsonify({'success': True, 'result': result})
        
    except json.JSONDecodeError as e:
        return jsonify({'success': False, 'error': f'Invalid JSON: {str(e)}'}), 400
    except Exception as e:
        print(f"Error in run_test_cases API: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

# Migration route - run once to add new columns
@app.route('/run-migration')
@login_required
def run_migration_route():
    """Run database migration for new features"""
    try:
        from migrate_new_features import migrate_database
        migrate_database()
        return """
        <!DOCTYPE html>
        <html>
        <head>
            <title>Migration Complete</title>
            <style>
                body { font-family: Arial, sans-serif; max-width: 600px; margin: 50px auto; padding: 20px; }
                .success { background: #d4edda; color: #155724; padding: 15px; border-radius: 5px; margin: 20px 0; }
                .btn { display: inline-block; padding: 10px 20px; background: #007bff; color: white; text-decoration: none; border-radius: 5px; margin: 10px 5px; }
            </style>
        </head>
        <body>
            <h1>✅ Migration Complete!</h1>
            <div class="success">
                Database migration has been completed successfully. All new columns have been added.
            </div>
            <p>You can now:</p>
            <ul>
                <li>Create quizzes with coding questions</li>
                <li>Use the 15-minute review unlock feature</li>
                <li>Enable one attempt per student</li>
                <li>Allow students to retake quizzes (teacher feature)</li>
            </ul>
            <a href="/dashboard" class="btn">Go to Dashboard</a>
            <a href="/teacher/quiz/new_simple" class="btn">Create Quiz</a>
        </body>
        </html>
        """
    except Exception as e:
        import traceback
        error_msg = str(e)
        traceback.print_exc()
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Migration Error</title>
            <style>
                body {{ font-family: Arial, sans-serif; max-width: 600px; margin: 50px auto; padding: 20px; }}
                .error {{ background: #f8d7da; color: #721c24; padding: 15px; border-radius: 5px; margin: 20px 0; }}
                .btn {{ display: inline-block; padding: 10px 20px; background: #007bff; color: white; text-decoration: none; border-radius: 5px; margin: 10px 5px; }}
                pre {{ background: #f4f4f4; padding: 10px; border-radius: 5px; overflow-x: auto; }}
            </style>
        </head>
        <body>
            <h1>❌ Migration Error</h1>
            <div class="error">
                <strong>Error:</strong> {error_msg}
            </div>
            <p>If you see a "column already exists" error, that's okay - the migration may have partially completed.</p>
            <p>Try creating a quiz now. If it still fails, run the SQL commands directly in NeonDB (see instructions below).</p>
            <h3>Alternative: Run SQL in NeonDB</h3>
            <p>Go to your NeonDB dashboard → SQL Editor → Run these commands:</p>
            <pre>
ALTER TABLE quiz_question ADD COLUMN IF NOT EXISTS test_cases_json TEXT;
ALTER TABLE quiz_question ADD COLUMN IF NOT EXISTS language_constraints TEXT;
ALTER TABLE quiz_question ADD COLUMN IF NOT EXISTS time_limit_seconds INTEGER;
ALTER TABLE quiz_question ADD COLUMN IF NOT EXISTS memory_limit_mb INTEGER;
ALTER TABLE quiz_question ADD COLUMN IF NOT EXISTS sample_input TEXT;
ALTER TABLE quiz_question ADD COLUMN IF NOT EXISTS sample_output TEXT;
ALTER TABLE quiz_question ADD COLUMN IF NOT EXISTS starter_code TEXT;
ALTER TABLE quiz_answer ADD COLUMN IF NOT EXISTS code_language VARCHAR(20);
ALTER TABLE quiz_answer ADD COLUMN IF NOT EXISTS test_results_json TEXT;
ALTER TABLE quiz_answer ADD COLUMN IF NOT EXISTS passed_test_cases INTEGER DEFAULT 0;
ALTER TABLE quiz_answer ADD COLUMN IF NOT EXISTS total_test_cases INTEGER DEFAULT 0;
            </pre>
            <a href="/dashboard" class="btn">Go to Dashboard</a>
        </body>
        </html>
        """

# Temporary helper: run lightweight migration for SQLite (adds missing columns/tables)
@app.route('/dev/migrate')
def dev_migrate():
    try:
        # Only for SQLite local use
        from sqlalchemy import text
        with db.engine.begin() as conn:
            # Check if 'role' column exists on user
            has_role = False
            try:
                res = conn.execute(text("PRAGMA table_info(user);"))
                for row in res:
                    if str(row[1]) == 'role':
                        has_role = True
                        break
            except Exception:
                pass

            if not has_role:
                try:
                    conn.execute(text("ALTER TABLE user ADD COLUMN role VARCHAR(20) NOT NULL DEFAULT 'student';"))
                except Exception as e:
                    print(f"ALTER TABLE role add failed (may already exist): {e}")

            # Add difficulty and duration to quiz if missing
            try:
                res = conn.execute(text("PRAGMA table_info(quiz);"))
                cols = [str(r[1]) for r in res]
                if 'difficulty' not in cols:
                    conn.execute(text("ALTER TABLE quiz ADD COLUMN difficulty VARCHAR(20) DEFAULT 'beginner';"))
                if 'duration_minutes' not in cols:
                    conn.execute(text("ALTER TABLE quiz ADD COLUMN duration_minutes INTEGER;"))
            except Exception as e:
                print(f"ALTER TABLE quiz add columns failed (may exist): {e}")

            # Add review_unlocked_at and fullscreen_exit_flag to quiz_submission if missing
            try:
                res = conn.execute(text("PRAGMA table_info(quiz_submission);"))
                cols = [str(r[1]) for r in res]
                if 'review_unlocked_at' not in cols:
                    conn.execute(text("ALTER TABLE quiz_submission ADD COLUMN review_unlocked_at DATETIME;"))
                if 'fullscreen_exit_flag' not in cols:
                    conn.execute(text("ALTER TABLE quiz_submission ADD COLUMN fullscreen_exit_flag BOOLEAN DEFAULT 0;"))
                if 'alt_tab_flag' not in cols:
                    conn.execute(text("ALTER TABLE quiz_submission ADD COLUMN alt_tab_flag BOOLEAN DEFAULT 0;"))
                if 'win_shift_s_flag' not in cols:
                    conn.execute(text("ALTER TABLE quiz_submission ADD COLUMN win_shift_s_flag BOOLEAN DEFAULT 0;"))
                if 'win_prtscn_flag' not in cols:
                    conn.execute(text("ALTER TABLE quiz_submission ADD COLUMN win_prtscn_flag BOOLEAN DEFAULT 0;"))
                if 'prtscn_flag' not in cols:
                    conn.execute(text("ALTER TABLE quiz_submission ADD COLUMN prtscn_flag BOOLEAN DEFAULT 0;"))
                if 'answered_count' not in cols:
                    conn.execute(text("ALTER TABLE quiz_submission ADD COLUMN answered_count INTEGER DEFAULT 0;"))
                if 'question_count' not in cols:
                    conn.execute(text("ALTER TABLE quiz_submission ADD COLUMN question_count INTEGER DEFAULT 0;"))
                if 'is_full_completion' not in cols:
                    conn.execute(text("ALTER TABLE quiz_submission ADD COLUMN is_full_completion BOOLEAN DEFAULT 0;"))
                if 'started_at' not in cols:
                    conn.execute(text("ALTER TABLE quiz_submission ADD COLUMN started_at DATETIME;"))
                if 'completed' not in cols:
                    conn.execute(text("ALTER TABLE quiz_submission ADD COLUMN completed BOOLEAN DEFAULT 0;"))
            except Exception as e:
                print(f"ALTER TABLE quiz_submission add columns failed (may exist): {e}")

            try:
                mi_res = conn.execute(text("PRAGMA table_info(mock_interview_session);"))
                mi_cols = [str(r[1]) for r in mi_res]
                if mi_cols:
                    if 'is_preview' not in mi_cols:
                        conn.execute(
                            text("ALTER TABLE mock_interview_session ADD COLUMN is_preview BOOLEAN DEFAULT 0;")
                        )
                    if 'question_count' not in mi_cols:
                        conn.execute(
                            text(
                                "ALTER TABLE mock_interview_session ADD COLUMN question_count INTEGER DEFAULT 5;"
                            )
                        )
            except Exception as mi_e:
                print(f"dev/migrate mock_interview_session: {mi_e}")

        # Create any new tables
        db.create_all()
        flash('Migration completed. If you were logged in, reload the page. Next, visit /dev/promote_me.', 'success')
    except Exception as e:
        print(f"Migration error: {e}")
        flash('Migration failed. See server logs.', 'error')
    return redirect(url_for('dashboard'))

# Temporary helper: promote current user to teacher (local/dev use)
@app.route('/dev/promote_me')
@login_required
def dev_promote_me():
    try:
        current_user.role = 'teacher'
        db.session.commit()
        flash('Your account is now a Teacher. You can create shared quizzes.', 'success')
    except Exception as e:
        db.session.rollback()
        print(f"Promote error: {e}")
        flash('Failed to promote user.', 'error')
    return redirect(url_for('dashboard'))

# Database migration helper
@app.route('/dev/migrate_db')
def migrate_database():
    """Manual database migration to fix column lengths"""
    try:
        from sqlalchemy import text
        with db.engine.begin() as conn:
            # Add role column if missing
            try:
                result = conn.execute(text("SELECT column_name FROM information_schema.columns WHERE table_name='user' AND column_name='role'"))
                if not result.fetchone():
                    conn.execute(text("ALTER TABLE \"user\" ADD COLUMN role VARCHAR(20) DEFAULT 'student'"))
                    print("✅ Added role column")
            except Exception as e:
                print(f"Role column: {e}")
            
            # Update password_hash column length
            try:
                result = conn.execute(text("SELECT character_maximum_length FROM information_schema.columns WHERE table_name='user' AND column_name='password_hash'"))
                row = result.fetchone()
                if row and row[0] < 255:
                    conn.execute(text("ALTER TABLE \"user\" ALTER COLUMN password_hash TYPE VARCHAR(255)"))
                    print("✅ Updated password_hash column to VARCHAR(255)")
                else:
                    print("✅ Password_hash column already correct length")
            except Exception as e:
                print(f"Password hash column: {e}")
            
            # Add reset_token column if missing
            try:
                result = conn.execute(text("SELECT column_name FROM information_schema.columns WHERE table_name='user' AND column_name='reset_token'"))
                if not result.fetchone():
                    conn.execute(text("ALTER TABLE \"user\" ADD COLUMN reset_token VARCHAR(100) UNIQUE"))
                    print("✅ Added reset_token column")
                else:
                    print("✅ reset_token column already exists")
            except Exception as e:
                print(f"reset_token column: {e}")
            
            # Add reset_token_expiry column if missing
            try:
                result = conn.execute(text("SELECT column_name FROM information_schema.columns WHERE table_name='user' AND column_name='reset_token_expiry'"))
                if not result.fetchone():
                    conn.execute(text("ALTER TABLE \"user\" ADD COLUMN reset_token_expiry TIMESTAMP"))
                    print("✅ Added reset_token_expiry column")
                else:
                    print("✅ reset_token_expiry column already exists")
            except Exception as e:
                print(f"reset_token_expiry column: {e}")
            
            # Update qtype column length in quiz_question table
            try:
                result = conn.execute(text("SELECT character_maximum_length FROM information_schema.columns WHERE table_name='quiz_question' AND column_name='qtype'"))
                row = result.fetchone()
                if row and row[0] < 20:
                    conn.execute(text("ALTER TABLE quiz_question ALTER COLUMN qtype TYPE VARCHAR(20)"))
                    print("✅ Updated qtype column to VARCHAR(20)")
                else:
                    print("✅ Qtype column already correct length")
            except Exception as e:
                print(f"Qtype column: {e}")
            
            # Update answer column to TEXT for subjective questions (can store long answers)
            try:
                result = conn.execute(text("SELECT data_type FROM information_schema.columns WHERE table_name='quiz_question' AND column_name='answer'"))
                row = result.fetchone()
                if row and row[0] != 'text':
                    # Check if it's VARCHAR with small size
                    result2 = conn.execute(text("SELECT character_maximum_length FROM information_schema.columns WHERE table_name='quiz_question' AND column_name='answer'"))
                    row2 = result2.fetchone()
                    if row2 and (row2[0] is None or row2[0] < 1000):
                        conn.execute(text("ALTER TABLE quiz_question ALTER COLUMN answer TYPE TEXT"))
                        print("✅ Updated answer column to TEXT")
                    else:
                        print("✅ Answer column already TEXT or large enough")
                else:
                    print("✅ Answer column already TEXT")
            except Exception as e:
                print(f"Answer column: {e}")
        
        return """
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Database Migration - UniTest</title>
            <link rel="stylesheet" href="/static/style.css">
            <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
        </head>
        <body>
            <div class="auth-container">
                <div class="auth-card">
                    <div class="auth-header">
                        <h1 class="auth-title">✅ Database Migration Complete</h1>
                        <p class="auth-subtitle">Database schema has been updated successfully</p>
                    </div>
                    
                    <div class="alert alert-success">
                        <span class="alert-icon">✅</span>
                        Database migration completed! The qtype and answer columns have been updated. You can now create quizzes with subjective questions.
                    </div>
                    
                    <div class="auth-links">
                        <a href="/signup" class="btn" style="display: inline-block; text-align: center; margin-bottom: 20px;">Try Sign Up Now</a><br>
                        <a href="/login" class="auth-link">Sign In</a><br>
                        <a href="/" class="auth-link">← Back to Home</a>
                    </div>
                </div>
            </div>
        </body>
        </html>
        """
    except Exception as e:
        return f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Migration Error - UniTest</title>
            <link rel="stylesheet" href="/static/style.css">
            <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
        </head>
        <body>
            <div class="auth-container">
                <div class="auth-card">
                    <div class="auth-header">
                        <h1 class="auth-title">❌ Migration Error</h1>
                        <p class="auth-subtitle">Database migration failed</p>
                    </div>
                    
                    <div class="alert alert-error">
                        <span class="alert-icon">⚠️</span>
                        Error: {str(e)}
                    </div>
                    
                    <div class="auth-links">
                        <a href="/" class="auth-link">← Back to Home</a>
                    </div>
                </div>
            </div>
        </body>
        </html>
        """

@app.route('/quiz', methods=['GET', 'POST'])
@login_required
def quiz():
    if request.method == 'GET':
        # Handle topic parameter from AI learning modal or next/retry level
        topic = request.args.get('topic', '')
        difficulty = request.args.get('difficulty', '')
        action = request.args.get('action', '')
        
        if topic:
            return render_template('quiz.html', 
                                 prefill_topic=topic, 
                                 prefill_difficulty=difficulty,
                                 action=action)
    
    if request.method == 'POST':
        topic = request.form.get('topic', '').strip()
        question_type = request.form.get('question_type', 'mcq')
        mcq_count = int(request.form.get('mcq_count', 3))
        subj_count = int(request.form.get('subj_count', 2))
        difficulty_level = request.form.get('difficulty_level', 'beginner')
        
        # Handle PDF file uploads (single or multiple)
        pdf_content = None
        pdf_file_paths = []
        
        # Check for PDF file uploads (supports multiple files)
        if 'file_upload' in request.files:
            files = request.files.getlist('file_upload')
            for file in files:
                if file and file.filename and file.filename.lower().endswith('.pdf'):
                    try:
                        import tempfile
                        import os
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
                            file.save(tmp_file.name)
                            pdf_file_paths.append(tmp_file.name)
                            print(f'Saved PDF: {file.filename} to {tmp_file.name}')
                    except Exception as e:
                        print(f'Error saving PDF {file.filename}: {str(e)}')
                        continue
        
        # Extract content from PDFs
        if pdf_file_paths:
            try:
                print(f"📄 Processing {len(pdf_file_paths)} PDF file(s)...")
                pdf_content = extract_pdf_content(pdf_file_paths)
                if pdf_content:
                    print(f"✅ PDF content extracted: {len(pdf_content)} characters")
                    # Extract a topic from PDF for tracking purposes (optional)
                    if not topic:
                        # Try to extract topic from first PDF filename
                        import os
                        first_pdf_name = os.path.basename(pdf_file_paths[0])
                        topic = extract_topic_from_filename(pdf_file_paths[0]) or "PDF Content"
                        print(f"📝 Extracted topic from PDF: {topic}")
                    
                    pdf_info = f'Successfully processed {len(pdf_file_paths)} PDF file(s). Questions will be generated from PDF content.'
                    if OCR_AVAILABLE:
                        pdf_info += ' (Local OCR + Cloud OCR enabled for scanned PDFs)'
                    else:
                        pdf_info += ' (Cloud OCR enabled for scanned PDFs)'
                    flash(pdf_info, 'success')
                else:
                    # Provide more helpful error message
                    error_msg = 'Could not extract content from PDF(s). '
                    error_msg += 'This may happen if the PDF is scanned/image-based and OCR services are unavailable. '
                    error_msg += 'Please try again or enter a topic manually to generate questions.'
                    flash(error_msg, 'error')
                    # Clean up temp files
                    import os
                    for path in pdf_file_paths:
                        try:
                            os.unlink(path)
                        except:
                            pass
                    return redirect(url_for('quiz'))
            except Exception as e:
                print(f"❌ Error processing PDF(s): {str(e)}")
                import traceback
                traceback.print_exc()
                flash(f'Error processing PDF(s): {str(e)}', 'error')
                # Clean up temp files
                import os
                for path in pdf_file_paths:
                    try:
                        os.unlink(path)
                    except:
                        pass
                return redirect(url_for('quiz'))
        
        # If no PDFs and no topic, require topic
        if not pdf_content and not topic:
            flash('Please either enter a topic OR upload PDF file(s) to generate questions from.', 'error')
            return redirect(url_for('quiz'))

        # Get user's current bloom level for this topic (for progress tracking)
        progress = db.session.query(Progress).filter_by(user_id=current_user.id, topic=topic).first()
        bloom_level = progress.bloom_level if progress else 1

        # Generate questions using difficulty level (with PDF content if available)
        questions = []
        try:
            if question_type == "both":
                mcq_questions = generate_quiz(topic or "PDF Content", difficulty_level, "mcq", mcq_count, pdf_content)
                subj_questions = generate_quiz(topic or "PDF Content", difficulty_level, "subjective", subj_count, pdf_content)
                if mcq_questions and subj_questions:
                    questions = mcq_questions + subj_questions
            else:
                num_q = mcq_count if question_type == "mcq" else subj_count
                questions = generate_quiz(topic or "PDF Content", difficulty_level, question_type, num_q, pdf_content)
        except Exception as gen_error:
            error_msg = str(gen_error)
            print(f"Quiz generation error: {error_msg}")
            print(f"Error type: {type(gen_error).__name__}")
            import traceback
            traceback.print_exc()
            
            # Provide more helpful error message
            if "API" in error_msg or "key" in error_msg.lower() or "configuration" in error_msg.lower():
                user_error_msg = "Failed to generate quiz questions: API configuration error. Please check your API keys in Vercel environment variables (OPENROUTER_API_KEY and/or GOOGLE_AI_API_KEY)."
            elif "PDF" in error_msg or "content" in error_msg.lower():
                user_error_msg = f"Failed to generate quiz from PDF: {error_msg}. Please try with a different PDF or enter a topic manually."
            else:
                user_error_msg = f"Failed to generate quiz questions: {error_msg}"
            
            flash(user_error_msg, 'error')
            # Clean up temporary PDF files
            if pdf_file_paths:
                import os
                for path in pdf_file_paths:
                    try:
                        os.unlink(path)
                    except:
                        pass
            return redirect(url_for('quiz'))
        
        # Clean up temporary PDF files after question generation
        if pdf_file_paths:
            import os
            for path in pdf_file_paths:
                try:
                    os.unlink(path)
                except Exception as cleanup_error:
                    print(f"Warning: Could not delete temp file {path}: {cleanup_error}")

        if questions:
            session['current_quiz'] = {
                'questions': questions,
                'topic': topic,
                'bloom_level': bloom_level,
                'difficulty_level': difficulty_level
            }
            return redirect(url_for('take_quiz'))
        else:
            flash('Failed to generate quiz questions. Please check your API key configuration or try again later.', 'error')

    return render_template('quiz.html')

@app.route('/take_quiz')
@login_required
def take_quiz():
    quiz_data = session.get('current_quiz')
    if not quiz_data:
        flash('No quiz available', 'error')
        return redirect(url_for('quiz'))
    # Re-normalize coding questions to enforce stable testcase shape/count
    changed = False
    questions = quiz_data.get('questions', [])
    for idx, q in enumerate(questions):
        if isinstance(q, dict) and q.get('type') == 'coding':
            questions[idx] = _normalize_coding_question(q)
            changed = True
    if changed:
        quiz_data['questions'] = questions
        session['current_quiz'] = quiz_data
        session.modified = True

    resp = make_response(render_template('take_quiz.html', quiz_data=quiz_data))
    if quiz_data.get('placement_module'):
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        resp.headers['Pragma'] = 'no-cache'
        resp.headers['Expires'] = '0'
    return resp

@app.route('/submit_quiz', methods=['POST'])
@login_required
def submit_quiz():
    quiz_data = session.get('current_quiz')
    if not quiz_data:
        return jsonify({'error': 'No quiz available'})

    questions = quiz_data['questions']
    topic = quiz_data['topic']
    bloom_level = quiz_data['bloom_level']
    difficulty_level = quiz_data.get('difficulty_level', 'beginner')
    practice_recommendations = quiz_data.get('practice_recommendations', [])
    
    # Get answers for each question
    user_answers = []
    for i in range(len(questions)):
        question = questions[i]
        if question.get('type') == 'mcq':
            # For MCQ questions, get from question group
            answer = request.form.get(f'question_{i}')
        elif question.get('type') == 'coding':
            answer = request.form.get(f'coding_answers[{i}]')
        else:
            # For subjective questions, get from subjective_answers array
            answer = request.form.get(f'subjective_answers[{i}]')
        
        if not answer or not str(answer).strip():
            flash(f'Please answer question {i+1} before submitting.', 'error')
            return redirect(url_for('take_quiz'))
        user_answers.append(answer)

    # Calculate scores
    correct_answers = 0
    total_marks = 0
    scored_marks = 0
    results = []

    for i, (q, user_ans) in enumerate(zip(questions, user_answers)):
        if q.get('type') == 'mcq':
            user_choice = user_ans.split(". ")[0] if user_ans else ""
            is_correct = user_choice == q["answer"]
            if is_correct:
                correct_answers += 1
            results.append({
                'question': q['question'],
                'user_answer': user_ans,
                'correct_answer': next((opt for opt in q["options"] if opt.startswith(f"{q['answer']}.")), ""),
                'is_correct': is_correct,
                'type': 'mcq',
                'solution': q.get('solution', '')
            })
        elif q.get('type') == 'coding':
            marks = q.get('marks', 10)
            total_marks += marks
            language = request.form.get(f'language_{i}', 'python')
            test_cases = q.get('test_cases') or []
            time_limit = int(q.get('time_limit_seconds', 2) or 2)
            memory_limit = int(q.get('memory_limit_mb', 256) or 256)
            ai_score = 0.0
            passed_test_cases = 0
            total_test_cases = len(test_cases)
            run_error = None
            if user_ans.strip():
                try:
                    test_run = run_test_cases(user_ans, language, test_cases, time_limit, memory_limit)
                    if test_run.get('executor_unavailable'):
                        # Avoid penalizing users when external compiler service is down.
                        ai_score, ai_feedback = evaluate_coding_answer_ai(
                            q.get('question', ''),
                            user_ans,
                            language,
                            q.get('sample_input', ''),
                            q.get('sample_output', ''),
                            q.get('solution', '')
                        )
                        scored_marks += ai_score * marks
                        if ai_score >= 0.6:
                            correct_answers += 1
                        run_error = (
                            (test_run.get('executor_message') or 'Code runner is temporarily unavailable.')
                            + f" AI evaluation used: {ai_feedback}"
                        )
                    else:
                        passed_test_cases = int(test_run.get('passed', 0))
                        total_test_cases = int(test_run.get('total', total_test_cases))
                        ai_score = (passed_test_cases / total_test_cases) if total_test_cases > 0 else 0.0
                        scored_marks += ai_score * marks
                        if ai_score >= 0.6:
                            correct_answers += 1
                except Exception as eval_err:
                    run_error = str(eval_err)

            results.append({
                'question': q['question'],
                'user_answer': user_ans,
                'marks': marks,
                'ai_score': ai_score,
                'scored_marks': ai_score * marks,
                'type': 'coding',
                'solution': q.get('solution', ''),
                'sample_input': q.get('sample_input', ''),
                'sample_output': q.get('sample_output', ''),
                'leetcode_id': q.get('leetcode_id'),
                'leetcode_title': q.get('leetcode_title'),
                'leetcode_url': q.get('leetcode_url'),
                'difficulty': q.get('difficulty', 'medium'),
                'language': language,
                'passed_test_cases': passed_test_cases,
                'total_test_cases': total_test_cases,
                'run_error': run_error
            })
        else:  # subjective
            marks = q.get('marks', 10)
            total_marks += marks
            
            if user_ans.strip():
                ai_score = evaluate_subjective_answer(q['question'], user_ans, q.get('answer', ''))
                scored_marks += ai_score * marks
                if ai_score >= 0.6:
                    correct_answers += 1
            else:
                ai_score = 0.0

            results.append({
                'question': q['question'],
                'user_answer': user_ans,
                'sample_answer': q.get('answer', 'N/A'),
                'marks': marks,
                'ai_score': ai_score,
                'scored_marks': ai_score * marks,
                'type': 'subjective',
                'solution': q.get('solution', '')
            })

    # Calculate final score
    has_graded_questions = any(q.get('type') in ('subjective', 'coding') for q in questions)
    
    if has_graded_questions:
        percentage = (scored_marks / total_marks) * 100 if total_marks > 0 else 0
        passed = percentage >= 60
        final_score = f"{scored_marks:.1f}/{total_marks} marks"
    else:
        percentage = (correct_answers / len(questions)) * 100 if questions else 0
        passed = percentage >= 60
        final_score = f"{correct_answers}/{len(questions)}"

    placement_module = quiz_data.get('placement_module')
    placement_level = (quiz_data.get('placement_level') or 'l1').strip().lower()
    if placement_level not in ('l1', 'l2'):
        placement_level = 'l1'
    if placement_module in PLACEMENT_SEQUENCE:
        state = _get_placement_state()
        lv = state['levels'].setdefault(placement_module, _default_placement_levels()[placement_module])
        if placement_level == 'l1':
            lv['l1_score'] = round(float(percentage), 1)
            lv['l1_passed'] = percentage >= _placement_pass_threshold(placement_module, 'l1')
            if lv['l1_passed']:
                lv['active_level'] = 'l2'
            passed = percentage >= _placement_pass_threshold(placement_module, 'l1')
        else:
            if lv.get('l1_passed', False):
                lv['l2_score'] = round(float(percentage), 1)
                lv['l2_passed'] = percentage >= _placement_pass_threshold(placement_module, 'l2')
                lv['active_level'] = 'l2'
                passed = percentage >= _placement_pass_threshold(placement_module, 'l2')
            else:
                passed = False
        state['scores'][placement_module] = round((float(lv.get('l1_score', 0.0)) + float(lv.get('l2_score', 0.0))) / 2.0, 1)
        state['completed'][placement_module] = bool(lv.get('l1_passed') and lv.get('l2_passed'))
        if state['completed'][placement_module]:
            if placement_module == 'aptitude':
                state['current_stage'] = 'group_discussion'
            elif placement_module == 'fundamentals':
                state['current_stage'] = 'basic_coding'
            elif placement_module == 'basic_coding':
                state['current_stage'] = 'coding'
        _save_placement_state(state)

    # Update progress
    progress = db.session.query(Progress).filter_by(user_id=current_user.id, topic=topic).first()
    if progress:
        if passed and bloom_level + 1 > progress.bloom_level:
            progress.bloom_level = bloom_level + 1
    else:
        new_progress = Progress(
            user_id=current_user.id,
            topic=topic,
            bloom_level=bloom_level + 1 if passed else bloom_level
        )
        db.session.add(new_progress)
    
    db.session.commit()

    # Clear quiz session
    session.pop('current_quiz', None)

    return render_template('quiz_results.html', 
                         results=results, 
                         final_score=final_score, 
                         percentage=percentage, 
                         passed=passed,
                         topic=topic,
                         bloom_level=bloom_level,
                         difficulty_level=difficulty_level,
                         placement_module=placement_module,
                         practice_recommendations=practice_recommendations)

@app.route('/next_level', methods=['POST'])
@login_required
def next_level():
    """Automatically generate next level quiz and redirect to take_quiz"""
    try:
        topic = request.form.get('topic', '').strip()
        difficulty_level = request.form.get('difficulty_level', 'beginner').strip()
        
        if not topic:
            flash('Topic is required', 'error')
            return redirect(url_for('quiz'))
        
        # Determine next difficulty level
        difficulty_mapping = {
            "beginner": "intermediate",
            "intermediate": "difficult", 
            "difficult": "difficult"  # Stay at difficult if already at highest level
        }
        next_difficulty = difficulty_mapping.get(difficulty_level, "intermediate")
        
        # Generate questions for next level
        questions = generate_quiz(topic, next_difficulty, "mcq", 5)
        
        if questions:
            session['current_quiz'] = {
                'questions': questions,
                'topic': topic,
                'bloom_level': 1,  # This will be updated based on difficulty
                'difficulty_level': next_difficulty
            }
            flash(f'Generated {next_difficulty.title()} level quiz for {topic}!', 'success')
            return redirect(url_for('take_quiz'))
        else:
            flash('Failed to generate next level quiz', 'error')
            return redirect(url_for('quiz'))
    except Exception as e:
        print(f"Error in next_level: {str(e)}")
        flash('An error occurred while generating the next level quiz', 'error')
        return redirect(url_for('quiz'))

@app.route('/retry_level', methods=['POST'])
@login_required
def retry_level():
    """Automatically generate retry quiz and redirect to take_quiz"""
    try:
        topic = request.form.get('topic', '').strip()
        difficulty_level = request.form.get('difficulty_level', 'beginner').strip()
        
        if not topic:
            flash('Topic is required', 'error')
            return redirect(url_for('quiz'))
        
        # Generate questions for the same level
        questions = generate_quiz(topic, difficulty_level, "mcq", 5)
        
        if questions:
            session['current_quiz'] = {
                'questions': questions,
                'topic': topic,
                'bloom_level': 1,  # This will be updated based on difficulty
                'difficulty_level': difficulty_level
            }
            flash(f'Generated new {difficulty_level.title()} level quiz for {topic}!', 'success')
            return redirect(url_for('take_quiz'))
        else:
            flash('Failed to generate retry quiz', 'error')
            return redirect(url_for('quiz'))
    except Exception as e:
        print(f"Error in retry_level: {str(e)}")
        flash('An error occurred while generating the retry quiz', 'error')
        return redirect(url_for('quiz'))

@app.route('/continue_learning', methods=['POST'])
@login_required
def continue_learning():
    """Continue learning from where user left off based on their progress"""
    try:
        topic = request.form.get('topic', '').strip()
        
        if not topic:
            flash('Topic is required', 'error')
            return redirect(url_for('dashboard'))
        
        # Get user's current progress for this topic
        progress = db.session.query(Progress).filter_by(user_id=current_user.id, topic=topic).first()
        
        if not progress:
            flash('No progress found for this topic', 'error')
            return redirect(url_for('dashboard'))
        
        # Map bloom level to difficulty level
        difficulty_level = get_difficulty_from_bloom_level(progress.bloom_level)
        
        # Generate questions for the current level
        questions = generate_quiz(topic, difficulty_level, "mcq", 5)
        
        if questions:
            session['current_quiz'] = {
                'questions': questions,
                'topic': topic,
                'bloom_level': progress.bloom_level,
                'difficulty_level': difficulty_level
            }
            flash(f'Continuing {topic} at {difficulty_level.title()} level (Bloom Level {progress.bloom_level})!', 'success')
            return redirect(url_for('take_quiz'))
        else:
            flash('Failed to generate quiz for continuing learning', 'error')
            return redirect(url_for('dashboard'))
    except Exception as e:
        print(f"Error in continue_learning: {str(e)}")
        flash('An error occurred while continuing your learning', 'error')
        return redirect(url_for('dashboard'))

@app.route('/upload_pdf', methods=['POST'])
@login_required
def upload_pdf():
    """Handle PDF upload and extract topic"""
    if 'file_upload' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'})
    
    file = request.files['file_upload']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'})
    
    if file and file.filename.lower().endswith('.pdf'):
        try:
            # Save file temporarily
            import tempfile
            import os
            
            # Ensure temp directory exists
            temp_dir = tempfile.gettempdir()
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf', dir=temp_dir) as tmp_file:
                file.save(tmp_file.name)
                tmp_path = tmp_file.name
            
            original_filename = file.filename
            print(f"Processing PDF: {original_filename} (saved to {tmp_path})")
            
            # Process the PDF to extract topic
            topic = process_document(tmp_path)
            
            # If no topic extracted, try filename extraction using original filename
            if not topic:
                print("No topic extracted from PDF content, trying filename extraction...")
                # Create a mock path with original filename for extraction
                import os
                temp_dir = os.path.dirname(tmp_path)
                mock_path = os.path.join(temp_dir, original_filename)
                topic = extract_topic_from_filename(mock_path)
            
            # Clean up temporary file
            try:
                os.unlink(tmp_path)
            except Exception as cleanup_error:
                print(f"Warning: Could not delete temp file: {cleanup_error}")
            
            if topic:
                print(f"Successfully extracted topic: {topic}")
                return jsonify({'success': True, 'topic': topic, 'message': f'Topic extracted: {topic}'})
            else:
                # Provide helpful error message
                error_msg = (
                    'Could not extract topic from PDF. This might happen if:\n'
                    '- The PDF is scanned (image-based)\n'
                    '- The PDF is encrypted or protected\n'
                    '- The PDF contains no text\n'
                    'Please enter the topic manually.'
                )
                return jsonify({'success': False, 'error': error_msg})
                
        except Exception as e:
            print(f"Error processing PDF: {str(e)}")
            import traceback
            traceback.print_exc()
            error_msg = f'Error processing PDF: {str(e)}. Please try again or enter topic manually.'
            return jsonify({'success': False, 'error': error_msg})
    
    return jsonify({'success': False, 'error': 'Invalid file format. Please upload a PDF file.'})

@app.route('/ai_learn', methods=['POST'])
@login_required
def ai_learn():
    """AI-powered learning content generation using Groq API"""
    try:
        data = request.get_json() or {}
        topic = (data.get('topic') or '').strip()
        level = (data.get('level') or 'intermediate').strip()
        style = (data.get('style') or 'theoretical').strip()
        image_data_url = (data.get('image_data_url') or '').strip()
        follow_up_question = (data.get('follow_up_question') or '').strip()
        prior_content = (data.get('prior_content') or '').strip()

        if not topic:
            return jsonify({'success': False, 'error': 'Topic is required'})

        groq_key = os.environ.get('GROQ_API_KEY')
        if not groq_key:
            return jsonify({
                'success': False,
                'error': 'GROQ_API_KEY is not set. Add it in your environment variables.'
            }), 500

        text_model = os.environ.get('GROQ_MODEL', 'llama-3.3-70b-versatile')
        vision_model = os.environ.get('GROQ_VISION_MODEL', text_model)
        use_vision = bool(image_data_url)
        selected_model = vision_model if use_vision else text_model

        if image_data_url and len(image_data_url) > 3_000_000:
            return jsonify({
                'success': False,
                'error': 'Image too large. Please upload a smaller image (preferably under 2MB).'
            }), 400

        base_prompt = f"""
Create a personalized learning path for topic: {topic}
Level: {level}
Learning style: {style}
"""
        if follow_up_question:
            base_prompt += f"""

The user has already seen this prior learning content:
{prior_content[:2500]}

Now answer this follow-up question interactively:
{follow_up_question}

Return concise, practical clarification while still using sectioned markdown.
"""
        else:
            base_prompt += f"""
IMPORTANT: Use this EXACT section format:

## OVERVIEW
[2-3 sentence overview]

## KEY CONCEPTS
• [concept 1]
• [concept 2]
• [concept 3]

## LEARNING OBJECTIVES
✓ [objective 1]
✓ [objective 2]
✓ [objective 3]

## STUDY APPROACH
[Recommendations based on {style}]

## COMMON MISCONCEPTIONS
⚠️ [misconception 1]
⚠️ [misconception 2]

## NEXT STEPS
[Immediate action plan]
"""

        user_content = [{"type": "text", "text": base_prompt.strip()}]
        if use_vision:
            user_content.append({
                "type": "image_url",
                "image_url": {"url": image_data_url}
            })

        payload = {
            "model": selected_model,
            "temperature": 0.3,
            "max_tokens": 1400,
            "messages": [
                {"role": "system", "content": "You are an expert tutor. Be structured, accurate, and practical."},
                {"role": "user", "content": user_content if use_vision else base_prompt.strip()}
            ]
        }

        response = requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {groq_key}",
                "Content-Type": "application/json"
            },
            json=payload,
            timeout=30
        )
        if response.status_code == 429:
            return jsonify({
                'success': False,
                'error': 'Groq rate limit reached. Please wait 10-15 seconds and try again.'
            }), 429
        if response.status_code >= 400:
            return jsonify({
                'success': False,
                'error': f"Groq API error ({response.status_code}): {response.text[:300]}"
            }), 500

        result = response.json()
        content = (
            result.get('choices', [{}])[0]
            .get('message', {})
            .get('content', '')
            .strip()
        )
        if not content:
            return jsonify({'success': False, 'error': 'Empty response from Groq API.'}), 500

        return jsonify({'success': True, 'content': content})
    except Exception as e:
        return jsonify({'success': False, 'error': f'Error generating learning content: {str(e)}'})

@app.route('/ai_doubt_resolver', methods=['POST'])
@login_required
def ai_doubt_resolver():
    try:
        data = request.get_json(silent=True) or {}
        question = (data.get('question') or '').strip()
        user_answer = (data.get('user_answer') or '').strip()
        correct_answer = (data.get('correct_answer') or '').strip()
        question_type = (data.get('question_type') or 'general').strip()

        if not question:
            return jsonify({'success': False, 'error': 'Question is required.'}), 400
        groq_key = os.environ.get('GROQ_API_KEY')
        if not groq_key:
            return jsonify({
                'success': False,
                'error': 'GROQ_API_KEY is not set. Add it in environment variables.'
            }), 500
        model = os.environ.get('GROQ_MODEL', 'llama-3.3-70b-versatile')

        prompt = f"""
You are an expert tutor. Resolve the student's doubt based on a wrong/weak answer.

Question type: {question_type}
Question: {question}
Student answer: {user_answer}
Expected/correct answer: {correct_answer}

Return STRICT JSON only:
{{
  "concept_explained": "3-6 sentence clear explanation",
  "why_wrong": "2-4 sentence reason why student's answer was weak",
  "improvement_steps": ["step1", "step2", "step3"],
  "practice_question": "one practice question based on same concept"
}}
"""
        response = requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {groq_key}",
                "Content-Type": "application/json"
            },
            json={
                "model": model,
                "temperature": 0.2,
                "max_tokens": 900,
                "messages": [
                    {"role": "system", "content": "You are a precise tutor who always returns valid JSON."},
                    {"role": "user", "content": prompt}
                ]
            },
            timeout=25
        )
        if response.status_code == 429:
            return jsonify({
                'success': False,
                'error': 'Groq rate limit reached. Please retry after a few seconds.'
            }), 429
        if response.status_code >= 400:
            return jsonify({
                'success': False,
                'error': f"Groq API error ({response.status_code}): {response.text[:250]}"
            }), 500
        raw = (
            response.json()
            .get('choices', [{}])[0]
            .get('message', {})
            .get('content', '')
            .strip()
        )
        cleaned = raw.replace("```json", "").replace("```", "").strip()
        parsed = json.loads(cleaned)

        return jsonify({
            'success': True,
            'concept_explained': parsed.get('concept_explained', ''),
            'why_wrong': parsed.get('why_wrong', ''),
            'improvement_steps': parsed.get('improvement_steps', []),
            'practice_question': parsed.get('practice_question', '')
        })
    except Exception as e:
        return jsonify({'success': False, 'error': f'Could not resolve doubt: {str(e)}'}), 500

@app.route('/download_pdf')
@login_required
def download_pdf():
    quiz_data = session.get('current_quiz')
    if not quiz_data:
        return jsonify({'error': 'No quiz available'})

    questions = quiz_data['questions']
    topic = quiz_data['topic']
    bloom_level = quiz_data['bloom_level']

    # Create PDF in memory
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30
    )
    question_style = ParagraphStyle(
        'Question',
        parent=styles['Normal'],
        fontSize=12,
        spaceAfter=12,
        textColor=colors.black
    )
    option_style = ParagraphStyle(
        'Option',
        parent=styles['Normal'],
        fontSize=11,
        leftIndent=20,
        spaceAfter=6,
        textColor=colors.black
    )
    
    content = []
    content.append(Paragraph(f"Quiz: {topic}", title_style))
    content.append(Paragraph(f"Bloom's Taxonomy Level: {bloom_level}", styles['Heading2']))
    content.append(Spacer(1, 20))
    
    for i, q in enumerate(questions, 1):
        question_text = f"Q{i}. {q['question']}"
        content.append(Paragraph(question_text, question_style))
        
        if q.get('type') == 'mcq':
            for opt in q['options']:
                option_text = f"□ {opt}"
                content.append(Paragraph(option_text, option_style))
        else:
            content.append(Paragraph(f"Marks: {q.get('marks', 10)}", option_style))
            content.append(Paragraph("Answer: ________________________", option_style))
        
        content.append(Spacer(1, 20))
    
    doc.build(content)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"Quiz_{topic}_Level{bloom_level}.pdf",
        mimetype='application/pdf'
    )

@app.route('/admin/users')
@login_required
def admin_users():
    """Admin route to view user statistics - ADMIN ONLY"""
    # Check if user is admin
    if not current_user.is_admin:
        flash('Access denied: Administrator privileges required.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        from sqlalchemy import text, func
        from datetime import datetime, timedelta
        
        # Get total users
        total_users = db.session.query(func.count(User.id)).scalar()
        
        # Get users by role
        users_by_role = db.session.query(
            User.role,
            func.count(User.id).label('count')
        ).group_by(User.role).all()
        
        # Get users this month
        this_month_start = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        users_this_month = db.session.query(func.count(User.id)).filter(
            User.created_at >= this_month_start
        ).scalar()
        
        # Get users this week
        this_week_start = datetime.utcnow() - timedelta(days=datetime.utcnow().weekday())
        this_week_start = this_week_start.replace(hour=0, minute=0, second=0, microsecond=0)
        users_this_week = db.session.query(func.count(User.id)).filter(
            User.created_at >= this_week_start
        ).scalar()
        
        # Get users today
        today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        users_today = db.session.query(func.count(User.id)).filter(
            User.created_at >= today_start
        ).scalar()
        
        # Get recent signups (last 10)
        recent_users = db.session.query(User).order_by(
            User.created_at.desc()
        ).limit(10).all()
        
        # Get signups by date (last 30 days)
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)
        signups_by_date = db.session.query(
            func.date(User.created_at).label('date'),
            func.count(User.id).label('count')
        ).filter(
            User.created_at >= thirty_days_ago
        ).group_by(
            func.date(User.created_at)
        ).order_by(
            func.date(User.created_at).desc()
        ).all()
        
        # === LOGIN STATISTICS ===
        # Total logins (all time)
        total_logins = db.session.query(func.count(LoginHistory.id)).scalar() or 0
        
        # Unique users who have logged in
        unique_logged_in_users = db.session.query(func.count(func.distinct(LoginHistory.user_id))).scalar() or 0
        
        # Logins today
        logins_today = db.session.query(func.count(LoginHistory.id)).filter(
            LoginHistory.login_time >= today_start
        ).scalar() or 0
        
        # Logins this week
        logins_this_week = db.session.query(func.count(LoginHistory.id)).filter(
            LoginHistory.login_time >= this_week_start
        ).scalar() or 0
        
        # Logins this month
        logins_this_month = db.session.query(func.count(LoginHistory.id)).filter(
            LoginHistory.login_time >= this_month_start
        ).scalar() or 0
        
        # Logins by date (last 30 days)
        logins_by_date = db.session.query(
            func.date(LoginHistory.login_time).label('date'),
            func.count(LoginHistory.id).label('count')
        ).filter(
            LoginHistory.login_time >= thirty_days_ago
        ).group_by(
            func.date(LoginHistory.login_time)
        ).order_by(
            func.date(LoginHistory.login_time).desc()
        ).all()
        
        # Recent logins (last 20)
        recent_logins = db.session.query(LoginHistory).join(User).order_by(
            LoginHistory.login_time.desc()
        ).limit(20).all()
        
        # Most active users (by login count)
        most_active_users = db.session.query(
            User.username,
            User.email,
            func.count(LoginHistory.id).label('login_count'),
            func.max(LoginHistory.login_time).label('last_login')
        ).outerjoin(
            LoginHistory, User.id == LoginHistory.user_id
        ).group_by(
            User.id, User.username, User.email
        ).having(
            func.count(LoginHistory.id) > 0
        ).order_by(
            func.count(LoginHistory.id).desc()
        ).limit(10).all()

        # === WEBSITE VISITS & USER ACTIVITIES ===
        total_site_visits = db.session.query(func.count(SiteActivity.id)).scalar() or 0
        site_visits_today = db.session.query(func.count(SiteActivity.id)).filter(
            SiteActivity.created_at >= today_start
        ).scalar() or 0
        unique_site_visitors = db.session.query(
            func.count(func.distinct(SiteActivity.user_id))
        ).filter(SiteActivity.user_id.isnot(None)).scalar() or 0

        recent_activities = db.session.query(SiteActivity).outerjoin(User).order_by(
            SiteActivity.created_at.desc()
        ).limit(25).all()
        
        # Convert users_by_role to dictionary
        users_by_role_dict = {role: count for role, count in users_by_role} if users_by_role else {}
        
        return render_template('admin_users.html',
            total_users=total_users or 0,
            users_by_role=users_by_role_dict,
            users_this_month=users_this_month or 0,
            users_this_week=users_this_week or 0,
            users_today=users_today or 0,
            recent_users=recent_users,
            signups_by_date=signups_by_date,
            # Login statistics
            total_logins=total_logins,
            unique_logged_in_users=unique_logged_in_users,
            logins_today=logins_today,
            logins_this_week=logins_this_week,
            logins_this_month=logins_this_month,
            logins_by_date=logins_by_date,
            recent_logins=recent_logins,
            most_active_users=most_active_users,
            total_site_visits=total_site_visits,
            site_visits_today=site_visits_today,
            unique_site_visitors=unique_site_visitors,
            recent_activities=recent_activities
        )
    except Exception as e:
        print(f"Error in admin_users: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'Error loading user statistics: {str(e)}', 'error')
        return redirect(url_for('dashboard'))

@app.route('/admin/metrics')
@login_required
def admin_metrics():
    """Admin route to view system capacity metrics - ADMIN ONLY"""
    # Check if user is admin
    if not current_user.is_admin:
        flash('Access denied: Administrator privileges required.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        from sqlalchemy import text, func
        import os
        
        metrics = {
            'database': {},
            'capacity': {},
            'current_usage': {}
        }
        
        # Check database type
        database_url = app.config.get('SQLALCHEMY_DATABASE_URI', '')
        is_postgres = 'postgresql' in database_url.lower() or 'postgres' in database_url.lower()
        
        if is_postgres:
            # Get active database connections
            try:
                result = db.session.execute(text("""
                    SELECT 
                        count(*) as active_connections,
                        max_conn as max_connections
                    FROM pg_stat_activity, 
                    (SELECT setting::int as max_conn FROM pg_settings WHERE name = 'max_connections') as max_conn_setting
                    WHERE datname = current_database()
                    GROUP BY max_conn
                """))
                row = result.fetchone()
                if row:
                    metrics['database']['active_connections'] = row[0]
                    metrics['database']['max_connections'] = row[1]
                    metrics['database']['usage_percent'] = round((row[0] / row[1]) * 100, 2) if row[1] > 0 else 0
                else:
                    # Fallback query
                    result = db.session.execute(text("""
                        SELECT count(*) FROM pg_stat_activity 
                        WHERE datname = current_database()
                    """))
                    active = result.scalar() or 0
                    metrics['database']['active_connections'] = active
                    metrics['database']['max_connections'] = 100  # Default for free tier
                    metrics['database']['usage_percent'] = round((active / 100) * 100, 2)
            except Exception as e:
                print(f"Error getting connection stats: {e}")
                metrics['database']['active_connections'] = 'N/A'
                metrics['database']['max_connections'] = 'N/A'
                metrics['database']['usage_percent'] = 'N/A'
            
            # Get database size
            try:
                result = db.session.execute(text("""
                    SELECT pg_size_pretty(pg_database_size(current_database())) as db_size
                """))
                row = result.fetchone()
                metrics['database']['size'] = row[0] if row else 'N/A'
            except Exception as e:
                metrics['database']['size'] = 'N/A'
        else:
            metrics['database']['type'] = 'SQLite (Local Development)'
            metrics['database']['active_connections'] = 'N/A'
            metrics['database']['max_connections'] = 'N/A'
            metrics['database']['usage_percent'] = 'N/A'
        
        # Get current logged-in users (active sessions in last 5 minutes)
        try:
            from datetime import datetime, timedelta
            five_minutes_ago = datetime.utcnow() - timedelta(minutes=5)
            
            # Count users with recent activity (approximate)
            active_users = db.session.query(func.count(func.distinct(QuizSubmission.user_id))).filter(
                QuizSubmission.submitted_at >= five_minutes_ago
            ).scalar() or 0
            
            # Also check recent logins
            recent_logins = db.session.query(func.count(func.distinct(LoginHistory.user_id))).filter(
                LoginHistory.login_time >= five_minutes_ago
            ).scalar() or 0
            
            metrics['current_usage']['recently_active_users'] = max(active_users, recent_logins)
        except Exception as e:
            metrics['current_usage']['recently_active_users'] = 'N/A'
        
        # Estimate capacity based on database connections
        if isinstance(metrics['database'].get('max_connections'), int):
            max_conn = metrics['database']['max_connections']
            # Estimate: 1.2 connections per user on average
            estimated_capacity = int((max_conn - 10) / 1.2)  # Reserve 10 connections
            metrics['capacity']['estimated_concurrent_users'] = estimated_capacity
            metrics['capacity']['safe_concurrent_users'] = int(estimated_capacity * 0.8)  # 80% of max for safety
        else:
            metrics['capacity']['estimated_concurrent_users'] = 'N/A'
            metrics['capacity']['safe_concurrent_users'] = 'N/A'
        
        # Get deployment info
        metrics['deployment'] = {
            'platform': 'Vercel' if os.environ.get('VERCEL') else 'Unknown',
            'database_provider': 'NeonDB' if 'neon.tech' in database_url else 'Other'
        }
        
        return jsonify(metrics)
        
    except Exception as e:
        print(f"Error in admin_metrics: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# Error handlers
@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return render_template('error.html', error="Internal Server Error"), 500

@app.errorhandler(404)
def not_found_error(error):
    return render_template('error.html', error="Page Not Found"), 404

def _ensure_mock_interview_session_columns(is_sqlite):
    """Add is_preview / question_count to mock_interview_session on existing deployments."""
    from sqlalchemy import text

    try:
        if is_sqlite:
            with db.engine.begin() as conn:
                mi_res = conn.execute(text("PRAGMA table_info(mock_interview_session);"))
                mi_cols = [str(r[1]) for r in mi_res]
                if not mi_cols:
                    return
                if 'is_preview' not in mi_cols:
                    conn.execute(
                        text("ALTER TABLE mock_interview_session ADD COLUMN is_preview BOOLEAN DEFAULT 0;")
                    )
                    print("✅ Added is_preview to mock_interview_session")
                if 'question_count' not in mi_cols:
                    conn.execute(
                        text(
                            "ALTER TABLE mock_interview_session ADD COLUMN question_count INTEGER DEFAULT 5;"
                        )
                    )
                    print("✅ Added question_count to mock_interview_session")
        else:
            tbl = db.session.execute(
                text(
                    "SELECT 1 FROM information_schema.tables WHERE table_schema='public' "
                    "AND table_name='mock_interview_session'"
                )
            )
            if not tbl.fetchone():
                return
            rcols = db.session.execute(
                text(
                    "SELECT column_name FROM information_schema.columns "
                    "WHERE table_schema='public' AND table_name='mock_interview_session'"
                )
            )
            existing_mi = {row[0] for row in rcols}
            if 'is_preview' not in existing_mi:
                db.session.execute(
                    text('ALTER TABLE mock_interview_session ADD COLUMN is_preview BOOLEAN DEFAULT FALSE')
                )
                db.session.commit()
                print("Added is_preview to mock_interview_session")
            if 'question_count' not in existing_mi:
                db.session.execute(
                    text('ALTER TABLE mock_interview_session ADD COLUMN question_count INTEGER DEFAULT 5')
                )
                db.session.commit()
                print("Added question_count to mock_interview_session")
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        print(f"ensure_mock_interview_session_columns: {e}")


# Initialize database tables
def init_db():
    try:
        with app.app_context():
            # Create all tables first
            db.create_all()
            
            # Check database type - PostgreSQL has information_schema, SQLite doesn't
            is_sqlite = 'sqlite' in app.config.get('SQLALCHEMY_DATABASE_URI', '').lower()
            
            # For PostgreSQL, check and add columns if needed
            if not is_sqlite:
                # Check if we need to add the role column to existing user table
                try:
                    from sqlalchemy import text
                    # Check if role column exists
                    result = db.session.execute(text("SELECT column_name FROM information_schema.columns WHERE table_name='user' AND column_name='role'"))
                    if not result.fetchone():
                        # Add role column if it doesn't exist
                        db.session.execute(text("ALTER TABLE \"user\" ADD COLUMN role VARCHAR(20) DEFAULT 'student'"))
                        db.session.commit()
                        print("Added role column to user table")
                except Exception as e:
                    print(f"Role column check/add failed (may already exist): {e}")
            
                # Check if we need to update password_hash column length
                try:
                    from sqlalchemy import text
                    # Check current password_hash column length
                    result = db.session.execute(text("SELECT character_maximum_length FROM information_schema.columns WHERE table_name='user' AND column_name='password_hash'"))
                    row = result.fetchone()
                    if row and row[0] < 255:
                        # Update password_hash column to be longer
                        db.session.execute(text("ALTER TABLE \"user\" ALTER COLUMN password_hash TYPE VARCHAR(255)"))
                        db.session.commit()
                        print("Updated password_hash column length to 255")
                except Exception as e:
                    print(f"Password hash column update failed (may already be correct): {e}")
                
                # Check if is_admin column exists, if not add it
                try:
                    from sqlalchemy import text
                    result = db.session.execute(text("SELECT column_name FROM information_schema.columns WHERE table_name='user' AND column_name='is_admin'"))
                    if not result.fetchone():
                        db.session.execute(text("ALTER TABLE \"user\" ADD COLUMN is_admin BOOLEAN DEFAULT FALSE"))
                        db.session.commit()
                        print("Added is_admin column to user table")
                except Exception as e:
                    print(f"is_admin column check/add failed (may already exist): {e}")
                
                # Check if last_login column exists, if not add it
                try:
                    from sqlalchemy import text
                    result = db.session.execute(text("SELECT column_name FROM information_schema.columns WHERE table_name='user' AND column_name='last_login'"))
                    if not result.fetchone():
                        db.session.execute(text("ALTER TABLE \"user\" ADD COLUMN last_login TIMESTAMP"))
                        db.session.commit()
                        print("Added last_login column to user table")
                except Exception as e:
                    print(f"last_login column check/add failed (may already exist): {e}")
                
                # Check if reset_token column exists, if not add it
                try:
                    from sqlalchemy import text
                    result = db.session.execute(text("SELECT column_name FROM information_schema.columns WHERE table_name='user' AND column_name='reset_token'"))
                    if not result.fetchone():
                        db.session.execute(text("ALTER TABLE \"user\" ADD COLUMN reset_token VARCHAR(100) UNIQUE"))
                        db.session.commit()
                        print("Added reset_token column to user table")
                except Exception as e:
                    print(f"reset_token column check/add failed (may already exist): {e}")
                
                # Check if reset_token_expiry column exists, if not add it
                try:
                    from sqlalchemy import text
                    result = db.session.execute(text("SELECT column_name FROM information_schema.columns WHERE table_name='user' AND column_name='reset_token_expiry'"))
                    if not result.fetchone():
                        db.session.execute(text("ALTER TABLE \"user\" ADD COLUMN reset_token_expiry TIMESTAMP"))
                        db.session.commit()
                        print("Added reset_token_expiry column to user table")
                except Exception as e:
                    print(f"reset_token_expiry column check/add failed (may already exist): {e}")

                # Auth extension columns (Google + OTP verification status)
                try:
                    from sqlalchemy import text
                    user_columns = [
                        ('phone_number', 'VARCHAR(20)'),
                        ('email_verified', 'BOOLEAN DEFAULT FALSE'),
                        ('phone_verified', 'BOOLEAN DEFAULT FALSE'),
                        ('google_id', 'VARCHAR(255)'),
                        ('auth_provider', "VARCHAR(20) DEFAULT 'local'")
                    ]
                    for col, defn in user_columns:
                        result = db.session.execute(text("SELECT column_name FROM information_schema.columns WHERE table_name='user' AND column_name=:c"), {'c': col})
                        if not result.fetchone():
                            db.session.execute(text(f'ALTER TABLE "user" ADD COLUMN {col} {defn}'))
                            db.session.commit()
                            print(f"Added {col} to user table")
                except Exception as e:
                    print(f"Auth extension columns check/add failed (may already exist): {e}")

                # Teacher quiz archive support
                try:
                    from sqlalchemy import text
                    result = db.session.execute(text("SELECT column_name FROM information_schema.columns WHERE table_name='quiz' AND column_name='is_archived'"))
                    if not result.fetchone():
                        db.session.execute(text("ALTER TABLE quiz ADD COLUMN is_archived BOOLEAN DEFAULT FALSE"))
                        db.session.commit()
                        print("Added is_archived to quiz table")
                except Exception as e:
                    print(f"Quiz archive column check/add failed (may already exist): {e}")
                
                # Check and add violation flag columns to quiz_submission table (PostgreSQL)
                try:
                    from sqlalchemy import text
                    # Check if alt_tab_flag column exists
                    result = db.session.execute(text("SELECT column_name FROM information_schema.columns WHERE table_name='quiz_submission' AND column_name='alt_tab_flag'"))
                    if not result.fetchone():
                        db.session.execute(text("ALTER TABLE quiz_submission ADD COLUMN alt_tab_flag BOOLEAN DEFAULT FALSE"))
                        db.session.commit()
                        print("Added alt_tab_flag column to quiz_submission table")
                    
                    # Check if win_shift_s_flag column exists
                    result = db.session.execute(text("SELECT column_name FROM information_schema.columns WHERE table_name='quiz_submission' AND column_name='win_shift_s_flag'"))
                    if not result.fetchone():
                        db.session.execute(text("ALTER TABLE quiz_submission ADD COLUMN win_shift_s_flag BOOLEAN DEFAULT FALSE"))
                        db.session.commit()
                        print("Added win_shift_s_flag column to quiz_submission table")
                    
                    # Check if win_prtscn_flag column exists
                    result = db.session.execute(text("SELECT column_name FROM information_schema.columns WHERE table_name='quiz_submission' AND column_name='win_prtscn_flag'"))
                    if not result.fetchone():
                        db.session.execute(text("ALTER TABLE quiz_submission ADD COLUMN win_prtscn_flag BOOLEAN DEFAULT FALSE"))
                        db.session.commit()
                        print("Added win_prtscn_flag column to quiz_submission table")
                    
                    # Check if prtscn_flag column exists
                    result = db.session.execute(text("SELECT column_name FROM information_schema.columns WHERE table_name='quiz_submission' AND column_name='prtscn_flag'"))
                    if not result.fetchone():
                        db.session.execute(text("ALTER TABLE quiz_submission ADD COLUMN prtscn_flag BOOLEAN DEFAULT FALSE"))
                        db.session.commit()
                        print("Added prtscn_flag column to quiz_submission table")
                    for col, defn in [('device_fingerprint','VARCHAR(512)'),('marked_as_cheating','BOOLEAN DEFAULT FALSE'),('proctor_notes','TEXT')]:
                        result = db.session.execute(text("SELECT column_name FROM information_schema.columns WHERE table_name='quiz_submission' AND column_name='"+col+"'"))
                        if not result.fetchone():
                            db.session.execute(text("ALTER TABLE quiz_submission ADD COLUMN "+col+" "+defn))
                            db.session.commit()
                            print("Added "+col+" to quiz_submission")
                    result = db.session.execute(text("SELECT column_name FROM information_schema.columns WHERE table_name='quiz_question' AND column_name='image_url'"))
                    if not result.fetchone():
                        db.session.execute(text("ALTER TABLE quiz_question ADD COLUMN image_url TEXT"))
                        db.session.commit()
                        print("Added image_url to quiz_question")
                    # Ensure legacy proctoring column can hold full base64 fallback data.
                    try:
                        db.session.execute(text("ALTER TABLE proctoring_snapshot ALTER COLUMN image_path TYPE TEXT"))
                        db.session.commit()
                        print("Ensured proctoring_snapshot.image_path is TEXT")
                    except Exception as e:
                        print(f"image_path type migration skipped/failed: {e}")
                except Exception as e:
                    print(f"Violation flag columns check/add failed (may already exist): {e}")
            else:
                # For SQLite, we need to manually add columns to existing tables
                try:
                    from sqlalchemy import text
                    with db.engine.begin() as conn:
                        # Check user table columns
                        user_res = conn.execute(text("PRAGMA table_info(user);"))
                        user_cols = [str(r[1]) for r in user_res]
                        if 'phone_number' not in user_cols:
                            conn.execute(text("ALTER TABLE user ADD COLUMN phone_number VARCHAR(20);"))
                            print("✅ Added phone_number to user")
                        if 'email_verified' not in user_cols:
                            conn.execute(text("ALTER TABLE user ADD COLUMN email_verified BOOLEAN DEFAULT 0;"))
                            print("✅ Added email_verified to user")
                        if 'phone_verified' not in user_cols:
                            conn.execute(text("ALTER TABLE user ADD COLUMN phone_verified BOOLEAN DEFAULT 0;"))
                            print("✅ Added phone_verified to user")
                        if 'google_id' not in user_cols:
                            conn.execute(text("ALTER TABLE user ADD COLUMN google_id VARCHAR(255);"))
                            print("✅ Added google_id to user")
                        if 'auth_provider' not in user_cols:
                            conn.execute(text("ALTER TABLE user ADD COLUMN auth_provider VARCHAR(20) DEFAULT 'local';"))
                            print("✅ Added auth_provider to user")

                        # Check quiz table columns
                        quiz_res = conn.execute(text("PRAGMA table_info(quiz);"))
                        quiz_cols = [str(r[1]) for r in quiz_res]
                        if 'is_archived' not in quiz_cols:
                            conn.execute(text("ALTER TABLE quiz ADD COLUMN is_archived BOOLEAN DEFAULT 0;"))
                            print("✅ Added is_archived to quiz")

                        # Check quiz_submission table columns
                        res = conn.execute(text("PRAGMA table_info(quiz_submission);"))
                        cols = [str(r[1]) for r in res]
                        
                        # Add violation flag columns if missing
                        if 'alt_tab_flag' not in cols:
                            conn.execute(text("ALTER TABLE quiz_submission ADD COLUMN alt_tab_flag BOOLEAN DEFAULT 0;"))
                            print("✅ Added alt_tab_flag column to quiz_submission")
                        if 'win_shift_s_flag' not in cols:
                            conn.execute(text("ALTER TABLE quiz_submission ADD COLUMN win_shift_s_flag BOOLEAN DEFAULT 0;"))
                            print("✅ Added win_shift_s_flag column to quiz_submission")
                        if 'win_prtscn_flag' not in cols:
                            conn.execute(text("ALTER TABLE quiz_submission ADD COLUMN win_prtscn_flag BOOLEAN DEFAULT 0;"))
                            print("✅ Added win_prtscn_flag column to quiz_submission")
                        if 'prtscn_flag' not in cols:
                            conn.execute(text("ALTER TABLE quiz_submission ADD COLUMN prtscn_flag BOOLEAN DEFAULT 0;"))
                            print("✅ Added prtscn_flag column to quiz_submission")
                        if 'device_fingerprint' not in cols:
                            conn.execute(text("ALTER TABLE quiz_submission ADD COLUMN device_fingerprint VARCHAR(512);"))
                            print("✅ Added device_fingerprint to quiz_submission")
                        if 'marked_as_cheating' not in cols:
                            conn.execute(text("ALTER TABLE quiz_submission ADD COLUMN marked_as_cheating BOOLEAN DEFAULT 0;"))
                            print("✅ Added marked_as_cheating to quiz_submission")
                        if 'proctor_notes' not in cols:
                            conn.execute(text("ALTER TABLE quiz_submission ADD COLUMN proctor_notes TEXT;"))
                            print("✅ Added proctor_notes to quiz_submission")
                        res2 = conn.execute(text("PRAGMA table_info(quiz_question);"))
                        qcols = [str(r[1]) for r in res2]
                        if 'image_url' not in qcols:
                            conn.execute(text("ALTER TABLE quiz_question ADD COLUMN image_url TEXT;"))
                            print("✅ Added image_url to quiz_question")
                except Exception as e:
                    print(f"SQLite migration check failed (may already exist): {e}")
                
                print("Using SQLite database - all tables created by db.create_all()")
            
            # Create login_history table if it doesn't exist
            try:
                db.create_all()
                print("Created login_history table if it didn't exist")
            except Exception as e:
                print(f"Login history table creation failed (may already exist): {e}")

            try:
                _ensure_mock_interview_session_columns(is_sqlite)
            except Exception as mi_ensure_e:
                print(f"Mock interview session column ensure: {mi_ensure_e}")
            
            print("Database tables created successfully!")
            print(f"Using database: {app.config['SQLALCHEMY_DATABASE_URI']}")
    except Exception as e:
        print(f"Database initialization error: {str(e)}")
        # Continue running the app even if database fails

# Initialize database on startup (runs migrations for both dev and production)
# This ensures new columns are added automatically
# Note: In serverless, this runs on first request, not at import time
def initialize_on_first_request():
    """Initialize database on first request (lazy initialization for serverless)"""
    try:
        init_db()
    except Exception as e:
        print(f"Warning: Database initialization failed: {e}")
        # Don't crash - continue without initialization

# Ensure migrations/db init actually runs in serverless (Flask 3 removed before_first_request)
_db_initialized = False

@app.before_request
def _ensure_db_initialized():
    global _db_initialized
    if _db_initialized:
        return
    initialize_on_first_request()
    _db_initialized = True

@app.before_request
def _track_site_activity():
    if request.endpoint == '_ensure_db_initialized':
        return
    try:
        _record_site_activity()
    except Exception as e:
        db.session.rollback()
        print(f"Site activity tracking failed: {e}")

# For Vercel/serverless: Don't initialize at import time
# Initialize will happen on first request via @app.before_first_request or similar
if os.environ.get('VERCEL') or os.environ.get('DATABASE_URL'):
    # Serverless: Initialize lazily
    pass
else:
    # Local: Initialize immediately
    try:
        init_db()
    except Exception as e:
        print(f"Warning: Database initialization failed: {e}")

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    # Enable debug mode for local development (not in production)
    debug_mode = os.environ.get('FLASK_ENV') == 'development' or os.environ.get('FLASK_DEBUG') == '1'
    if debug_mode:
        print("🔧 Running in DEVELOPMENT mode (debug=True)")
        print("⚠️  Never use debug=True in production!")
    app.run(host='0.0.0.0', port=port, debug=debug_mode)

# Deployment trigger - commit 82835ee reverted to stable state

